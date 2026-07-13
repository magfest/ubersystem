"""Shared hotel confirmation/cancellation file import.

Used by both the uber-vault hotel portal (uber.api.HotelLookup.import_confirmation_file)
and the hotel lottery admin upload. Parses a CSV or XLSX of confirmation and/or
cancellation numbers, applies them to room assignments keyed by confirmation_num,
and retains the raw uploaded file (on disk under UPLOADED_FILES_DIR) for later
debugging so odd hotel formats can be fixed after the fact.
"""
import csv
import io
import os
import uuid

from uber.config import c


def _normalize(value):
    return str(value if value is not None else '').strip().lower().replace(' ', '_')


def parse_confirmation_rows(raw, filename):
    """Parse a CSV or XLSX file into a list of normalized-key dict rows.

    Returns (rows, error). Columns are matched case-insensitively with spaces
    treated as underscores. Parse failures come back as an error string rather
    than raising.
    """
    name = (filename or '').lower()
    rows = []
    try:
        if name.endswith(('.xlsx', '.xlsm')) or raw[:2] == b'PK':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            header = None
            for excel_row in wb.active.iter_rows(values_only=True):
                if header is None:
                    header = [_normalize(cell) for cell in excel_row]
                    continue
                rows.append({header[i]: ('' if v is None else str(v))
                             for i, v in enumerate(excel_row)
                             if i < len(header) and header[i]})
        else:
            text = raw.decode('utf-8-sig', errors='replace')
            for record in csv.DictReader(io.StringIO(text)):
                rows.append({_normalize(k): ('' if v is None else str(v))
                             for k, v in record.items() if k})
    except Exception as e:
        return [], f'Could not parse file: {e}'
    return rows, None


def _store_file(session, raw, filename, content_type, hotel, source, uploaded_by):
    from uber.models.hotel import HotelImportFile

    ext = os.path.splitext(filename or '')[1][:10]
    stored_name = f"hotel_import_{uuid.uuid4().hex}{ext}"
    os.makedirs(c.UPLOADED_FILES_DIR, exist_ok=True)
    filepath = os.path.join(c.UPLOADED_FILES_DIR, stored_name)
    with open(filepath, 'wb') as f:
        f.write(raw)

    record = HotelImportFile(
        hotel_id=hotel.id if hotel else None,
        filename=filename or stored_name,
        content_type=content_type or '',
        filepath=filepath,
        size=len(raw),
        source=source,
        uploaded_by=uploaded_by or '',
    )
    session.add(record)
    session.flush()
    return record


def import_confirmation_file(session, raw, filename, hotel=None, source='',
                             uploaded_by='', content_type=''):
    """Store the raw file, then apply confirmation/cancellation numbers.

    Recognized columns (case-insensitive, spaces treated as underscores):
      - confirmation_num (required key; identifies the attendee booking)
      - hotel_confirmation_number (optional)
      - hotel_cancellation_number (optional)

    Whichever hotel-number columns are present are applied, keyed by
    confirmation_num. Unknown columns are ignored, empty cells leave existing
    values alone, and rows that don't match a booking are skipped. The file is
    retained even when parsing fails. Returns {updated, unchanged, changes,
    error}.
    """
    from uber.models import LotteryApplication
    from uber.models.hotel import RoomAssignment, HotelExportLog

    record = _store_file(session, raw, filename, content_type, hotel, source, uploaded_by)

    rows, error = parse_confirmation_rows(raw, filename)
    if error:
        record.note = error
        session.commit()
        return {'updated': 0, 'unchanged': 0, 'changes': [], 'error': error}

    # (file column, RoomAssignment attribute)
    fields = [('hotel_confirmation_number', 'hotel_confirmation_number'),
              ('hotel_cancellation_number', 'cancellation_confirmation_number')]

    updated = 0
    unchanged = 0
    changes = []
    hotels_imported = set()
    for row in rows:
        conf_num = (row.get('confirmation_num') or '').strip()
        if not conf_num:
            continue
        app = session.query(LotteryApplication).filter(
            LotteryApplication.confirmation_num == conf_num).one_or_none()
        ras = session.query(RoomAssignment).filter_by(
            lottery_application_id=app.id).all() if app else []
        if not ras:
            continue  # no matching booking; skip the row

        row_present = False
        row_changed = False
        for col, attr in fields:
            if col not in row:
                continue
            new_val = (row.get(col) or '').strip()
            if not new_val:
                continue  # empty cell: don't clear an existing value
            row_present = True
            old_val = getattr(ras[0], attr) or ''
            field_changed = False
            for ra in ras:
                if (getattr(ra, attr) or '') != new_val:
                    setattr(ra, attr, new_val)
                    session.add(ra)
                    field_changed = True
                    if ra.inventory and ra.inventory.hotel_id:
                        hotels_imported.add(str(ra.inventory.hotel_id))
            if field_changed:
                row_changed = True
                changes.append({'confirmation_num': conf_num, 'field': col,
                                'old': old_val, 'new': new_val})
        if row_present:
            updated += 1 if row_changed else 0
            unchanged += 0 if row_changed else 1

    record.updated_count = updated
    record.unchanged_count = unchanged
    record.note = f"{updated} updated, {unchanged} unchanged"

    for hotel_id in hotels_imported:
        session.add(HotelExportLog(
            hotel_id=hotel_id, export_type='confirmation_import', record_count=updated))
    session.commit()

    return {'updated': updated, 'unchanged': unchanged, 'changes': changes, 'error': None}
