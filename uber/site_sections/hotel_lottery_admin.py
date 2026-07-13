import base64
import csv
import os
import pycountry
import cherrypy
import logging
from cherrypy.lib.static import serve_file
import random
import math
from copy import deepcopy
from collections import defaultdict
from datetime import date, datetime, timedelta
from pytz import UTC
from dateutil import parser as dateparser
import sqlalchemy as sa
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import joinedload
from sqlalchemy.types import String
from ortools.linear_solver import pywraplp

from uber.config import c
from uber.custom_tags import datetime_local_filter
from uber.decorators import all_renderable, log_pageview, ajax, xlsx_file, csv_file, multifile_zipfile, render
from uber.errors import HTTPRedirect
from uber.forms import load_forms
from uber.models import Attendee, Group, LotteryApplication, Email, Tracking, PageViewTracking
from uber.lottery_perms import record_partition_audit
from uber.models.hotel import (HotelRoomInventory, InventoryNightQuantity, InventoryPartition,
                               InventoryPartitionBlock, LotteryRun, HotelExportLog, LotteryHotel, LotteryRoomType,
                               PartitionAuditLog, PartitionOwner, RoomAssignment,
                               WaitlistReveal, WaitlistRevealLink, HotelRoomIssueNote,
                               HotelImportFile)
from uber.email import EmailService
from uber.utils import (Order, check_csrf, get_page, localized_now,
                        validate_model, get_age_from_birthday,
                        normalize_email_legacy)

log = logging.getLogger(__name__)

def _search(session, text):
    applications = session.query(LotteryApplication)

    terms = text.split()
    if len(terms) == 1 and terms[0].isdigit():
        if len(terms[0]) == 10:
            return applications.filter(or_(LotteryApplication.confirmation_num == terms[0])), ''

    check_list = []

    # Skip columns that will raise unexpected applications
    skip_columns = {'id', 'parent_application_id',
                    'lottery_run_id', 'former_parent_id'}
    for attr in [col for col in LotteryApplication.__table__.columns if isinstance(col.type, String)]:
        if attr.name not in skip_columns:
            check_list.append(attr.ilike('%' + text + '%'))

    # Search by hotel / room-type name through inventory. Room assignments
    # live on RoomAssignment, so the inventory match goes through the
    # RoomAssignment join (applications that have any awarded
    # RoomAssignment at a matching inventory row).
    from uber.models.hotel import RoomAssignment

    matching_inventory_ids = set()
    hotel_matches = session.query(HotelRoomInventory.id).join(
        LotteryHotel, HotelRoomInventory.hotel_id == LotteryHotel.id
    ).filter(LotteryHotel.name.ilike('%' + text + '%')).all()
    matching_inventory_ids.update(str(row[0]) for row in hotel_matches)

    rt_matches = session.query(HotelRoomInventory.id).join(
        LotteryRoomType, or_(
            HotelRoomInventory.room_type_id == LotteryRoomType.id,
            HotelRoomInventory.suite_type_id == LotteryRoomType.id,
        )
    ).filter(LotteryRoomType.name.ilike('%' + text + '%')).all()
    matching_inventory_ids.update(str(row[0]) for row in rt_matches)

    if matching_inventory_ids:
        app_ids_with_inventory_match = session.query(
            RoomAssignment.lottery_application_id
        ).filter(
            RoomAssignment.inventory_id.in_(matching_inventory_ids),
            RoomAssignment.lottery_application_id.isnot(None),
        ).distinct().all()
        if app_ids_with_inventory_match:
            check_list.append(
                LotteryApplication.id.in_(
                    str(row[0]) for row in app_ids_with_inventory_match))

    for col_name in ['entry_type', 'status']:
        col = getattr(LotteryApplication, col_name).type
        label_list = [choice for choice in col.choices.values()]
        for label in label_list:
            if text.lower() in label.lower():
                check_list.append(getattr(LotteryApplication, col_name) == col.convert_if_label(label))

    if not check_list:
        return applications.filter(sa.false()), 'No matches found.'

    return applications.filter(or_(*check_list)), ''


def _partition_capacity(session, inv, night, partition_id):
    """Compute the effective capacity and assigned count for a block/night respecting partitions.

    If partition_id is set, the capacity is the partition's allocation for
    this block and only same-partition assignments count. If partition_id
    is None, the capacity is the block's total minus all partition
    allocations, and only non-partitioned assignments count.

    Assignment count is sourced from RoomAssignment (per multi-room).

    Returns (capacity, assigned_count, open_slots).
    """
    from uber.models.hotel import RoomAssignment

    nq_map = inv.night_quantity_map
    block_qty = nq_map.get(night, inv.quantity) if nq_map else inv.quantity

    base_filters = [
        RoomAssignment.inventory_id == str(inv.id),
        RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
        RoomAssignment.assigned_check_in_date <= night,
        RoomAssignment.assigned_check_out_date > night,
    ]

    if partition_id:
        pb = session.query(InventoryPartitionBlock).filter_by(
            partition_id=partition_id, inventory_id=inv.id).first()
        capacity = min(pb.quantity, block_qty) if pb else 0
        assigned_count = session.query(RoomAssignment).filter(
            *base_filters,
            RoomAssignment.partition_id == partition_id,
        ).count()
    else:
        total_partitioned = session.query(
            func.coalesce(func.sum(InventoryPartitionBlock.quantity), 0)
        ).filter(
            InventoryPartitionBlock.inventory_id == str(inv.id),
        ).scalar()
        capacity = max(0, block_qty - total_partitioned)
        assigned_count = session.query(RoomAssignment).filter(
            *base_filters,
            RoomAssignment.partition_id == None,  # noqa: E711
        ).count()

    return capacity, assigned_count, max(0, capacity - assigned_count)


def _fulfill_waitlist(session, inventory_id=None, night_date=None):
    """Process waitlist by extending the assigned dates on SECURED
    RoomAssignment rows that have unfulfilled per-room waitlist demand.

    Waitlist demand is the delta between the assignment's
    `waitlisted_check_in_date` / `waitlisted_check_out_date` and the
    confirmed `assigned_check_in_date` / `assigned_check_out_date`.
    Both waitlist columns NULL -> no demand -> row is skipped.

    The per-room columns are the source of truth: the application's
    `earliest_checkin_date` / `latest_checkout_date` represent the
    original lottery entry, which isn't meaningful once attendees can
    edit per-room dates post-award.

    Args:
        inventory_id: If provided, only process this inventory block.
        night_date: If provided, only process this specific night.
    """
    from uber.models.hotel import RoomAssignment

    total_fulfilled = 0
    total_skipped_locked = 0
    fulfilled_assignments = set()

    # Only SECURED, non-group, inventory-bound rows with at least one
    # waitlist column populated are candidates. (Group-entry sub-apps
    # don't get their own RoomAssignment; the leader's row covers the
    # group's nights.)
    base_q = (session.query(RoomAssignment)
              .outerjoin(LotteryApplication,
                         RoomAssignment.lottery_application_id == LotteryApplication.id)
              .filter(RoomAssignment.status == c.SECURED,
                      RoomAssignment.inventory_id.isnot(None),
                      sa.or_(LotteryApplication.id.is_(None),
                             LotteryApplication.entry_type != c.GROUP_ENTRY),
                      sa.or_(RoomAssignment.waitlisted_check_in_date.isnot(None),
                             RoomAssignment.waitlisted_check_out_date.isnot(None))))

    def _wl_ci(ra):
        return ra.waitlisted_check_in_date or ra.assigned_check_in_date

    def _wl_co(ra):
        return ra.waitlisted_check_out_date or ra.assigned_check_out_date

    if inventory_id and night_date:
        pairs_to_process = [(str(inventory_id), night_date)]
    else:
        candidates = base_q.all()
        if inventory_id:
            candidates = [ra for ra in candidates
                          if str(ra.inventory_id) == str(inventory_id)]

        pairs = set()
        for ra in candidates:
            block_id = str(ra.inventory_id)
            wl_ci = _wl_ci(ra)
            wl_co = _wl_co(ra)
            if wl_ci and ra.assigned_check_in_date and wl_ci < ra.assigned_check_in_date:
                d = wl_ci
                while d < ra.assigned_check_in_date:
                    pairs.add((block_id, d))
                    d += timedelta(days=1)
            if wl_co and ra.assigned_check_out_date and wl_co > ra.assigned_check_out_date:
                d = ra.assigned_check_out_date
                while d < wl_co:
                    pairs.add((block_id, d))
                    d += timedelta(days=1)
        pairs_to_process = sorted(pairs, key=lambda p: p[1])

    for block_id, night in pairs_to_process:
        inv = session.query(HotelRoomInventory).get(block_id)
        if not inv:
            continue

        # Discover the set of partition_ids covered by candidates for this block.
        # Exclude export-locked applications (those rows can't be edited locally
        # anymore - they live with the hotel).
        block_candidates = base_q.filter(
            RoomAssignment.inventory_id == block_id,
            sa.or_(LotteryApplication.id.is_(None),
                   LotteryApplication.export_locked == False),
        ).all()

        candidate_partitions = set()
        for ra in block_candidates:
            wl_ci = _wl_ci(ra)
            wl_co = _wl_co(ra)
            has_demand = (
                (wl_ci and ra.assigned_check_in_date and wl_ci < ra.assigned_check_in_date)
                or
                (wl_co and ra.assigned_check_out_date and wl_co > ra.assigned_check_out_date))
            if has_demand:
                candidate_partitions.add(ra.partition_id)

        skipped_locked = base_q.filter(
            RoomAssignment.inventory_id == block_id,
            LotteryApplication.export_locked == True,
        ).count()
        total_skipped_locked += skipped_locked

        for part_id in candidate_partitions:
            max_iterations = 500
            for _iteration in range(max_iterations):
                capacity, assigned_count, open_slots = _partition_capacity(
                    session, inv, night, part_id)
                if open_slots <= 0:
                    break

                part_filter = ((RoomAssignment.partition_id == part_id) if part_id
                               else (RoomAssignment.partition_id == None))  # noqa: E711
                candidates = base_q.filter(
                    RoomAssignment.inventory_id == block_id,
                    sa.or_(LotteryApplication.id.is_(None),
                           LotteryApplication.export_locked == False),
                    part_filter,
                ).all()

                eligible = []
                for ra in candidates:
                    wl_ci = _wl_ci(ra)
                    wl_co = _wl_co(ra)
                    # Walk the gap one night at a time on either end -
                    # we only extend by one contiguous night per pass.
                    if (wl_ci and ra.assigned_check_in_date
                            and night < ra.assigned_check_in_date
                            and night >= wl_ci
                            and night == ra.assigned_check_in_date - timedelta(days=1)):
                        eligible.append(('checkin', ra))
                    elif (wl_co and ra.assigned_check_out_date
                            and night >= ra.assigned_check_out_date
                            and night < wl_co
                            and night == ra.assigned_check_out_date):
                        eligible.append(('checkout', ra))

                if not eligible:
                    break

                # FIFO: earliest waitlist_started_at first. Stable on
                # `id` so concurrent same-millisecond entries (rare but
                # possible during a solver re-run) get a deterministic
                # order. NULL `waitlist_started_at` is treated as the
                # epoch - should never happen post-migration, but if a
                # row ever ends up with waitlisted_* set and no start
                # timestamp, we still want it served (FIFO can't
                # reasonably gauge "when did they join" if it's missing,
                # so we default to "as early as possible" rather than
                # silently dropping the row).
                from datetime import datetime as _dt, timezone as _tz
                _epoch = _dt(1970, 1, 1, tzinfo=_tz.utc)
                eligible.sort(key=lambda dr: (
                    dr[1].waitlist_started_at or _epoch,
                    str(dr[1].id)))
                selected = eligible[:open_slots]
                if not selected:
                    break

                for direction, ra in selected:
                    if direction == 'checkin':
                        ra.assigned_check_in_date = night
                    else:
                        ra.assigned_check_out_date = night + timedelta(days=1)
                    # The model's `clear_waitlist_when_satisfied` presave
                    # zeros the waitlist columns when the assigned range
                    # fully covers the request, so the row drops out of
                    # subsequent scans without an extra branch here.
                    session.add(ra)
                    total_fulfilled += 1
                    fulfilled_assignments.add(ra)

                session.flush()

    session.commit()

    for ra in fulfilled_assignments:
        if ra.attendee and ra.lottery_application:
            EmailService.queue_email(
                session, 'hotel_lottery_waitlist_fulfilled', ra.lottery_application,
                subject=f'{c.EVENT_NAME} Hotel Lottery - Room Dates Updated',
                data={'assignment': ra, 'app': ra.lottery_application})

    return {
        "success": True,
        "fulfilled": total_fulfilled,
        "skipped_locked": total_skipped_locked,
        "message": f"Fulfilled {total_fulfilled} waitlist entries." + (
            f" Skipped {total_skipped_locked} locked entries." if total_skipped_locked else "")
    }


def weight_entry(entry, hotel_room, base_weight):
    """Takes a lottery entry and a hotel room and returns an arbitrary score for how likely that applicant
        should be to get that particular room.
    """
    weight = 0

    # Give 10 points for being the first choice hotel, 9 points for the second, etc
    hotel_choice_rank = 10 - entry["hotels"].index(hotel_room["hotel_id"])
    weight += hotel_choice_rank

    # Give 10 points for being the first choice room type, 9 points for the second, etc
    try:
        room_type_rank = 10 - entry["room_types"].index(hotel_room["room_type"])
        assert room_type_rank >= 0
        weight += room_type_rank
    except ValueError:
        # room types are optional, so we need to figure out how much weight to give people who don't choose any
        weight += 9 # Probably fine?

    return weight + base_weight

def solve_lottery(applications, hotel_rooms, lottery_type=c.ROOM_ENTRY,
                  connector_map=None):
    """Takes a set of hotel_rooms and applications and assigns the
    hotel_rooms with mandatory connector-room coupling.

    Each inventory block can be tagged as a "connector" via
    `connector_map`: a dict mapping `child_type_id` ->
    `(parent_type_id, qty)`. Connector inventory does not participate
    in the per-app "max 1 room" cap, but each connector still respects
    its own per-inventory capacity. The solver adds an equality coupling
    constraint per app and parent inventory:

        sum(child_vars for app over child_type inventory)
            == parent_var[app, p] * qty

    so awarding the parent forces exactly `qty` connectors to the same
    app, and a parent cannot be awarded if its connectors can't be
    satisfied.

    Parameters:
        applications List[Application]: Iterable set of Application
            objects to assign.
        hotel_rooms List[dict]: Iterable set of hotel rooms; each dict
            has id, hotel_id, capacity, min_capacity, room_type,
            quantity, night_quantities.
        lottery_type: c.ROOM_ENTRY or c.SUITE_ENTRY.
        connector_map: dict {child_type_id: (parent_type_id, qty)}.
            Empty / None when no types are configured as connectors.

    Returns:
        List[Tuple[application_id, inventory_id, role]] where role is
        'primary' or 'connector'. The same application id can appear
        more than once. Returns None on solver failure.
    """
    connector_map = connector_map or {}
    connector_types = set(connector_map.keys())
    # parent_type -> list of (child_type, qty)
    parent_to_children = {}
    for child_type, (parent_type, qty) in connector_map.items():
        parent_to_children.setdefault(parent_type, []).append((child_type, qty))

    random.shuffle(applications)
    solver = pywraplp.Solver.CreateSolver("SAT")
    solver.SetSolverSpecificParametersAsString("log_search_progress: true")

    # Collect all nights across all inventory blocks.
    all_nights = set()
    inventory_by_id = {hr["id"]: hr for hr in hotel_rooms}
    inventory_by_type = {}  # type_id -> [hotel_room_dict, ...]
    for hr in hotel_rooms:
        hr["primary_constraints"] = []   # [(BoolVar, entry)] for per-app cap + per-inv cap
        hr["connector_constraints"] = []  # [(BoolVar, entry, parent_inv_id)] for per-inv cap only
        if hr.get("night_quantities"):
            all_nights.update(hr["night_quantities"].keys())
        inventory_by_type.setdefault(hr["room_type"], []).append(hr)

    # Build entries (one per non-group app), then absorb group members.
    entries = {}
    for app in applications:
        if app.entry_type == lottery_type or (
                lottery_type == c.ROOM_ENTRY
                and app.entry_type == c.SUITE_ENTRY
                and app.room_opt_out is False):
            type_pref = (app.room_type_preference if lottery_type == c.ROOM_ENTRY
                         else app.suite_type_preference)
            entries[app.id] = {
                "app": app,
                "members": [app],
                "hotels": app.hotel_preference.split(","),
                "room_types": type_pref.split(","),
                "primary_vars": [],     # [(BoolVar, weight, hotel_room)]
                "connector_vars": [],   # [(BoolVar, hotel_room, parent_inv_id, child_type, qty)]
                "check_in": app.earliest_checkin_date,
                "check_out": app.latest_checkout_date,
            }

    for app in applications:
        if app.parent_application and app.parent_application.id in entries:
            entries[app.parent_application.id]["members"].append(app)

    # Create BoolVars: primary for every eligible (app, non-connector inv),
    # plus connector vars for every (app, parent_inv, child_inv) where the
    # app might win the parent.
    for app_id, entry in entries.items():
        # Bias weights based on group size.
        base_weight = 0
        weights_cfg = c.HOTEL_LOTTERY["weights"]
        if random.random() < weights_cfg[f"group_weight_{len(entry['members'])}"]:
            base_weight = weights_cfg[f"group_base_{len(entry['members'])}"]

        for hr in hotel_rooms:
            # An app's preferences only consider primary (non-connector)
            # types. Connector rooms can never be a primary preference.
            if hr["room_type"] in connector_types:
                continue
            if hr["hotel_id"] not in entry["hotels"]:
                continue
            if hr["room_type"] not in entry["room_types"]:
                continue
            if not (hr["min_capacity"] <= len(entry["members"]) <= hr["capacity"]):
                continue

            weight = weight_entry(entry, hr, base_weight)
            primary_var = solver.BoolVar(f'{app_id}_primary_{hr["id"]}')
            entry["primary_vars"].append((primary_var, weight, hr))
            hr["primary_constraints"].append((primary_var, entry))

            # For each child type of this primary's room type, also create
            # connector BoolVars over every child-type inventory. We index
            # by parent_inventory_id so the coupling constraint can be
            # local to (app, parent_inv).
            for child_type, qty in parent_to_children.get(hr["room_type"], []):
                for child_hr in inventory_by_type.get(child_type, []):
                    cvar = solver.BoolVar(
                        f'{app_id}_connector_{child_hr["id"]}_for_{hr["id"]}')
                    entry["connector_vars"].append(
                        (cvar, child_hr, hr["id"], child_type, qty))
                    child_hr["connector_constraints"].append(
                        (cvar, entry, hr["id"]))

    # Per-inventory capacity. Each inventory's pool is the union of its
    # primary BoolVars (apps using it as their main award) plus its
    # connector BoolVars (apps using it as a ride-along for some parent).
    def _vars_for_inventory(hr):
        primary = [(cv, entry) for cv, entry in hr["primary_constraints"]]
        connector = [(cv, entry) for cv, entry, _ in hr["connector_constraints"]]
        return primary + connector

    if all_nights:
        for hr in hotel_rooms:
            inv_vars = _vars_for_inventory(hr)
            if not inv_vars:
                continue
            nq = hr.get("night_quantities", {})
            for night_iso in sorted(all_nights):
                night_qty = nq.get(night_iso, 0)
                if night_qty <= 0:
                    continue
                night_date = date.fromisoformat(night_iso)
                night_vars = [
                    cv for cv, entry in inv_vars
                    if entry["check_in"] and entry["check_out"]
                    and entry["check_in"] <= night_date < entry["check_out"]
                ]
                if night_vars:
                    solver.Add(sum(night_vars) <= night_qty)
    else:
        # Fallback when no per-night data is available.
        for hr in hotel_rooms:
            inv_vars = _vars_for_inventory(hr)
            if inv_vars:
                solver.Add(sum(cv for cv, _ in inv_vars) <= hr["quantity"])

    # Per-app "max one primary award" cap. Connector BoolVars are
    # intentionally excluded - connector rooms ride along with the
    # parent and don't count as separate awards.
    for app_id, entry in entries.items():
        if entry["primary_vars"]:
            solver.Add(sum(v for v, _, _ in entry["primary_vars"]) <= 1)

    # Connector coupling. For each (app, parent_inventory, child_type),
    # the sum of connector BoolVars over all inventory of that child
    # type for that (app, parent_inv) must equal qty if the parent is
    # awarded, and 0 otherwise. We express this as a single equality
    # constraint per (app, parent_inv, child_type).
    for app_id, entry in entries.items():
        # Bucket connector vars by (parent_inv_id, child_type, qty).
        bucket = {}  # (parent_inv_id, child_type) -> (qty, [child_vars])
        for cvar, child_hr, parent_inv_id, child_type, qty in entry["connector_vars"]:
            bucket.setdefault((parent_inv_id, child_type), (qty, []))
            bucket[(parent_inv_id, child_type)][1].append(cvar)

        for (parent_inv_id, child_type), (qty, cvars) in bucket.items():
            # Find the parent BoolVar for this app and parent inventory.
            parent_var = None
            for pvar, _w, hr in entry["primary_vars"]:
                if hr["id"] == parent_inv_id:
                    parent_var = pvar
                    break
            if parent_var is None or not cvars:
                continue
            # sum(connectors) == parent * qty
            solver.Add(sum(cvars) == parent_var * qty)

    # Objective: weighted sum on primary BoolVars only. Connectors ride
    # along and don't contribute to the maximization signal.
    objective = solver.Objective()
    for entry in entries.values():
        for pvar, weight, _hr in entry["primary_vars"]:
            objective.SetCoefficient(pvar, weight)
    objective.SetMaximization()

    status = solver.Solve()
    if status not in (pywraplp.Solver.OPTIMAL, pywraplp.Solver.FEASIBLE):
        log.error(f"Error solving room lottery: {status}")
        return None

    # Output: list of (leader_application_id, inventory_id, role). Group
    # members do NOT appear separately - they get added as occupants on
    # each leader-owned RoomAssignment during materialization.
    allocations = []
    for app_id, entry in entries.items():
        leader_id = entry["app"].id
        for pvar, _weight, hr in entry["primary_vars"]:
            if pvar.solution_value() > 0.5:
                allocations.append((leader_id, hr["id"], 'primary'))

        for cvar, child_hr, _pid, _ct, _qty in entry["connector_vars"]:
            if cvar.solution_value() > 0.5:
                allocations.append((leader_id, child_hr["id"], 'connector'))

    return allocations

def _notify_applicants_of_inventory_change(session, inventory):
    """When an inventory row is deactivated, email applicants whose
    preferences referenced it. Matches by UUID inside the comma-separated
    preference strings on LotteryApplication.
    """
    if not inventory:
        return
    inv_id = str(inventory.id)
    hotel_id = str(inventory.hotel_id) if inventory.hotel_id else None
    type_id = (str(inventory.suite_type_id) if inventory.is_suite
               else str(inventory.room_type_id))

    candidates = session.query(LotteryApplication).filter(
        LotteryApplication.status.in_([c.COMPLETE, c.PROCESSED]),
    ).all()

    for app in candidates:
        hotels = {x.strip() for x in (app.hotel_preference or '').split(',') if x.strip()}
        rooms = {x.strip() for x in (app.room_type_preference or '').split(',') if x.strip()}
        suites = {x.strip() for x in (app.suite_type_preference or '').split(',') if x.strip()}
        if hotel_id and hotel_id not in hotels:
            continue
        if type_id and type_id not in (suites if inventory.is_suite else rooms):
            continue
        if not app.attendee:
            continue
        EmailService.queue_email(
            session, 'hotel_lottery_inventory_changed_applicant', app,
            subject=f"{c.EVENT_NAME_AND_YEAR}: One of your lottery preferences is no longer available",
            data={
            'attendee': app.attendee, 'application': app, 'inventory': inventory,
        })


def _notify_partition_owners_of_inventory_change(session, partition, change_description):
    """Notify partition owners with can_view_inventory of edits to
    blocks in their partition."""
    if not partition:
        return
    grants = session.query(PartitionOwner).filter_by(
        partition_id=partition.id, can_view_inventory=True).all()
    for grant in grants:
        if not grant.admin_account or not grant.admin_account.attendee:
            continue
        recipient = grant.admin_account.attendee
        EmailService.queue_email(
            session, 'hotel_lottery_inventory_changed_owner', partition,
            subject=f"{c.EVENT_NAME_AND_YEAR}: Inventory change in {partition.name}",
            data={
            'attendee': recipient, 'partition': partition,
            'change_description': change_description,
        })


def _send_confirmation_updated_email(session, assignment):
    """Notify the attendee that hotel_confirmation_number has changed.

    Direct send rather than AutomatedEmailFixture so it fires exactly once
    at the change site (the import endpoint or the API). Called after the
    field is set on `assignment` but before commit.
    """
    if not assignment or not assignment.attendee_id:
        return
    attendee = session.query(Attendee).get(assignment.attendee_id)
    if not attendee:
        return
    EmailService.queue_email(
        session, 'hotel_lottery_confirmation_updated', assignment,
        subject=f"{c.EVENT_NAME_AND_YEAR}: Hotel confirmation number updated",
        data={
        'attendee': attendee, 'assignment': assignment,
    })


def _room_issues_url(message='', severity='all', kind='all', search='',
                     show_hidden=''):
    """Build a room_issues URL preserving the active filters, so the
    hide/unhide/note POST handlers redirect back to the same view. Built
    with urlencode and passed to HTTPRedirect as one pre-formatted
    string (HTTPRedirect quotes each `{}` substitution, which would
    double-encode a hand-built query string)."""
    from urllib.parse import urlencode
    params = {}
    if severity and severity != 'all':
        params['severity'] = severity
    if kind and kind not in ('all', ''):
        params['kind'] = kind
    if search:
        params['search'] = search
    if show_hidden:
        params['show_hidden'] = '1'
    if message:
        params['message'] = message
    qs = urlencode(params)
    return 'room_issues' + ('?' + qs if qs else '')


@all_renderable()
class Root:
    def _materialize_room_assignments(self, session, applications, allocations,
                                       lottery_run, run_deadline, partition_filter):
        """Create RoomAssignment rows for the solver output.

        `allocations` is the list of (leader_application_id, inventory_id,
        role) tuples produced by solve_lottery. Group members are added as
        occupants on each leader-owned RoomAssignment.

        Connector rows are created with parent_assignment_id pointing at
        the primary RoomAssignment for the same (app, parent_inventory).
        Each connector lives as its own row so the hotel export emits one
        line per physical room.

        Returns the number of distinct leaders that got at least one
        primary award (so the caller can stamp lottery_run.rooms_assigned).
        """
        from uber.models import RoomAssignment

        app_by_id = {a.id: a for a in applications}

        # Group allocations by leader app + role.
        by_leader = {}  # leader_id -> {'primary': [inv_id...], 'connector': [inv_id...]}
        for leader_id, inv_id, role in allocations:
            by_leader.setdefault(leader_id, {'primary': [], 'connector': []})[role].append(inv_id)

        leaders_with_primary = 0
        for leader_id, roles in by_leader.items():
            leader = app_by_id.get(leader_id)
            if not leader or not leader.attendee_id:
                continue

            # Default occupants: the leader's attendee + every valid group
            # member's attendee. Leaders edit per-room later.
            occupant_ids = [leader.attendee_id]
            for member in leader.valid_group_members or []:
                if member.attendee_id and member.attendee_id not in occupant_ids:
                    occupant_ids.append(member.attendee_id)

            # Primary first so we can hang connectors off its id.
            primaries_by_inv = {}
            for inv_id in roles['primary']:
                primary = RoomAssignment(
                    attendee_id=leader.attendee_id,
                    inventory_id=inv_id,
                    lottery_application_id=leader.id,
                    lottery_run_id=lottery_run.id,
                    partition_id=partition_filter or None,
                    assignment_reason=c.LOTTERY_AWARD,
                    status=c.ASSIGNED,
                    require_cc=True,
                    assigned_check_in_date=leader.earliest_checkin_date,
                    assigned_check_out_date=leader.latest_checkout_date,
                    deposit_cutoff_date=run_deadline,
                )
                session.add(primary)
                session.flush()  # need primary.id
                primaries_by_inv[inv_id] = primary
                self._set_occupants(session, primary, occupant_ids)
                leaders_with_primary += 1

            # Connectors - for each, find the parent primary in this leader's
            # set (any primary will do as the structural parent; the solver's
            # coupling already guaranteed there's one).
            parent_primary = next(iter(primaries_by_inv.values()), None)
            for inv_id in roles['connector']:
                child = RoomAssignment(
                    attendee_id=leader.attendee_id,
                    inventory_id=inv_id,
                    lottery_application_id=leader.id,
                    lottery_run_id=lottery_run.id,
                    parent_assignment_id=parent_primary.id if parent_primary else None,
                    partition_id=partition_filter or None,
                    assignment_reason=c.SUITE_CONNECTOR,
                    status=c.ASSIGNED,
                    require_cc=True,
                    assigned_check_in_date=leader.earliest_checkin_date,
                    assigned_check_out_date=leader.latest_checkout_date,
                    deposit_cutoff_date=run_deadline,
                )
                session.add(child)
                session.flush()
                self._set_occupants(session, child, occupant_ids)

        return leaders_with_primary

    def _set_occupants(self, session, assignment, attendee_ids):
        """Replace the room_assignment_occupant rows for `assignment`."""
        from uber.models.hotel import room_assignment_occupant
        session.execute(room_assignment_occupant.delete().where(
            room_assignment_occupant.c.room_assignment_id == assignment.id))
        for aid in attendee_ids:
            session.execute(room_assignment_occupant.insert().values(
                room_assignment_id=assignment.id, attendee_id=aid))

    def index(self, session, message='', page='0', search_text='', order='status', **params):
        if c.DEV_BOX and not int(page):
            page = 1

        total_count = session.query(LotteryApplication.id).count()
        complete_valid_entries = session.query(LotteryApplication.id).filter(LotteryApplication.status == c.COMPLETE).join(
            LotteryApplication.attendee).filter(Attendee.hotel_lottery_eligible == True)
        room_count_base = complete_valid_entries.filter(LotteryApplication.entry_type != c.GROUP_ENTRY)
        count = 0
        search_text = search_text.strip()
        advanced_filters = {}

        if search_text:
            search_results, message = _search(session, search_text)
            if search_results and search_results.count():
                applications = search_results
                count = applications.count()
                if count == total_count:
                    message = 'Every lottery application matched this search.'
            elif not message:
                message = 'No matches found. Try searching the lottery tracking history instead.'

        filter_status = params.get('filter_status', '')
        filter_entry_type = params.get('filter_entry_type', '')
        filter_hotel = params.get('filter_hotel', '')
        filter_inventory = params.get('filter_inventory', '')
        filter_partition = params.get('filter_partition', '')
        filter_export_locked = params.get('filter_export_locked', '')
        filter_staff = params.get('filter_staff', '')

        has_advanced = any([filter_status, filter_entry_type, filter_hotel,
                           filter_inventory, filter_partition, filter_export_locked, filter_staff])

        if has_advanced:
            if not count:
                applications = session.query(LotteryApplication)
            if filter_status:
                applications = applications.filter(LotteryApplication.status == int(filter_status))
            if filter_entry_type:
                applications = applications.filter(LotteryApplication.entry_type == int(filter_entry_type))
            if filter_hotel:
                inv_ids = [str(inv.id) for inv in
                           session.query(HotelRoomInventory).filter_by(hotel_id=filter_hotel).all()]
                if inv_ids:
                    matched_app_ids = [
                        row[0] for row in session.query(
                            RoomAssignment.lottery_application_id
                        ).filter(
                            RoomAssignment.inventory_id.in_(inv_ids),
                            RoomAssignment.lottery_application_id.isnot(None),
                        ).distinct().all()
                    ]
                    if matched_app_ids:
                        applications = applications.filter(
                            LotteryApplication.id.in_(matched_app_ids))
                    else:
                        applications = applications.filter(sa.false())
                else:
                    applications = applications.filter(sa.false())
            if filter_inventory:
                matched_app_ids = [
                    row[0] for row in session.query(
                        RoomAssignment.lottery_application_id
                    ).filter(
                        RoomAssignment.inventory_id == filter_inventory,
                        RoomAssignment.lottery_application_id.isnot(None),
                    ).distinct().all()
                ]
                if matched_app_ids:
                    applications = applications.filter(
                        LotteryApplication.id.in_(matched_app_ids))
                else:
                    applications = applications.filter(sa.false())
            if filter_partition:
                applications = applications.filter(LotteryApplication.partition_id == filter_partition)
            if filter_export_locked == 'true':
                applications = applications.filter(LotteryApplication.export_locked == True)
            elif filter_export_locked == 'false':
                applications = applications.filter(LotteryApplication.export_locked == False)
            if filter_staff == 'true':
                applications = applications.filter(LotteryApplication.is_staff_entry == True)
            elif filter_staff == 'false':
                applications = applications.filter(LotteryApplication.is_staff_entry == False)
            count = applications.count()
            advanced_filters = {k: v for k, v in params.items() if k.startswith('filter_') and v}

        if not count:
            applications = session.query(LotteryApplication)
            count = applications.count()

        applications = applications.order(order).options(joinedload(LotteryApplication.attendee))

        page = int(page)
        if search_text:
            page = page or 1

        pages = range(1, int(math.ceil(count / 100)) + 1)
        applications = applications[-100 + 100*page: 100*page] if page else []

        return {
            'message':        message if isinstance(message, str) else message[-1],
            'page':           page,
            'pages':          pages,
            'search_text':    search_text,
            'search_results': bool(search_text) or has_advanced,
            'applications':   applications,
            'order':          Order(order),
            'search_count':   count,
            'total_count':    total_count,
            'complete_count': complete_valid_entries.count(),
            'suite_count': room_count_base.filter(LotteryApplication.entry_type == c.SUITE_ENTRY).count(),
            'room_count': room_count_base.filter(or_(LotteryApplication.entry_type == c.ROOM_ENTRY,
                                                     LotteryApplication.room_opt_out == False)).count(),
            'advanced_filters': advanced_filters,
            'hotels': session.query(LotteryHotel).filter_by(active=True).order_by(LotteryHotel.name).all(),
            'inventory_blocks': session.query(HotelRoomInventory).filter_by(active=True).all(),
            'partitions': session.query(InventoryPartition).filter_by(active=True).order_by(InventoryPartition.name).all(),
        }  # noqa: E711

    def feed(self, session, message='', page='1', who='', what='', action=''):
        feed = session.query(Tracking).filter(Tracking.model == 'LotteryApplication').order_by(Tracking.when.desc())
        what = what.strip()
        if who:
            feed = feed.filter_by(who=who)
        if what:
            like = '%' + what + '%'
            or_filters = [Tracking.page.ilike(like),
                          Tracking.which.ilike(like),
                          Tracking.data.ilike(like)]
            feed = feed.filter(or_(*or_filters))
        if action:
            feed = feed.filter_by(action=action)
        return {
            'message': message,
            'who': who,
            'what': what,
            'page': page,
            'action': action,
            'count': feed.count(),
            'feed': get_page(page, feed),
            'action_opts': c.TRACKING_OPTS,
            'who_opts': [
                who for [who] in session.query(Tracking).filter(
                    Tracking.model == 'LotteryApplication').distinct().order_by(Tracking.who).values(Tracking.who)]
        }
    
    @ajax
    def validate_hotel_lottery(self, session, id=None, form_list=[], **params):
        application = session.lottery_application(id)

        if not form_list:
            form_list = ["LotteryAdminInfo"]
        elif isinstance(form_list, str):
            form_list = [form_list]
        forms = load_forms(params, application, form_list)
        all_errors = validate_model(session, forms, application, is_admin=True)
        if all_errors:
            return {"error": all_errors}

        return {"success": True}

    @log_pageview
    def form(self, session, message='', return_to='', **params):
        id = params.get('id', None)

        if id in [None, '', 'None']:
            application = LotteryApplication()
        else:
            application = session.lottery_application(id)

        forms = load_forms(params, application, ['LotteryAdminInfo'])

        if cherrypy.request.method == 'POST':
            for form in forms.values():
                form.populate_obj(application, is_admin=True)
            # hotel_confirmation_number is per-RoomAssignment and edited in
            # the form's "Rooms" section, not on the application here.

            message = '{}\'s entry (conf # {}) has been saved.'.format(application.attendee_name,
                                                                       application.confirmation_num)
            stay_on_form = params.get('save_return_to_search', False) is False
            session.add(application)
            if application.orig_value_of('status') != application.status and application.status in [
                    c.REJECTED, c.CANCELLED, c.REMOVED, c.WITHDRAWN]:
                application.attendee.hotel_eligible = True
                session.add(application.attendee)
            session.commit()
            if stay_on_form:
                    raise HTTPRedirect('form?id={}&message={}&return_to={}', application.id, message, return_to)
            else:
                if return_to:
                    raise HTTPRedirect(return_to + '&message={}', 'Application updated.')
                else:
                    raise HTTPRedirect('index?message={}', message)

        # Partition + inventory picker data for the Rooms section's
        # add/edit modals. The template renders both selects with the
        # partition on top; JS filters the inventory options to those in
        # the selected partition (or all unpartitioned + every block when
        # "no partition" is chosen).
        partitions = session.query(InventoryPartition).filter_by(
            active=True).order_by(InventoryPartition.name).all()
        inventory_blocks = (session.query(HotelRoomInventory)
                            .filter_by(active=True)
                            .order_by(HotelRoomInventory.hotel_id,
                                      HotelRoomInventory.name).all())
        # {inventory_id: [partition_id, ...]} - drives the JS filter.
        # An inventory with no entry in this dict has no partition
        # restriction and is always offered.
        partition_blocks = session.query(InventoryPartitionBlock).all()
        inventory_partitions_map = {}
        for pb in partition_blocks:
            inventory_partitions_map.setdefault(
                str(pb.inventory_id), []).append(str(pb.partition_id))

        return {
            'message':    message,
            'application':   application,
            'forms': forms,
            'return_to':  return_to,
            'partitions': partitions,
            'inventory_blocks': inventory_blocks,
            'inventory_partitions_map': inventory_partitions_map,
        }

    def history(self, session, id):
        application = session.lottery_application(id)
        return {
            'application':  application,
            'emails': session.query(Email).filter(Email.model == 'LotteryApplication',
                                                  Email.fk_id == id
                                                  ).order_by(Email.when).all(),
            'changes': session.query(Tracking).filter(Tracking.model == 'LotteryApplication', Tracking.fk_id == id
                                                      ).order_by(Tracking.when).all(),
            'pageviews': session.query(PageViewTracking).filter(PageViewTracking.which == repr(application)
                                                                ).order_by(PageViewTracking.when).all(),
        }
    
    def lottery_runs(self, session, message=''):
        runs = session.query(LotteryRun).order_by(LotteryRun.run_at.desc()).all()
        hotels = session.query(LotteryHotel).filter_by(active=True).order_by(LotteryHotel.name).all()
        room_types = session.query(LotteryRoomType).filter_by(active=True, is_suite=False).order_by(LotteryRoomType.name).all()
        suite_types = session.query(LotteryRoomType).filter_by(active=True, is_suite=True).order_by(LotteryRoomType.name).all()
        inventory_blocks = session.query(HotelRoomInventory).filter_by(active=True).order_by(
            HotelRoomInventory.hotel_id, HotelRoomInventory.name).all()
        partitions = session.query(InventoryPartition).filter_by(active=True).order_by(InventoryPartition.name).all()
        return {
            'runs': runs,
            'hotels': hotels,
            'room_types': room_types,
            'suite_types': suite_types,
            'inventory_blocks': inventory_blocks,
            'partitions': partitions,
            'message': message,
        }

    def lottery_run_detail(self, session, id, message=''):
        lottery_run = session.query(LotteryRun).get(id)
        applications = session.query(LotteryApplication).filter(
            LotteryApplication.lottery_run_id == id,
            LotteryApplication.entry_type != c.GROUP_ENTRY,
        ).order_by(LotteryApplication.confirmation_num).all()
        partitions = session.query(InventoryPartition).filter_by(active=True).order_by(InventoryPartition.name).all()
        partition_lookup = {str(p.id): p.name for p in partitions}
        return {
            'lottery_run': lottery_run,
            'applications': applications,
            'hotels': session.query(LotteryHotel).filter_by(active=True).order_by(LotteryHotel.name).all(),
            'room_types': session.query(LotteryRoomType).filter_by(is_suite=False, active=True).order_by(LotteryRoomType.name).all(),
            'suite_types': session.query(LotteryRoomType).filter_by(is_suite=True, active=True).order_by(LotteryRoomType.name).all(),
            'partition_lookup': partition_lookup,
            'message': message,
        }

    def update_lottery_run(self, session, id, name, **params):
        lottery_run = session.query(LotteryRun).get(id)
        lottery_run.name = name
        session.commit()
        raise HTTPRedirect('lottery_run_detail?id={}&message={}', id, 'Run name updated.')

    def update_run_card_deadline(self, session, id, card_deadline='',
                                 propagate='', csrf_token=None):
        """Edit LotteryRun.card_deadline. Optionally retroactively apply
        the new deadline to RoomAssignments produced by this run that
        haven't been individually overridden.
        """
        from uber.utils import check_csrf
        from dateutil import parser as dateparser
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('lottery_run_detail?id={}', id)
        check_csrf(csrf_token)

        lottery_run = session.query(LotteryRun).get(id)
        if not lottery_run:
            raise HTTPRedirect('lottery_runs?message={}', 'Run not found.')

        new_deadline = None
        if card_deadline.strip():
            try:
                new_deadline = dateparser.parse(card_deadline).replace(tzinfo=c.EVENT_TIMEZONE)
            except (ValueError, TypeError):
                raise HTTPRedirect(
                    'lottery_run_detail?id={}&message={}',
                    id, 'Could not parse the deadline.')

        original = lottery_run.card_deadline
        lottery_run.card_deadline = new_deadline
        session.add(lottery_run)

        propagated_count = 0
        if propagate == '1' and new_deadline:
            from uber.models import RoomAssignment
            target_date = new_deadline.date()
            assignments = session.query(RoomAssignment).filter_by(
                lottery_run_id=lottery_run.id).all()
            original_date = original.date() if original else None
            for ra in assignments:
                # Only push the new deadline onto rows that match the prior
                # run-level deadline - leaving per-assignment overrides
                # alone, per the plan.
                if (ra.deposit_cutoff_date is None
                        or ra.deposit_cutoff_date == original_date):
                    ra.deposit_cutoff_date = target_date
                    session.add(ra)
                    propagated_count += 1

        session.commit()
        msg = "Card deadline updated."
        if propagated_count:
            msg += f" Pushed to {propagated_count} assignment(s)."
        raise HTTPRedirect('lottery_run_detail?id={}&message={}', id, msg)

    def update_assignment_deadline(self, session, id, deposit_cutoff_date='',
                                   csrf_token=None):
        """Override deposit_cutoff_date on a single RoomAssignment.

        Empty value clears the override, letting the run-level deadline
        govern again.
        """
        from uber.utils import check_csrf
        from uber.models import RoomAssignment
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('index')
        check_csrf(csrf_token)

        assignment = session.query(RoomAssignment).get(id)
        if not assignment:
            raise HTTPRedirect('index?message={}', 'Assignment not found.')

        if deposit_cutoff_date.strip():
            try:
                assignment.deposit_cutoff_date = date.fromisoformat(
                    deposit_cutoff_date.strip())
            except ValueError:
                raise HTTPRedirect('assign_room?id={}&message={}', id,
                                   'Could not parse the deadline.')
        else:
            assignment.deposit_cutoff_date = None
        session.add(assignment)
        session.commit()
        raise HTTPRedirect('assign_room?id={}&message={}', id,
                           'Deadline updated.')

    def award_run(self, session, id, **params):
        from uber.models import RoomAssignment

        lottery_run = session.query(LotteryRun).get(id)
        if lottery_run.status != c.LOTTERY_PENDING:
            raise HTTPRedirect('lottery_run_detail?id={}&message={}', id, 'This run cannot be awarded.')

        applications = session.query(LotteryApplication).join(LotteryApplication.attendee).filter(
            LotteryApplication.lottery_run_id == id,
            LotteryApplication.status == c.PROCESSED,
            Attendee.hotel_lottery_eligible == True,
        ).all()

        # Compute the per-run deposit_cutoff_date once and stamp it on
        # every RoomAssignment in this run that doesn't already carry an
        # override. The application's own status flips to AWARDED.
        run_deadline_date = None
        if c.HOTEL_LOTTERY_GUARANTEE_HOURS:
            dt = (localized_now() + timedelta(hours=c.HOTEL_LOTTERY_GUARANTEE_HOURS)).strftime('%Y-%m-%d')
            run_deadline_date = datetime.strptime(dt + ' 23:59', '%Y-%m-%d %H:%M').date()

        for app in applications:
            app.status = c.AWARDED
            session.add(app)
            if run_deadline_date:
                for ra in session.query(RoomAssignment).filter_by(
                        lottery_application_id=app.id,
                        lottery_run_id=lottery_run.id).all():
                    if not ra.deposit_cutoff_date:
                        ra.deposit_cutoff_date = run_deadline_date
                        session.add(ra)

        lottery_run.status = c.LOTTERY_AWARDED
        lottery_run.awarded_at = datetime.now(UTC)
        session.commit()
        raise HTTPRedirect('lottery_run_detail?id={}&message={}', id,
                           f"{len(applications)} entries awarded.")

    def revert_run(self, session, id, **params):
        from uber.models import RoomAssignment

        lottery_run = session.query(LotteryRun).get(id)
        if lottery_run.status != c.LOTTERY_PENDING:
            raise HTTPRedirect('lottery_run_detail?id={}&message={}', id, 'This run cannot be reverted.')

        applications = session.query(LotteryApplication).filter(
            LotteryApplication.lottery_run_id == id,
            LotteryApplication.status == c.PROCESSED,
        ).all()

        for app in applications:
            app.status = c.COMPLETE
            app.partition_id = None
            app.lottery_run_id = None
            session.add(app)

        # Drop the RoomAssignment rows materialized by this run - connectors
        # included. parent_assignment_id rows resolve through CASCADE on the
        # FK once we delete the parent, but we delete connectors explicitly
        # first for clarity.
        deleted_assignments = session.query(RoomAssignment).filter_by(
            lottery_run_id=lottery_run.id).count()
        session.query(RoomAssignment).filter_by(
            lottery_run_id=lottery_run.id).delete(synchronize_session=False)

        lottery_run.status = c.LOTTERY_REVERTED
        lottery_run.reverted_at = datetime.now(UTC)
        session.commit()
        raise HTTPRedirect('lottery_runs?message={}',
                           f"Run '{lottery_run.name}' reverted. {len(applications)} entries reset to complete, "
                           f"{deleted_assignments} room assignment(s) cleared.")

    def delete_run(self, session, id, **params):
        from uber.models import RoomAssignment

        lottery_run = session.query(LotteryRun).get(id)
        if lottery_run.status != c.LOTTERY_REVERTED:
            raise HTTPRedirect('lottery_run_detail?id={}&message={}', id, 'Only reverted runs can be deleted.')

        # Defensive: clear any stray RoomAssignments still pointing at this run.
        session.query(RoomAssignment).filter_by(
            lottery_run_id=lottery_run.id).delete(synchronize_session=False)

        name = lottery_run.name
        session.delete(lottery_run)
        session.commit()
        raise HTTPRedirect('lottery_runs?message={}', f"Run '{name}' has been deleted.")

    def manage_inventory(self, session, message=''):
        inventory = session.query(HotelRoomInventory).order_by(
            HotelRoomInventory.hotel_id, HotelRoomInventory.is_suite, HotelRoomInventory.name).all()

        # Count assigned per block
        assigned_counts = session.query(
            RoomAssignment.inventory_id, func.count(RoomAssignment.id)
        ).filter(
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            RoomAssignment.inventory_id.isnot(None),
        ).group_by(RoomAssignment.inventory_id).all()
        assigned_per_block = defaultdict(int, {str(inv_id): cnt for inv_id, cnt in assigned_counts})

        return {
            'inventory': inventory,
            'assigned_per_block': assigned_per_block,
            'hotels': session.query(LotteryHotel).filter_by(active=True).order_by(LotteryHotel.name).all(),
            'room_types': session.query(LotteryRoomType).filter_by(is_suite=False, active=True).order_by(LotteryRoomType.name).all(),
            'suite_types': session.query(LotteryRoomType).filter_by(is_suite=True, active=True).order_by(LotteryRoomType.name).all(),
            'message': message,
        }

    def edit_inventory_item(self, session, id=None, message='', **params):
        if id and id != 'None' and id != '':
            item = session.query(HotelRoomInventory).get(id)
        else:
            item = None

        if not item:
            item = HotelRoomInventory()

        # Build hotel night dates for per-night quantity grid
        event_nights = []
        day = c.HOTEL_LOTTERY_CHECKIN_START.date()
        end = c.HOTEL_LOTTERY_CHECKOUT_END.date()
        while day < end:
            event_nights.append(day)
            day += timedelta(days=1)

        if cherrypy.request.method == 'POST':
            was_active = bool(item.id) and item.active
            item.hotel_id = params['hotel']
            item.is_suite = params.get('is_suite') == 'true'
            if item.is_suite:
                item.suite_type_id = params.get('suite_type') or None
                item.room_type_id = None
            else:
                item.room_type_id = params.get('room_type') or None
                item.suite_type_id = None
            try:
                item.quantity = int(params.get('quantity', 0))
                item.capacity = int(params.get('capacity', 2))
                item.min_capacity = int(params.get('min_capacity', 1))
            except (ValueError, TypeError):
                raise HTTPRedirect('edit_inventory_item?id={}&message={}', item.id if item.id else '',
                                   'Quantity, capacity, and min capacity must be numbers.')
            item.name = params.get('name', '')
            item.active = params.get('active') == 'true'
            became_inactive = was_active and not item.active
            item.vault_reference = params.get('vault_reference', '') or None
            item.info_url = params.get('info_url', '').strip()
            item.price = params.get('price', '').strip()
            item.staff_price = params.get('staff_price', '').strip()
            session.add(item)
            session.flush()

            # Save per-night quantities
            existing_nq = {nq.night_date: nq for nq in item.night_quantities}
            for night in event_nights:
                qty_str = params.get(f'night_qty_{night.isoformat()}', '')
                if qty_str != '':
                    try:
                        qty = int(qty_str)
                    except (ValueError, TypeError):
                        continue
                    if night in existing_nq:
                        existing_nq[night].quantity = qty
                    else:
                        nq = InventoryNightQuantity(inventory_id=item.id, night_date=night, quantity=qty)
                        session.add(nq)

            session.commit()

            # Notify applicants whose preferences referenced this block if it
            # was just deactivated. Async + best-effort - failures don't roll
            # back the inventory change.
            if became_inactive:
                _notify_applicants_of_inventory_change(session, item)

            # Auto-process waitlist for this inventory block
            waitlist_result = _fulfill_waitlist(session, inventory_id=str(item.id))
            save_msg = 'Inventory item saved.'
            if waitlist_result['fulfilled'] > 0:
                save_msg += f" Waitlist: {waitlist_result['fulfilled']} entries fulfilled."

            raise HTTPRedirect('manage_inventory?message={}', save_msg)

        return {
            'item': item,
            'hotels': session.query(LotteryHotel).filter_by(active=True).all(),
            'room_types': session.query(LotteryRoomType).filter_by(is_suite=False, active=True).all(),
            'suite_types': session.query(LotteryRoomType).filter_by(is_suite=True, active=True).all(),
            'event_nights': event_nights,
            'message': message,
        }

    def settings(self, session, message=''):
        return {
            'message': message,
        }

    @ajax
    def vault_usage(self, session, month=''):
        if not c.VAULT_ENABLED:
            return {'error': 'Vault integration is not enabled.'}

        from uber.vault import get_usage, get_billing

        try:
            usage = get_usage(month=month if month else None)
            billing = get_billing()
            return {'success': True, 'usage': usage, 'billing': billing}
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def manage_hotels(self, session, message=''):
        hotels = session.query(LotteryHotel).order_by(LotteryHotel.name).all()
        return {
            'hotels': hotels,
            'message': message,
        }

    def edit_hotel(self, session, id=None, message='', **params):
        if id and id not in ('None', ''):
            hotel = session.query(LotteryHotel).get(id)
        else:
            hotel = None

        if not hotel:
            hotel = LotteryHotel()

        if cherrypy.request.method == 'POST':
            hotel.name = params.get('name', '').strip()
            hotel.export_name = params.get('export_name', '').strip()
            hotel.description = params.get('description', '').strip()
            hotel.description_right = params.get('description_right', '').strip()
            hotel.footnote = params.get('footnote', '').strip()
            hotel.active = params.get('active') == 'true'
            session.add(hotel)
            session.commit()
            raise HTTPRedirect('manage_hotels?message={}', f"Hotel '{hotel.name}' saved.")

        return {
            'hotel': hotel,
            'message': message,
        }

    def manage_room_types(self, session, message=''):
        room_types = session.query(LotteryRoomType).order_by(
            LotteryRoomType.is_suite, LotteryRoomType.name).all()
        # Pre-compute parent/children maps the template uses for the chain
        # column - saves N+1 lookups when rendering. A parent's children
        # are all rows whose `connects_to_type_id` points back at it.
        by_id = {rt.id: rt for rt in room_types}
        children_by_parent = {rt.id: [] for rt in room_types}
        for rt in room_types:
            if rt.connects_to_type_id and rt.connects_to_type_id in children_by_parent:
                children_by_parent[rt.connects_to_type_id].append(rt)
        return {
            'room_types': room_types,
            'by_id': by_id,
            'children_by_parent': children_by_parent,
            'message': message,
        }

    def edit_room_type(self, session, id=None, message='', **params):
        if id and id not in ('None', ''):
            room_type = session.query(LotteryRoomType).get(id)
        else:
            room_type = None

        if not room_type:
            room_type = LotteryRoomType()

        # All other active types - drives the "Follows another room type"
        # select. Sorted by suite-first then name so visually grouped.
        siblings = (session.query(LotteryRoomType)
                    .filter(LotteryRoomType.id != room_type.id)
                    .order_by(LotteryRoomType.is_suite.desc(), LotteryRoomType.name)
                    .all()) if room_type.id else (
                        session.query(LotteryRoomType)
                        .order_by(LotteryRoomType.is_suite.desc(), LotteryRoomType.name)
                        .all())
        # Types that follow *this* one - when non-empty, this type is a
        # parent and cannot itself be a child (no chaining), so the
        # template renders the connector controls read-only.
        children = (session.query(LotteryRoomType)
                    .filter(LotteryRoomType.connects_to_type_id == room_type.id)
                    .order_by(LotteryRoomType.name).all()) if room_type.id else []

        if cherrypy.request.method == 'POST':
            room_type.name = params.get('name', '').strip()
            room_type.export_name = params.get('export_name', '').strip()
            room_type.description = params.get('description', '').strip()
            room_type.description_right = params.get('description_right', '').strip()
            room_type.footnote = params.get('footnote', '').strip()
            room_type.capacity = int(params.get('capacity', 4))
            room_type.min_capacity = int(params.get('min_capacity', 1))
            room_type.is_suite = params.get('is_suite') == 'true'
            room_type.active = params.get('active') == 'true'

            # Connector ("follows") config. The cycle/chain guard runs
            # both here (clear feedback) and in the model_checks.py
            # validator (catches API/back-door writes).
            raw_parent = (params.get('connects_to_type_id') or '').strip()
            raw_qty = (params.get('connector_quantity') or '').strip()
            if raw_parent:
                if children:
                    raise HTTPRedirect(
                        'edit_room_type?id={}&message={}', room_type.id,
                        "Cannot make this room type follow another while other "
                        "room types follow it. Detach the children first.")
                if raw_parent == room_type.id:
                    raise HTTPRedirect(
                        'edit_room_type?id={}&message={}', room_type.id,
                        "A room type cannot follow itself.")
                parent = session.query(LotteryRoomType).get(raw_parent)
                if not parent:
                    raise HTTPRedirect(
                        'edit_room_type?id={}&message={}', room_type.id,
                        "Selected parent room type not found.")
                # Don't allow chains: refuse if the chosen parent itself
                # follows another type.
                if parent.connects_to_type_id:
                    raise HTTPRedirect(
                        'edit_room_type?id={}&message={}', room_type.id,
                        "Cannot follow a room type that already follows another. "
                        "Chains are not supported.")
                room_type.connects_to_type_id = parent.id
                try:
                    room_type.connector_quantity = max(1, int(raw_qty or '1'))
                except ValueError:
                    room_type.connector_quantity = 1
            else:
                room_type.connects_to_type_id = None
                room_type.connector_quantity = 0

            session.add(room_type)
            session.commit()
            raise HTTPRedirect('manage_room_types?message={}', f"Room type '{room_type.name}' saved.")

        return {
            'room_type': room_type,
            'siblings': siblings,
            'children': children,
            'message': message,
        }

    # CSV/XLSX export + CSV import used by the export-tracking modal. The
    # column layout intentionally mirrors the JSON shape of the
    # `HotelLookup.export_room_bookings` API so that a hotel that prefers
    # spreadsheets to API calls can use the same field names.
    #
    # Credit-card vault tokens are deliberately omitted on export and
    # actively refused on import - the rest of the system treats them as
    # PCI-sensitive, so they never leave the database in a spreadsheet
    # and we won't accept new ones from one either.

    _BOOKING_BASE_COLS = [
        'assignment_id', 'lottery_application_id', 'parent_assignment_id',
        'confirmation_num', 'assignment_reason', 'status',
        'hotel', 'room_type', 'suite_type',
        'check_in_date', 'check_out_date',
        'hotel_confirmation_number', 'cancellation_confirmation_number',
        'legal_first_name', 'legal_last_name', 'cellphone', 'email',
        'address1', 'address2', 'city', 'region', 'zip_code', 'country',
        'wants_ada', 'ada_requests', 'special_requests',
        'last_modified_at', 'cc_captured_at', 'cc_last_four',
    ]
    _BOOKING_GUEST_FIELDS = [
        'legal_first_name', 'legal_last_name', 'cellphone', 'email',
    ]
    _BOOKING_MAX_GUESTS = 4

    def _booking_columns(self):
        cols = list(self._BOOKING_BASE_COLS)
        for i in range(1, self._BOOKING_MAX_GUESTS + 1):
            for f in self._BOOKING_GUEST_FIELDS:
                cols.append(f'guest{i}_{f}')
        return cols

    def _booking_row(self, ra, app):
        """Build the base (no-guest) column values for one RoomAssignment.

        Dates and datetimes are explicitly serialized to ISO 8601 strings
        (`YYYY-MM-DD` / `YYYY-MM-DDTHH:MM:SS+00:00`) - without this, xlsxwriter
        treats date objects as numeric serial values, which Excel displays as
        bare integers that look like opaque IDs to anyone opening the file.
        """
        def iso(v):
            return v.isoformat() if v else ''

        inv = ra.inventory
        hotel_name = inv.hotel.name if inv and inv.hotel else ''
        room_type_name = (inv.room_type.name
                          if inv and not inv.is_suite and inv.room_type else '')
        suite_type_name = (inv.suite_type.name
                           if inv and inv.is_suite and inv.suite_type else '')

        return [
            ra.id, ra.lottery_application_id or '', ra.parent_assignment_id or '',
            (app.confirmation_num if app else ''),
            ra.assignment_reason_label if hasattr(ra, 'assignment_reason_label') else ra.assignment_reason,
            ra.status_label if hasattr(ra, 'status_label') else ra.status,
            hotel_name, room_type_name, suite_type_name,
            iso(ra.assigned_check_in_date),
            iso(ra.assigned_check_out_date),
            ra.hotel_confirmation_number or '',
            ra.cancellation_confirmation_number or '',
            # Legal-name fields live on the attendee (hotel_first_name /
            # hotel_last_name with the legal/first/last fallback chain).
            (app.attendee.effective_hotel_first_name if app and app.attendee else '') or '',
            (app.attendee.effective_hotel_last_name if app and app.attendee else '') or '',
            (app.cellphone if app else '') or '',
            (app.email if app else '') or '',
            ra.address1 or '', ra.address2 or '', ra.city or '',
            ra.region or '', ra.zip_code or '', ra.country or '',
            ('yes' if (app and app.wants_ada) else ''),
            (app.ada_requests if app else '') or '',
            ra.special_requests or '',
            iso(ra.last_modified_at),
            iso(ra.cc_captured_at),
            ra.cc_last_four or '',
        ]

    def _booking_export_data(self, session, hotel_id):
        """Common query + per-row construction for both CSV and XLSX export.

        Source is now RoomAssignment - one row per assigned room.
        Connectors get their own line; their `parent_assignment_id`
        column points at the parent (suite) assignment's id so the hotel
        can group them.
        """
        hotel = session.query(LotteryHotel).filter_by(id=hotel_id).first()
        if not hotel:
            return None, []

        inv_ids = [str(inv.id) for inv in
                   session.query(HotelRoomInventory).filter_by(hotel_id=hotel.id).all()]
        if not inv_ids:
            return hotel, []

        assignments = (session.query(RoomAssignment)
                       .filter(RoomAssignment.inventory_id.in_(inv_ids),
                               RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]))
                       .order_by(RoomAssignment.parent_assignment_id.asc().nullsfirst(),
                                 RoomAssignment.created.asc())
                       .all())

        # Bulk-fetch each assignment's source LotteryApplication.
        app_ids = {ra.lottery_application_id for ra in assignments
                   if ra.lottery_application_id}
        apps_by_id = {}
        if app_ids:
            for app in session.query(LotteryApplication).filter(
                    LotteryApplication.id.in_(app_ids)).all():
                apps_by_id[app.id] = app

        rows = []
        for ra in assignments:
            app = apps_by_id.get(ra.lottery_application_id)
            row = self._booking_row(ra, app)

            # Guest columns: occupants take precedence (multi-room
            # explicit occupants); fall back to the application's
            # valid_group_members for back-compat with old data.
            members = list(getattr(ra, 'occupants', None) or [])
            if not members and app:
                members = list(app.valid_group_members or [])
            for i in range(self._BOOKING_MAX_GUESTS):
                if i < len(members):
                    m = members[i]
                    # `m` may be either an Attendee (from `ra.occupants`)
                    # or a LotteryApplication (from valid_group_members);
                    # both expose `.attendee` (the LA has a FK to one;
                    # Attendees return themselves via a tiny shim below).
                    a = getattr(m, 'attendee', None) or m
                    row += [a.effective_hotel_first_name or '',
                            a.effective_hotel_last_name or '',
                            (getattr(m, 'cellphone', '') or getattr(a, 'cellphone', '') or ''),
                            (getattr(m, 'email', '') or getattr(a, 'email', '') or '')]
                else:
                    row += ['', '', '', '']
            rows.append(row)

        return hotel, rows

    @csv_file
    def export_hotel_bookings_csv(self, out, session, hotel_id):
        """Per-hotel booking CSV. CC tokens omitted by design."""
        hotel, rows = self._booking_export_data(session, hotel_id)
        if hotel is None:
            return
        out.writerow(self._booking_columns())
        for row in rows:
            out.writerow(row)

    @xlsx_file
    def export_hotel_bookings_xlsx(self, out, session, hotel_id):
        """Per-hotel booking XLSX. CC tokens omitted by design."""
        hotel, rows = self._booking_export_data(session, hotel_id)
        if hotel is None:
            return
        out.writerow(self._booking_columns())
        for row in rows:
            out.writerow(row)

    def import_hotel_bookings_csv(self, session, hotel_id, csrf_token=None,
                                  bookings_csv=None, bookings_file=None, **params):
        """Back-import the same CSV/XLSX layout we export, populating
        hotel_confirmation_number (and cancellation_confirmation_number on
        the corresponding RoomAssignment, when present). Refuses any file
        that includes a column starting with `cc_token` - those don't
        belong in a spreadsheet.

        Accepts either a `.csv` or `.xlsx` upload. The form's file input
        is named `bookings_file`; the older `bookings_csv` name is kept
        as a fallback so any in-flight bookmarks keep working.
        """
        from uber.models import RoomAssignment
        from uber.utils import check_csrf as _check_csrf

        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('export_tracking')
        _check_csrf(csrf_token)

        upload = bookings_file or bookings_csv
        if not upload or not getattr(upload, 'file', None):
            raise HTTPRedirect(
                'export_tracking?message={}',
                'No file uploaded.')

        filename = (getattr(upload, 'filename', '') or '').lower()
        is_xlsx = filename.endswith('.xlsx') or filename.endswith('.xlsm')

        if is_xlsx:
            # XLSX path: openpyxl read-only mode to keep memory bounded.
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise HTTPRedirect(
                    'export_tracking?message={}',
                    "XLSX import is unavailable on this server: openpyxl "
                    "is not installed.")
            try:
                wb = load_workbook(upload.file, read_only=True, data_only=True)
            except Exception as e:
                raise HTTPRedirect(
                    'export_tracking?message={}',
                    f"Could not read uploaded XLSX file: {e}")
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                header = ()

            def _cell_to_str(v):
                # openpyxl returns datetimes for date cells - render them
                # as ISO so the existing _parse_iso_date matcher works.
                if v is None:
                    return ''
                from datetime import date as _d, datetime as _dt
                if isinstance(v, (_d, _dt)):
                    return v.isoformat()
                return str(v)

            fieldnames = [(h or '').strip() for h in header]
            rows = []
            for raw in rows_iter:
                if raw is None:
                    continue
                cells = list(raw)
                # Pad short rows so zip() doesn't drop trailing blanks.
                if len(cells) < len(fieldnames):
                    cells = cells + [None] * (len(fieldnames) - len(cells))
                rows.append({fn: _cell_to_str(cells[i])
                             for i, fn in enumerate(fieldnames) if fn})
            reader_fieldnames = fieldnames
            reader_iter = rows
        else:
            raw = upload.file.read()
            if isinstance(raw, bytes):
                raw = raw.decode('utf-8', errors='replace')
            reader = csv.DictReader(raw.splitlines())
            reader_fieldnames = reader.fieldnames or []
            reader_iter = reader

        # Guard against any column carrying CC vault tokens.
        sensitive = [
            f for f in (reader_fieldnames or [])
            if f and f.lower().startswith('cc_token')
        ]
        if sensitive:
            raise HTTPRedirect(
                'export_tracking?message={}',
                f"Refusing to import: file contains credit-card token "
                f"column(s) ({', '.join(sensitive)}). "
                f"Strip them and re-upload.")

        def _parse_iso_date(raw):
            """Accept ISO 8601 dates (YYYY-MM-DD) and any ISO 8601 datetime
            with optional offset - the hotel may quote either back. Returns
            None if the cell is blank or unparseable."""
            if not raw:
                return None
            raw = raw.strip()
            if not raw:
                return None
            # date.fromisoformat handles 'YYYY-MM-DD' on its own. For
            # datetimes we parse and project to date.
            try:
                return date.fromisoformat(raw[:10])
            except ValueError:
                return None

        updated = 0
        cancelled = 0
        date_updates = 0
        unmatched = []

        for row in reader_iter:
            app_id = (row.get('lottery_application_id') or '').strip()
            conf = (row.get('confirmation_num') or '').strip()
            new_conf = (row.get('hotel_confirmation_number') or '').strip()
            cancel_num = (row.get('cancellation_confirmation_number') or '').strip()
            new_ci = _parse_iso_date(row.get('check_in_date'))
            new_co = _parse_iso_date(row.get('check_out_date'))

            if not (new_conf or cancel_num or new_ci or new_co):
                continue  # Nothing to update for this row.

            app = None
            if app_id:
                app = session.query(LotteryApplication).filter_by(id=app_id).first()
            if not app and conf:
                app = (session.query(LotteryApplication)
                       .filter_by(confirmation_num=conf).first())
            if not app:
                unmatched.append(conf or app_id or '(blank)')
                continue

            if new_conf:
                # Hotel confirmation number lives on RoomAssignment only.
                # Write to every assignment for this app whose value differs.
                ras = (session.query(RoomAssignment)
                       .filter_by(lottery_application_id=app.id).all())
                touched_any = False
                for ra in ras:
                    if (ra.hotel_confirmation_number or '') != new_conf:
                        ra.hotel_confirmation_number = new_conf
                        session.add(ra)
                        _send_confirmation_updated_email(session, ra)
                        touched_any = True
                if touched_any:
                    updated += 1

            if cancel_num:
                ras = (session.query(RoomAssignment)
                       .filter_by(lottery_application_id=app.id).all())
                for ra in ras:
                    if ra.cancellation_confirmation_number != cancel_num:
                        ra.cancellation_confirmation_number = cancel_num
                        # The model's presave flips status to CANCELLED.
                        session.add(ra)
                        cancelled += 1

            # Date columns: persist directly to every matching
            # RoomAssignment row whose dates differ.
            ras_for_app = (session.query(RoomAssignment)
                           .filter_by(lottery_application_id=app.id).all())
            for ra in ras_for_app:
                touched = False
                if new_ci and ra.assigned_check_in_date != new_ci:
                    ra.assigned_check_in_date = new_ci
                    touched = True
                if new_co and ra.assigned_check_out_date != new_co:
                    ra.assigned_check_out_date = new_co
                    touched = True
                if touched:
                    session.add(ra)
                    date_updates += 1

        # One HotelExportLog entry summarizing the import.
        if updated or cancelled or date_updates:
            session.add(HotelExportLog(
                hotel_id=hotel_id,
                export_type='confirmation_import',
                record_count=updated + cancelled + date_updates,
                notes=(f"{('XLSX' if is_xlsx else 'CSV')} upload: "
                       f"{updated} confirmation(s), "
                       f"{cancelled} cancellation(s), "
                       f"{date_updates} date update(s)"),
            ))
        session.commit()

        msg_parts = []
        if updated:
            msg_parts.append(f"{updated} confirmation update(s)")
        if cancelled:
            msg_parts.append(f"{cancelled} cancellation(s)")
        if date_updates:
            msg_parts.append(f"{date_updates} date update(s)")
        if unmatched:
            msg_parts.append(
                f"{len(unmatched)} unmatched row(s): "
                f"{', '.join(unmatched[:5])}"
                f"{'...' if len(unmatched) > 5 else ''}")
        message = "; ".join(msg_parts) or "No matching rows to import."
        raise HTTPRedirect('export_tracking?message={}', message)

    def hotel_export_details(self, session, hotel_id, page='1', page_size='25'):
        """Per-room export/import detail for a single hotel, used by the
        modal on the export tracking page. Returns a server-rendered
        partial (no base layout) for direct injection into the modal body.

        Each booking gets a `sync_status` of:
          - in_sync:               exported, has confirmation #, not modified since
          - pending_export:        exported but modified after the export ran
          - awaiting_confirmation: exported, no hotel confirmation # yet
          - never_exported:        no export log for this hotel yet

        The list paginates server-side so a hotel with hundreds of bookings
        doesn't pile the full table into the modal at once. The modal's JS
        re-injects the partial when page links are clicked.
        """
        hotel = session.query(LotteryHotel).filter_by(id=hotel_id).first()
        if not hotel:
            return {'hotel': None, 'bookings': [], 'page': 1,
                    'page_size': 25, 'total': 0, 'page_count': 0}

        last_export = (session.query(HotelExportLog)
                       .filter(HotelExportLog.hotel_id == hotel.id,
                               HotelExportLog.export_type == 'room_export')
                       .order_by(HotelExportLog.exported_at.desc())
                       .first())

        try:
            page_num = max(1, int(page))
        except (TypeError, ValueError):
            page_num = 1
        try:
            ps = max(5, min(200, int(page_size)))
        except (TypeError, ValueError):
            ps = 25

        hotel_inventory_ids = [str(inv.id) for inv in
                               session.query(HotelRoomInventory).filter_by(hotel_id=hotel.id).all()]

        base_q = (session.query(RoomAssignment)
                  .filter(RoomAssignment.inventory_id.in_(hotel_inventory_ids),
                          RoomAssignment.status.in_([c.ASSIGNED, c.SECURED])))
        total = base_q.count()
        page_count = max(1, (total + ps - 1) // ps)
        if page_num > page_count:
            page_num = page_count

        assignments = (base_q
                       .order_by(RoomAssignment.parent_assignment_id.asc().nullsfirst(),
                                 RoomAssignment.created.asc())
                       .offset((page_num - 1) * ps)
                       .limit(ps)
                       .all())

        bookings = []
        for ra in assignments:
            has_conf = bool(ra.hotel_confirmation_number and ra.hotel_confirmation_number.strip())
            modified = ra.last_modified_at
            if not last_export:
                status = 'never_exported'
            elif modified and modified > last_export.exported_at:
                status = 'pending_export'
            elif not has_conf:
                status = 'awaiting_confirmation'
            else:
                status = 'in_sync'
            bookings.append({
                'assignment': ra,
                'app': ra.lottery_application,
                'sync_status': status,
            })

        return {
            'hotel': hotel,
            'last_export': last_export,
            'bookings': bookings,
            'page': page_num,
            'page_size': ps,
            'total': total,
            'page_count': page_count,
        }

    def export_tracking(self, session, message=''):
        hotels = []
        for hotel in session.query(LotteryHotel).filter_by(active=True).all():
            last_export = session.query(HotelExportLog).filter(
                HotelExportLog.hotel_id == hotel.id, HotelExportLog.export_type == 'room_export'
            ).order_by(HotelExportLog.exported_at.desc()).first()

            last_import = session.query(HotelExportLog).filter(
                HotelExportLog.hotel_id == hotel.id, HotelExportLog.export_type == 'confirmation_import'
            ).order_by(HotelExportLog.exported_at.desc()).first()

            hotel_inventory_ids = [str(inv.id) for inv in
                                    session.query(HotelRoomInventory).filter_by(hotel_id=hotel.id).all()]
            bookings = session.query(RoomAssignment).filter(
                RoomAssignment.inventory_id.in_(hotel_inventory_ids),
                RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            )

            total_bookings = bookings.count()
            missing_confirmation = bookings.filter(
                or_(RoomAssignment.hotel_confirmation_number == None,  # noqa: E711
                    RoomAssignment.hotel_confirmation_number == '')
            ).count()

            dirty_count = 0
            if last_export:
                dirty_count = bookings.filter(
                    RoomAssignment.last_modified_at > last_export.exported_at
                ).count()

            hotels.append({
                'hotel': hotel,
                'last_export': last_export,
                'last_import': last_import,
                'total_bookings': total_bookings,
                'missing_confirmation': missing_confirmation,
                'dirty_count': dirty_count,
            })

        import_files = session.query(HotelImportFile).order_by(
            HotelImportFile.uploaded_at.desc()).all()

        return {
            'hotels': hotels,
            'message': message,
            'import_files': import_files,
            'all_hotels': session.query(LotteryHotel).order_by(LotteryHotel.name).all(),
        }

    def upload_confirmation_file(self, session, hotel_id=None, message='', **params):
        """Admin upload of a hotel confirmation/cancellation file.

        Uses the same parsing, application, and file retention as the
        uber-vault hotel portal, so admin-uploaded files appear in the exports
        list alongside portal uploads.
        """
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('export_tracking')

        upload = params.get('import_file')
        if upload is None or not getattr(upload, 'file', None):
            raise HTTPRedirect('export_tracking?message={}', 'Please choose a file to upload.')

        raw = upload.file.read()
        if len(raw) > 5 * 1024 * 1024:
            raise HTTPRedirect('export_tracking?message={}', 'File is too large (5 MB max).')

        hotel = session.query(LotteryHotel).get(hotel_id) if hotel_id else None
        account = session.current_admin_account()
        uploaded_by = account.attendee.full_name if account and account.attendee else 'Admin'

        from uber.hotel_imports import import_confirmation_file
        result = import_confirmation_file(
            session, raw, getattr(upload, 'filename', ''), hotel=hotel,
            source='admin', uploaded_by=uploaded_by,
            content_type=getattr(upload, 'content_type', '') or '')

        if result.get('error'):
            message = f"File saved, but could not be parsed: {result['error']}"
        else:
            message = f"Imported {result['updated']} update(s), {result['unchanged']} unchanged."
        raise HTTPRedirect('export_tracking?message={}', message)

    def download_import_file(self, session, id):
        """Download a previously uploaded hotel import file."""
        record = session.query(HotelImportFile).get(id)
        if not record or not record.filepath or not os.path.exists(record.filepath):
            raise cherrypy.HTTPError(404, "File not found")
        return serve_file(record.filepath, disposition='attachment',
                          name=record.filename or os.path.basename(record.filepath),
                          content_type=record.content_type or 'application/octet-stream')

    def run_lottery(self, session, lottery_group="attendee", lottery_type="room", run_name="", **params):
        if lottery_type == "room":
            lottery_type_val = c.ROOM_ENTRY
        elif lottery_type == "suite":
            lottery_type_val = c.SUITE_ENTRY
        else:
            return {'error': f'Invalid lottery type: {lottery_type}'}
        applications = session.query(LotteryApplication).join(LotteryApplication.attendee
                                                              ).filter(LotteryApplication.status == c.COMPLETE,
                                                                       Attendee.hotel_lottery_eligible == True)

        cutoff = None
        if params.get('cutoff', ''):
            cutoff = dateparser.parse(params['cutoff']).replace(tzinfo=c.EVENT_TIMEZONE)
            applications = applications.filter(LotteryApplication.last_submitted < cutoff)

        # Optional re-confirmation gate. When set, only apps whose
        # attendee has clicked Confirm since this datetime are considered.
        # Apps that were awarded and then expired sit in COMPLETE and must
        # re-confirm when the admin sets this filter.
        confirmation_window_start = None
        if params.get('confirmation_window_start', ''):
            confirmation_window_start = dateparser.parse(
                params['confirmation_window_start']).replace(tzinfo=c.EVENT_TIMEZONE)
            applications = applications.filter(
                LotteryApplication.last_confirmed_at.isnot(None),
                LotteryApplication.last_confirmed_at >= confirmation_window_start,
            )

        # We always grab all roommate entries, but the solver only looks at those that have a matching parent
        # in the lottery batch.
        if lottery_type_val == c.SUITE_ENTRY:
            applications = applications.filter(LotteryApplication.entry_type.in_([lottery_type_val, c.GROUP_ENTRY]))
        else:
            applications = applications.filter(or_(LotteryApplication.entry_type.in_([lottery_type_val, c.GROUP_ENTRY]),
                                                   LotteryApplication.room_opt_out == False))

        # If lottery_group is "both" don't filter either way
        if lottery_group == "staff":
            applications = applications.filter(LotteryApplication.is_staff_entry == True)
        elif lottery_group == "attendee":
            applications = applications.filter(LotteryApplication.is_staff_entry == False)

        applications = applications.all()

        # Count already-assigned rooms per inventory block per night,
        # sourced from RoomAssignment. Connector rooms count against their
        # own inventory's capacity; primary rooms count against theirs -
        # both are RoomAssignment rows.
        already_assigned_query = session.query(RoomAssignment).filter(
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            RoomAssignment.inventory_id.isnot(None),
        )

        # Partition filter: only count same-partition assignments toward capacity
        partition_filter = params.get('partition_filter', '')
        inventory_filter = params.get('inventory_filter', '')

        if partition_filter:
            already_assigned = already_assigned_query.filter(
                RoomAssignment.partition_id == partition_filter).all()
        else:
            already_assigned = already_assigned_query.filter(
                RoomAssignment.partition_id == None).all()  # noqa: E711

        assigned_per_block_night = defaultdict(lambda: defaultdict(int))
        for ra in already_assigned:
            if ra.assigned_check_in_date and ra.assigned_check_out_date:
                day = ra.assigned_check_in_date
                while day < ra.assigned_check_out_date:
                    assigned_per_block_night[str(ra.inventory_id)][day.isoformat()] += 1
                    day += timedelta(days=1)

        is_suite = lottery_type_val == c.SUITE_ENTRY
        inventory_table = HotelRoomInventory.get_inventory(session, is_suite=is_suite)

        # Apply filters
        hotel_filter = params.get('hotel_filter', '')
        room_type_filter = params.get('room_type_filter', '')
        if hotel_filter:
            hotel_filter_set = set(hotel_filter.split(','))
            inventory_table = [r for r in inventory_table if r['hotel_id'] in hotel_filter_set]
        if room_type_filter:
            room_type_filter_set = set(room_type_filter.split(','))
            inventory_table = [r for r in inventory_table if r['room_type'] in room_type_filter_set]
        if inventory_filter:
            inventory_filter_set = set(inventory_filter.split(','))
            inventory_table = [r for r in inventory_table if r['id'] in inventory_filter_set]

        # Build partition allocation maps
        partition_qty_map = {}  # {block_id: cap} for the selected partition
        total_partitioned_map = {}  # {block_id: total across all partitions}
        if partition_filter:
            for pb in session.query(InventoryPartitionBlock).filter_by(partition_id=partition_filter).all():
                partition_qty_map[str(pb.inventory_id)] = pb.quantity
        else:
            # For non-partitioned runs, compute total partitioned allocation per block
            for pb in session.query(InventoryPartitionBlock).all():
                bid = str(pb.inventory_id)
                total_partitioned_map[bid] = total_partitioned_map.get(bid, 0) + pb.quantity

        available_rooms = deepcopy(inventory_table)
        for block in available_rooms:
            block_id = block['id']

            if partition_filter:
                # Partitioned run: cap at partition allocation
                partition_cap = partition_qty_map.get(block_id, 0)
                if partition_cap <= 0:
                    block['quantity'] = 0
                    if block.get('night_quantities'):
                        block['night_quantities'] = {k: 0 for k in block['night_quantities']}
                    continue
                block['quantity'] = min(block['quantity'], partition_cap)
                if block.get('night_quantities'):
                    block['night_quantities'] = {k: min(v, partition_cap) for k, v in block['night_quantities'].items()}
            else:
                # Non-partitioned run: subtract total partition allocations from capacity
                reserved = total_partitioned_map.get(block_id, 0)
                if reserved:
                    block['quantity'] = max(0, block['quantity'] - reserved)
                    if block.get('night_quantities'):
                        block['night_quantities'] = {k: max(0, v - reserved) for k, v in block['night_quantities'].items()}

            if block.get('night_quantities'):
                for night_iso, qty in block['night_quantities'].items():
                    already = assigned_per_block_night.get(block_id, {}).get(night_iso, 0)
                    block['night_quantities'][night_iso] = max(0, qty - already)
            else:
                total_assigned = sum(assigned_per_block_night.get(block_id, {}).values())
                if total_assigned:
                    block['quantity'] = max(0, block['quantity'] - total_assigned)

        rooms_available_before = sum([x['quantity'] for x in available_rooms])

        # Build type-level connector map for the solver. Each child type
        # points at its parent + how many of itself the parent needs.
        connector_map = {}
        for rt in session.query(LotteryRoomType).filter(
                LotteryRoomType.connects_to_type_id.isnot(None),
                LotteryRoomType.connector_quantity > 0).all():
            connector_map[str(rt.id)] = (str(rt.connects_to_type_id), rt.connector_quantity)

        allocations = solve_lottery(
            applications, available_rooms,
            lottery_type=lottery_type_val,
            connector_map=connector_map,
        ) or []

        # Create LotteryRun record
        lottery_run = LotteryRun(
            name=run_name or f"{lottery_group}_{lottery_type}_{localized_now().strftime('%Y%m%d_%H%M%S')}",
            lottery_group=lottery_group,
            lottery_type=lottery_type,
            cutoff=cutoff,
            confirmation_window_start=confirmation_window_start,
            hotel_filter=hotel_filter or None,
            room_type_filter=room_type_filter or None,
            inventory_filter=inventory_filter or None,
            partition_filter=partition_filter or None,
            entries_considered=len([x for x in applications if x.entry_type != c.GROUP_ENTRY]),
            rooms_available_before=rooms_available_before,
        )
        session.add(lottery_run)
        session.flush()

        # Deadline stamped on every RoomAssignment from this run (per-assignment
        # override remains possible later).
        run_deadline = lottery_run.card_deadline.date() if lottery_run.card_deadline else None

        # Move the awarded leaders' apps to PROCESSED so they can be moved
        # to AWARDED in the award_run step. Group members come along via
        # parent_application_id; they stay COMPLETE on the app side and
        # are added as occupants of the leader's RoomAssignment rows.
        awarded_leader_ids = {leader_id for leader_id, _inv, _role in allocations}
        for application in applications:
            if application.id in awarded_leader_ids and not application.parent_application:
                application.lottery_run_id = lottery_run.id
                application.status = c.PROCESSED
                if partition_filter:
                    application.partition_id = partition_filter
                session.add(application)

        # Materialize RoomAssignment rows (primary + connectors) per leader.
        num_rooms_assigned = self._materialize_room_assignments(
            session, applications, allocations, lottery_run, run_deadline,
            partition_filter)

        lottery_run.rooms_assigned = num_rooms_assigned
        session.commit()

        raise HTTPRedirect('lottery_run_detail?id={}&message={}', lottery_run.id,
                           f"Lottery run complete: {num_rooms_assigned} rooms assigned.")
    
    def hotel_inventory(self, session, message='', partition='all'):
        # Build event night dates
        event_nights = []
        day = c.HOTEL_LOTTERY_CHECKIN_START.date()
        end = c.HOTEL_LOTTERY_CHECKOUT_END.date()
        while day < end:
            event_nights.append(day)
            day += timedelta(days=1)

        partitions = session.query(InventoryPartition).filter_by(active=True).order_by(InventoryPartition.name).all()

        # Determine partition filter for assigned/waitlist counting
        # partition='all' -> count all apps, show full capacity
        # partition='default' -> count only non-partitioned apps, show capacity minus partition allocations
        # partition=<uuid> -> count only that partition's apps, show partition allocation as capacity
        filter_partition_id = None
        filtering_default = False
        if partition == 'default':
            filtering_default = True
        elif partition not in ('all', ''):
            filter_partition_id = partition

        # Get assigned rooms, optionally filtered by partition, sourced from
        # RoomAssignment. Waitlist demand is read straight off the
        # assignment's waitlisted_* columns, which reflect the current
        # per-room request rather than the original lottery entry.
        ra_query = (session.query(RoomAssignment)
                    .filter(RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
                            RoomAssignment.inventory_id.isnot(None)))
        if filter_partition_id:
            ra_query = ra_query.filter(RoomAssignment.partition_id == filter_partition_id)
        elif filtering_default:
            ra_query = ra_query.filter(RoomAssignment.partition_id == None)  # noqa: E711
        assigned_ras = ra_query.all()

        # Build per-block per-night assignment counts + status counts.
        assigned_per_block_night = defaultdict(lambda: defaultdict(int))
        status_per_block = defaultdict(lambda: defaultdict(int))
        for ra in assigned_ras:
            block_id = str(ra.inventory_id)
            status_per_block[block_id][ra.status] += 1
            if ra.assigned_check_in_date and ra.assigned_check_out_date:
                d = ra.assigned_check_in_date
                while d < ra.assigned_check_out_date:
                    assigned_per_block_night[block_id][d] += 1
                    d += timedelta(days=1)

        # Build per-block per-night waitlist demand from each
        # assignment's own waitlisted_* range. Either column NULL means
        # no demand on that end; we coalesce to the assigned date so the
        # gap calculation is symmetric.
        waitlist_per_block_night = defaultdict(lambda: defaultdict(int))
        for ra in assigned_ras:
            if ra.status != c.SECURED:
                continue
            if not (ra.waitlisted_check_in_date or ra.waitlisted_check_out_date):
                continue
            block_id = str(ra.inventory_id)
            wl_ci = ra.waitlisted_check_in_date or ra.assigned_check_in_date
            wl_co = ra.waitlisted_check_out_date or ra.assigned_check_out_date
            if wl_ci and ra.assigned_check_in_date and wl_ci < ra.assigned_check_in_date:
                d = wl_ci
                while d < ra.assigned_check_in_date:
                    waitlist_per_block_night[block_id][d] += 1
                    d += timedelta(days=1)
            if wl_co and ra.assigned_check_out_date and wl_co > ra.assigned_check_out_date:
                d = ra.assigned_check_out_date
                while d < wl_co:
                    waitlist_per_block_night[block_id][d] += 1
                    d += timedelta(days=1)

        # Build partition allocation maps for capacity adjustments
        partition_alloc_per_block = {}  # {block_id: allocation} for the selected partition
        total_partitioned_per_block = defaultdict(int)  # {block_id: sum of all partition allocations}
        if filter_partition_id:
            for pb in session.query(InventoryPartitionBlock).filter_by(partition_id=filter_partition_id).all():
                partition_alloc_per_block[str(pb.inventory_id)] = pb.quantity
        for pb in session.query(InventoryPartitionBlock).all():
            total_partitioned_per_block[str(pb.inventory_id)] += pb.quantity

        hotel_lookup = {str(h.id): h for h in session.query(LotteryHotel).all()}

        def effective_capacity(block_id, block_qty):
            if filter_partition_id:
                return min(partition_alloc_per_block.get(block_id, 0), block_qty)
            elif filtering_default:
                return max(0, block_qty - total_partitioned_per_block.get(block_id, 0))
            else:
                return block_qty

        def build_inventory_data(is_suite):
            inventory = defaultdict(list)
            for inv in session.query(HotelRoomInventory).filter_by(is_suite=is_suite, active=True).all():
                hotel_obj = hotel_lookup.get(str(inv.hotel_id))
                block_id = str(inv.id)
                nq_map = inv.night_quantity_map

                night_data = []
                for night in event_nights:
                    raw_qty = nq_map.get(night, inv.quantity) if nq_map else inv.quantity
                    available = effective_capacity(block_id, raw_qty)
                    assigned = assigned_per_block_night.get(block_id, {}).get(night, 0)
                    waitlisted = waitlist_per_block_night.get(block_id, {}).get(night, 0)
                    night_data.append({
                        'night': night,
                        'available': available,
                        'assigned': assigned,
                        'remaining': max(0, available - assigned),
                        'waitlisted': waitlisted,
                    })

                total_assigned = sum(status_per_block.get(block_id, {}).values())
                info = {
                    'inventory': inv,
                    'room_type': inv.suite_type if is_suite else inv.room_type,
                    'quantity': effective_capacity(block_id, inv.quantity),
                    'nights': night_data,
                    'total_assigned': total_assigned,
                    c.PROCESSED: status_per_block.get(block_id, {}).get(c.PROCESSED, 0),
                    c.AWARDED: status_per_block.get(block_id, {}).get(c.AWARDED, 0),
                    c.SECURED: status_per_block.get(block_id, {}).get(c.SECURED, 0),
                }
                inventory[hotel_obj].append(info)
            return inventory

        # Build chart data: per-night totals by hotel
        chart_data = {}
        for inv in session.query(HotelRoomInventory).filter_by(active=True).all():
            hotel_name = hotel_lookup.get(str(inv.hotel_id), None)
            hotel_name = hotel_name.name if hotel_name else 'Unknown'
            if hotel_name not in chart_data:
                chart_data[hotel_name] = {
                    'available': [0] * len(event_nights),
                    'assigned': [0] * len(event_nights),
                    'waitlisted': [0] * len(event_nights),
                }
            nq_map = inv.night_quantity_map
            block_id = str(inv.id)
            for i, night in enumerate(event_nights):
                raw_qty = nq_map.get(night, inv.quantity) if nq_map else inv.quantity
                chart_data[hotel_name]['available'][i] += effective_capacity(block_id, raw_qty)
                chart_data[hotel_name]['assigned'][i] += assigned_per_block_night.get(block_id, {}).get(night, 0)
                chart_data[hotel_name]['waitlisted'][i] += waitlist_per_block_night.get(block_id, {}).get(night, 0)

        # Build combined total across all hotels
        total = {'available': [0] * len(event_nights), 'assigned': [0] * len(event_nights), 'waitlisted': [0] * len(event_nights)}
        for data in chart_data.values():
            for i in range(len(event_nights)):
                total['available'][i] += data['available'][i]
                total['assigned'][i] += data['assigned'][i]
                total['waitlisted'][i] += data['waitlisted'][i]
        chart_data['_total'] = total

        return {
            'room_inventory': build_inventory_data(is_suite=False),
            'suite_inventory': build_inventory_data(is_suite=True),
            'event_nights': event_nights,
            'partitions': partitions,
            'chart_data': chart_data,
            'now': localized_now(),
            'current_partition': partition,
        }

    @ajax
    def inventory_assignees(self, session, inventory_id, night_date='', partition='all'):
        query = session.query(RoomAssignment).filter(
            RoomAssignment.inventory_id == inventory_id,
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
        )
        if night_date:
            nd = date.fromisoformat(night_date)
            query = query.filter(
                RoomAssignment.assigned_check_in_date <= nd,
                RoomAssignment.assigned_check_out_date > nd,
            )
        if partition == 'default':
            query = query.filter(RoomAssignment.partition_id == None)  # noqa: E711
        elif partition not in ('all', ''):
            query = query.filter(RoomAssignment.partition_id == partition)

        assignments = query.order_by(RoomAssignment.assigned_check_in_date).all()
        assignees = []
        for ra in assignments:
            app = ra.lottery_application
            assignees.append({
                'assignment_id': ra.id,
                'app_id': ra.lottery_application_id or '',
                'attendee_id': str(ra.attendee_id) if ra.attendee_id else '',
                'name': (app.attendee_name if app else (ra.attendee.full_name if ra.attendee else '')),
                'conf_num': (app.confirmation_num if app else '') or '',
                'status': ra.status_label,
                'check_in': ra.assigned_check_in_date.strftime('%a %-m/%-d') if ra.assigned_check_in_date else '',
                'check_out': ra.assigned_check_out_date.strftime('%a %-m/%-d') if ra.assigned_check_out_date else '',
                'partition': ra.partition.name if ra.partition_id and ra.partition else '',
            })
        return {'assignees': assignees}

    @csv_file
    def assigned_entries(self, out, session, lock_entries=''):
        out.writerow(['LotteryRunName', 'StaffEntry?', 'AssignmentReason',
                      'CheckInDate', 'CheckOutDate', 'NumberofGuests', 'HotelName', 'RoomType', 'SpecialRequest', 'AccessibleRoom',
                      'RewardsNumber',
                      'Guest1CheckInDate', 'Guest1CheckOutDate', 'Guest1FirstName', 'Guest1LastName', 'Guest1Phone', 'Guest1Email',
                      'Guest2CheckInDate', 'Guest2CheckOutDate', 'Guest2FirstName', 'Guest2LastName', 'Guest2Phone', 'Guest2Email',
                      'Guest3CheckInDate', 'Guest3CheckOutDate', 'Guest3FirstName', 'Guest3LastName', 'Guest3Phone', 'Guest3Email',
                      'Guest4CheckInDate', 'Guest4CheckOutDate', 'Guest4FirstName', 'Guest4LastName', 'Guest4Phone', 'Guest4Email',])

        # Source is RoomAssignment: one row per assigned room
        # (connectors included). Each line is the hotel's view of one
        # physical reservation.
        assignments = (session.query(RoomAssignment)
                       .filter(RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]))
                       .order_by(RoomAssignment.inventory_id, RoomAssignment.created))

        if lock_entries:
            # Lock the SOURCE applications (one app may produce many rooms).
            app_ids_to_lock = {ra.lottery_application_id for ra in assignments
                               if ra.lottery_application_id}
            if app_ids_to_lock:
                for app in session.query(LotteryApplication).filter(
                        LotteryApplication.id.in_(app_ids_to_lock)).all():
                    if not app.export_locked:
                        app.export_locked = True
                        session.add(app)
                session.commit()

        for ra in assignments:
            app = ra.lottery_application
            inv = ra.inventory
            check_in_date = ra.assigned_check_in_date
            check_out_date = ra.assigned_check_out_date
            occupants = list(getattr(ra, 'occupants', None) or [])
            if not occupants and app:
                occupants = [app.attendee] + list(app.valid_group_members or [])
            num_guests = len(occupants)
            hotel_name = (inv.hotel.name if inv and inv.hotel else '')
            if inv and inv.is_suite and inv.suite_type:
                room_type_name = inv.suite_type.name
            elif inv and not inv.is_suite and inv.room_type:
                room_type_name = inv.room_type.name
            else:
                room_type_name = ''
            row = [
                lottery_run := (ra.lottery_run.name if ra.lottery_run else ''),
                (app.is_staff_entry if app else False),
                ra.assignment_reason_label,
                check_in_date, check_out_date, num_guests, hotel_name, room_type_name,
                (app.ada_requests if app else '') or '',
                (app.wants_ada if app else False),
                ra.hotel_rewards_number or '',
            ]
            for i in range(4):
                if i < len(occupants):
                    o = occupants[i]
                    # `o` is an Attendee or a LotteryApplication.attendee;
                    # either way the effective-hotel-name properties
                    # resolve to the right hotel-facing legal name.
                    a = getattr(o, 'attendee', None) or o
                    row.extend([check_in_date, check_out_date,
                                getattr(a, 'effective_hotel_first_name', '') or '',
                                getattr(a, 'effective_hotel_last_name', '') or '',
                                getattr(o, 'cellphone', '') or '',
                                getattr(o, 'email', '') or ''])
                else:
                    row.extend(['', '', '', '', '', ''])
            out.writerow(row)
    
    @xlsx_file
    def hotel_inventory_xlsx(self, out, session, hotel_id):
        rows = []

        hotel_inventory_ids = [str(inv.id) for inv in
                               session.query(HotelRoomInventory).filter_by(hotel_id=hotel_id).all()]
        assignments_q = session.query(RoomAssignment).filter(
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            RoomAssignment.inventory_id.in_(hotel_inventory_ids),
        )

        first_assignment = assignments_q.order_by(RoomAssignment.assigned_check_in_date).first()
        if not first_assignment:
            return  # No assignments for this hotel
        earliest_check_in = first_assignment.assigned_check_in_date
        latest_check_out = assignments_q.order_by(
            RoomAssignment.assigned_check_out_date.desc()).first().assigned_check_out_date
        date_range = [earliest_check_in + timedelta(days=x)
                      for x in range(0, (latest_check_out - earliest_check_in).days)] + [latest_check_out]

        inv_by_room_type = defaultdict(list)
        inv_by_suite_type = defaultdict(list)
        for inv in session.query(HotelRoomInventory).filter(
                HotelRoomInventory.id.in_(hotel_inventory_ids)).all():
            if inv.is_suite:
                inv_by_suite_type[str(inv.suite_type_id)].append(str(inv.id))
            else:
                inv_by_room_type[str(inv.room_type_id)].append(str(inv.id))

        header_row = [''] + [d.strftime("%A %-m/%-d") for d in date_range]
        for rt in session.query(LotteryRoomType).filter_by(
                is_suite=False, active=True).order_by(LotteryRoomType.name).all():
            inv_ids = inv_by_room_type.get(str(rt.id), [])
            row = [rt.name]
            for d in date_range:
                row.append(
                    assignments_q.filter(
                        RoomAssignment.inventory_id.in_(inv_ids),
                        RoomAssignment.assigned_check_in_date <= d,
                        RoomAssignment.assigned_check_out_date >= d,
                    ).count() if inv_ids else 0)
            rows.append(row)

        has_suites = any(inv_by_suite_type.values())
        if has_suites:
            for st in session.query(LotteryRoomType).filter_by(
                    is_suite=True, active=True).order_by(LotteryRoomType.name).all():
                inv_ids = inv_by_suite_type.get(str(st.id), [])
                row = [st.name]
                for d in date_range:
                    row.append(
                        assignments_q.filter(
                            RoomAssignment.inventory_id.in_(inv_ids),
                            RoomAssignment.assigned_check_in_date <= d,
                            RoomAssignment.assigned_check_out_date >= d,
                        ).count() if inv_ids else 0)
                rows.append(row)

        out.writerows(header_row, rows)

    @multifile_zipfile
    def hotel_inventory_zip(self, zip_file, session):
        for hotel in session.query(LotteryHotel).filter_by(active=True).all():
            hotel_inv_ids = [str(inv.id) for inv in
                             session.query(HotelRoomInventory).filter_by(hotel_id=hotel.id).all()]
            has_assignments = session.query(RoomAssignment).filter(
                RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
                RoomAssignment.inventory_id.in_(hotel_inv_ids),
            ).first()
            if has_assignments:
                output = self.hotel_inventory_xlsx(hotel_id=hotel.id, set_headers=False)
                zip_file.writestr(f'hotel_inventory_{hotel.name}.xlsx', output)

    @csv_file
    def accepted_dealers(self, out, session):
        out.writerow(['Group Name', 'Group ID', 'Reg ID'])

        for dealer in session.query(Attendee).join(Group, Attendee.group_id == Group.id).filter(
            Group.is_dealer, Group.status.in_(c.DEALER_ACCEPTED_STATUSES)):
            out.writerow([dealer.group.name, dealer.group.id, dealer.id])

    @csv_file
    def interchange_export(self, out, session, staff_lottery=False):
        def print_dt(dt):
            if not dt:
                return ""

            if isinstance(dt, datetime):
                return dt.astimezone(c.EVENT_TIMEZONE).strftime('%m/%d/%Y %H:%M:%S')
            else:
                return dt.strftime('%m/%d/%Y')
        
        def print_bool(bool):
            return "TRUE" if bool else "FALSE"

        country_codes = {}
        for country in list(pycountry.countries):
            value = country.name if "Taiwan" not in country.name else "Taiwan"
            country_codes[value] = f"{country.alpha_2};{country.name}"

        header_row = []
        # Config data and IDs
        header_row.extend(["Lottery Close", "suite_cutoff", "year", "Response ID", "Confirmation Code",
                           "SessionID", "Survey ID", "entry_id", "dealer_group_id"])

        # Contact data
        header_row.extend(["is_staff", "email", "first_name:contact", "last_name:contact", "Title:contact",
                           "Company Name:contact", "street_address:contact", "apt_suite_office:contact",
                           "city:contact", "state:contact", "zip:contact", "country:contact", "phone:contact",
                           "Mobile Phone:contact"])

        # Entry metadata
        header_row.extend(["Time Started", "Date Submitted", "entry_confirmed", "Status", "edit_link", "I agree:agree",
                           "Comments", "payment_valid", "reg_conf_code", "entry_type", "Referer",
                           "User Agent", "IP Address", "Longitude", "Latitude", "Country", "City", "State/Region",
                           "Postal"])

        # Entry data
        header_row.extend(["group_conf", "group_email", "Yes:age_ack", "special_room", "ada_req_text",
                           "I agree, understand and will comply:suite_agree",
                           "desired_arrival", "latest_arrival", "desired_departure", "earliest_departure"])

        all_hotels = session.query(LotteryHotel).filter_by(active=True).order_by(LotteryHotel.name).all()
        all_room_types = session.query(LotteryRoomType).filter_by(is_suite=False, active=True).order_by(LotteryRoomType.name).all()
        all_suite_types = session.query(LotteryRoomType).filter_by(is_suite=True, active=True).order_by(LotteryRoomType.name).all()

        for hotel in all_hotels:
            header_row.append(f"{hotel.export_name or hotel.name}:hotel_pref")

        for rt in all_room_types:
            header_row.append(f"{rt.export_name or rt.name}:room_pref")

        for st in all_suite_types:
            header_row.append(f"{st.export_name or st.name}:suite_type")

        out.writerow(header_row)

        applications = session.query(LotteryApplication).join(LotteryApplication.attendee
                                                              ).filter(LotteryApplication.status != c.PROCESSED,
                                                                       Attendee.hotel_lottery_eligible == True)
        if staff_lottery:
            applications = applications.filter(LotteryApplication.is_staff_entry == True)
        else:
            applications = applications.filter(LotteryApplication.is_staff_entry == False)

        for app in applications:
            attendee = app.attendee
            row = []

            # Config data and IDs
            dealer_id = ''
            if app.attendee.is_dealer and app.attendee.group and app.attendee.group.status in c.DEALER_ACCEPTED_STATUSES:
                dealer_id = app.attendee.group.id
            current_lottery_deadline = c.HOTEL_LOTTERY_STAFF_DEADLINE if app.is_staff_entry else c.HOTEL_LOTTERY_FORM_DEADLINE
            row.extend([datetime_local_filter(current_lottery_deadline), datetime_local_filter(c.HOTEL_LOTTERY_SUITE_CUTOFF),
                        c.EVENT_YEAR, app.response_id, app.confirmation_num, app.id, "RAMS_1", app.id, dealer_id])

            # Contact data
            base_cellphone = app.cellphone or app.attendee.cellphone
            row.extend([print_bool(attendee.badge_type == c.STAFF_BADGE or c.STAFF_RIBBON in attendee.ribbon_ints),
                        attendee.email,
                        attendee.effective_hotel_first_name,
                        attendee.effective_hotel_last_name,
                        "", "", attendee.address1,
                        attendee.address2, attendee.city, attendee.region, attendee.zip_code,
                        country_codes.get(attendee.country, attendee.country),
                        ''.join(filter(str.isdigit, base_cellphone)) if base_cellphone else "", ""])

            # Entry metadata
            if app.entry_type:
                type_str = "I am entering as a roommate" if app.entry_type == c.GROUP_ENTRY else "I am requesting a room"
            else:
                type_str = "I am withdrawing from the lottery"
            row.extend([print_dt(app.entry_started), print_dt(app.last_submitted), print_bool(app.status == c.COMPLETE),
                        app.status_label, f"{c.URL_BASE}/hotel_lottery/index?attendee_id={app.attendee.id}",
                        print_bool(app.terms_accepted), app.admin_notes, "FALSE", attendee.id, type_str])
            if app.entry_metadata:
                row.extend([app.entry_metadata.get('referer'), app.entry_metadata.get('user_agent'), app.entry_metadata.get('ip_address')])
            else:
                row.extend(['', '', ''])
            row.extend(['', '', '', '', '', ''])

            # Entry data
            if app.parent_application:
                row.extend([app.parent_application.confirmation_num, app.parent_application.email,
                            '', '', '', '', '', '', '', ''])
            else:
                row.extend(['', '', print_bool(app.entry_form_completed)])

                entry_type_base = app.entry_type or c.ROOM_ENTRY
                if entry_type_base == c.ROOM_ENTRY:
                    if app.wants_ada:
                        row.extend(['ADA Room', app.ada_requests, ''])
                    else:
                        row.extend(['Standard Rooms with no Special Requests', '', ''])
                elif entry_type_base == c.SUITE_ENTRY:
                    row.extend(['Hyatt Regency O\'Hare Suites', '', print_bool(app.suite_terms_accepted)])
                
                row.extend([print_dt(app.earliest_checkin_date), print_dt(app.latest_checkin_date),
                            print_dt(app.latest_checkout_date), print_dt(app.earliest_checkout_date)])

            if app.parent_application or not app.hotel_preference or (
                    app.entry_type and app.entry_type == c.SUITE_ENTRY and app.room_opt_out):
                row.extend(['' for _ in range(len(all_hotels))])
            else:
                hotels_ranking = {}
                for index, item in enumerate(app.hotel_preference.split(','), start=1):
                    hotels_ranking[item] = index

                for hotel in all_hotels:
                    row.append(hotels_ranking.get(str(hotel.id), ''))

            if app.parent_application or not app.room_type_preference or (
                    app.entry_type and app.entry_type == c.SUITE_ENTRY and app.room_opt_out):
                row.extend(['' for _ in range(len(all_room_types))])
            else:
                room_types_ranking = {}
                for index, item in enumerate(app.room_type_preference.split(','), start=1):
                    room_types_ranking[item] = index

                for rt in all_room_types:
                    row.append(room_types_ranking.get(str(rt.id), ''))

            if app.parent_application or not app.suite_type_preference or (
                    app.entry_type and app.entry_type == c.ROOM_ENTRY):
                row.extend(['' for _ in range(len(all_suite_types))])
            else:
                suite_types_ranking = {}
                for index, item in enumerate(app.suite_type_preference.split(','), start=1):
                    suite_types_ranking[item] = index

                for st in all_suite_types:
                    row.append(suite_types_ranking.get(str(st.id), ''))

            out.writerow(row)

    def manage_partitions(self, session, message=''):
        partitions = session.query(InventoryPartition).order_by(InventoryPartition.name).all()

        # Count assigned per partition (RoomAssignment-sourced).
        assigned_per_partition = defaultdict(int)
        for ra in session.query(RoomAssignment).filter(
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            RoomAssignment.inventory_id.isnot(None),
        ).all():
            key = str(ra.partition_id) if ra.partition_id else '_none'
            assigned_per_partition[key] += 1

        # Compute total allocation and non-partitioned capacity
        total_partition_alloc = 0
        for p in partitions:
            for b in p.blocks:
                total_partition_alloc += b.quantity

        total_inventory = sum(inv.quantity for inv in session.query(HotelRoomInventory).filter_by(active=True).all())

        return {
            'partitions': partitions,
            'assigned_per_partition': assigned_per_partition,
            'non_partitioned_capacity': max(0, total_inventory - total_partition_alloc),
            'message': message,
        }

    def request_confirmation(self, session, id=None, clear='', csrf_token=None):
        """Set or clear LotteryApplication.confirmation_requested_at.

        When set, the attendee's status page surfaces the Confirm / Withdraw
        prompt and the reconfirm email fires. Clearing removes the prompt
        without touching last_confirmed_at.
        """
        from uber.utils import check_csrf
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('form?id={}', id)
        check_csrf(csrf_token)

        app = session.query(LotteryApplication).get(id)
        if not app:
            raise HTTPRedirect('index?message={}', 'Application not found.')
        if clear:
            app.confirmation_requested_at = None
            msg = 'Confirmation request cleared.'
        else:
            app.confirmation_requested_at = datetime.now(UTC)
            msg = 'Confirmation request sent.'
        session.add(app)
        session.commit()
        raise HTTPRedirect('form?id={}&message={}', id, msg)

    def bulk_request_confirmation(self, session, app_ids='', csrf_token=None):
        """Stamp confirmation_requested_at on many applications at once.

        Accepts a comma-separated string of application IDs from a search-
        results checkbox UI.
        """
        from uber.utils import check_csrf
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('index')
        check_csrf(csrf_token)

        ids = [x.strip() for x in app_ids.split(',') if x.strip()]
        if not ids:
            raise HTTPRedirect('index?message={}', 'No applications selected.')

        apps = session.query(LotteryApplication).filter(
            LotteryApplication.id.in_(ids)).all()
        now = datetime.now(UTC)
        for app in apps:
            app.confirmation_requested_at = now
            session.add(app)
        session.commit()
        raise HTTPRedirect(
            'index?message={}',
            f'Confirmation requested for {len(apps)} application(s).')

    def import_hotel_cancellations(self, session, message='', **params):
        """Back-import cancellations from the hotel.

        Expected CSV columns (header row required):
          - confirmation_num         (matches RoomAssignment.hotel_confirmation_number)
          - cancellation_confirmation_number  (the hotel's cancel record id, optional)

        Two-step UX: upload shows a preview of matched / already-cancelled /
        unmatched rows. The admin then ticks "apply" and resubmits to write.
        Setting cancellation_confirmation_number triggers the model presave
        that flips status to CANCELLED. The cancellation email fires on
        the next email tick.
        """
        from uber.models import RoomAssignment

        upload = params.get('cancellations_csv')
        apply_changes = params.get('apply') == 'true'

        rows = []
        preview = None

        if upload is not None and getattr(upload, 'file', None):
            raw = upload.file.read()
            if isinstance(raw, bytes):
                raw = raw.decode('utf-8', errors='replace')
            reader = csv.DictReader(raw.splitlines())
            preview = {'matched': [], 'already': [], 'unmatched': [], 'applied': 0}

            for row in reader:
                conf = (row.get('confirmation_num') or '').strip()
                cancel_num = (row.get('cancellation_confirmation_number') or '').strip()
                if not conf:
                    continue

                assignment = session.query(RoomAssignment).filter_by(
                    hotel_confirmation_number=conf).first()
                if not assignment:
                    preview['unmatched'].append({
                        'confirmation_num': conf,
                        'cancellation_confirmation_number': cancel_num,
                    })
                    continue

                if assignment.status == c.CANCELLED:
                    preview['already'].append({
                        'assignment': assignment,
                        'cancellation_confirmation_number': cancel_num,
                    })
                    if apply_changes and cancel_num and not assignment.cancellation_confirmation_number:
                        assignment.cancellation_confirmation_number = cancel_num
                        session.add(assignment)
                        preview['applied'] += 1
                    continue

                preview['matched'].append({
                    'assignment': assignment,
                    'cancellation_confirmation_number': cancel_num,
                })

                if apply_changes:
                    assignment.cancellation_confirmation_number = cancel_num or 'imported'
                    # cancellation_flips_status presave on the model takes
                    # care of status = CANCELLED.
                    session.add(assignment)
                    preview['applied'] += 1

            if apply_changes and preview['applied']:
                session.commit()
                message = (
                    f"Applied {preview['applied']} cancellation update(s). "
                    f"{len(preview['unmatched'])} row(s) couldn't be matched."
                )

        return {
            'preview': preview,
            'message': message,
        }

    def import_hotel_confirmations(self, session, message='', **params):
        """Back-import confirmation numbers from the hotel.

        Expected CSV columns:
          - lottery_application_id   (preferred) OR
          - confirmation_num         (matches RoomAssignment.hotel_confirmation_number,
                                      when updating an existing record)
          - new_confirmation_num     (the value to write)

        Same two-step UX as cancellations.
        """
        from uber.models import RoomAssignment

        upload = params.get('confirmations_csv')
        apply_changes = params.get('apply') == 'true'

        preview = None

        if upload is not None and getattr(upload, 'file', None):
            raw = upload.file.read()
            if isinstance(raw, bytes):
                raw = raw.decode('utf-8', errors='replace')
            reader = csv.DictReader(raw.splitlines())
            preview = {'new': [], 'changed': [], 'unchanged': [], 'unmatched': [], 'applied': 0}

            for row in reader:
                app_id = (row.get('lottery_application_id') or '').strip()
                conf = (row.get('confirmation_num') or '').strip()
                new_conf = (row.get('new_confirmation_num') or '').strip()
                if not new_conf:
                    continue

                assignment = None
                if app_id:
                    assignment = session.query(RoomAssignment).filter_by(
                        lottery_application_id=app_id).first()
                if not assignment and conf:
                    assignment = session.query(RoomAssignment).filter_by(
                        hotel_confirmation_number=conf).first()

                if not assignment:
                    preview['unmatched'].append({
                        'lottery_application_id': app_id,
                        'confirmation_num': conf,
                        'new_confirmation_num': new_conf,
                    })
                    continue

                existing = assignment.hotel_confirmation_number or ''
                if existing == new_conf:
                    preview['unchanged'].append({'assignment': assignment})
                    continue

                bucket = 'changed' if existing else 'new'
                preview[bucket].append({
                    'assignment': assignment,
                    'old': existing,
                    'new': new_conf,
                })

                if apply_changes:
                    assignment.hotel_confirmation_number = new_conf
                    session.add(assignment)
                    preview['applied'] += 1
                    _send_confirmation_updated_email(session, assignment)

            if apply_changes and preview['applied']:
                session.commit()
                message = (
                    f"Applied {preview['applied']} confirmation update(s). "
                    f"{len(preview['unmatched'])} row(s) couldn't be matched."
                )

        return {
            'preview': preview,
            'message': message,
        }

    def waitlist_reveals(self, session, message=''):
        """List configured waitlist reveals."""
        reveals = session.query(WaitlistReveal).order_by(
            WaitlistReveal.reveal_at.desc().nullsfirst()).all()
        return {'reveals': reveals, 'message': message}

    def edit_waitlist_reveal(self, session, id=None, message='', **params):
        """Create or edit one WaitlistReveal."""
        reveal = None
        if id and id not in ('None', ''):
            reveal = session.query(WaitlistReveal).get(id)
        if reveal is None:
            reveal = WaitlistReveal()

        if cherrypy.request.method == 'POST':
            reveal.name = params.get('name', '').strip()
            reveal.external_url = params.get('external_url', '').strip()
            reveal.audience_description = params.get('audience_description', '').strip()
            reveal.active = params.get('active') == 'true'
            raw = params.get('reveal_at', '').strip()
            if raw:
                from dateutil import parser as dateparser
                try:
                    reveal.reveal_at = dateparser.parse(raw).replace(tzinfo=c.EVENT_TIMEZONE)
                except (ValueError, TypeError):
                    message = "Could not parse reveal time."
            else:
                reveal.reveal_at = None

            if not message:
                session.add(reveal)
                session.commit()
                raise HTTPRedirect('waitlist_reveals?message={}',
                                   f"Reveal '{reveal.name}' saved.")

        return {'reveal': reveal, 'message': message}

    def send_waitlist_reveal_emails(self, session, id, csrf_token=None):
        """Materialize one WaitlistRevealLink per eligible attendee (anyone
        hotel-lottery-eligible without an active RoomAssignment) and queue
        the reveal email. Idempotent for already-emailed (attendee, reveal)
        pairs - running this again only emails new candidates.
        """
        from uber.utils import check_csrf
        from uber.models import Attendee, RoomAssignment
        import secrets

        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('waitlist_reveals')
        check_csrf(csrf_token)

        reveal = session.query(WaitlistReveal).get(id)
        if not reveal or not reveal.active:
            raise HTTPRedirect('waitlist_reveals?message={}',
                               'Reveal is missing or inactive.')

        eligible_subq = session.query(Attendee.id).outerjoin(
            RoomAssignment,
            sa.and_(
                RoomAssignment.attendee_id == Attendee.id,
                RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            )
        ).filter(
            Attendee.hotel_lottery_eligible == True,  # noqa: E712
            RoomAssignment.id.is_(None),
        ).subquery()

        eligible_ids = [row[0] for row in session.query(eligible_subq.c.id).all()]
        existing_attendee_ids = {
            row[0] for row in session.query(WaitlistRevealLink.attendee_id).filter_by(
                waitlist_reveal_id=reveal.id).all()}

        new_links = []
        for aid in eligible_ids:
            if aid in existing_attendee_ids:
                continue
            link = WaitlistRevealLink(
                waitlist_reveal_id=reveal.id,
                attendee_id=aid,
                token=secrets.token_urlsafe(24),
            )
            session.add(link)
            new_links.append(link)
        session.flush()

        for link in new_links:
            attendee = session.query(Attendee).get(link.attendee_id)
            if not attendee:
                continue
            EmailService.queue_email(
                session, 'hotel_lottery_waitlist_reveal', attendee,
                subject=f"{c.EVENT_NAME_AND_YEAR}: Hotel waitlist link",
                data={'attendee': attendee, 'reveal': reveal, 'link': link})
            link.emailed_at = datetime.now(UTC)
            session.add(link)

        session.commit()
        raise HTTPRedirect(
            'waitlist_reveals?message={}',
            f"Queued {len(new_links)} new waitlist email{'s' if len(new_links) != 1 else ''}.")

    def assign_room(self, session, id=None, message='', **params):
        """Create or edit a RoomAssignment outside the lottery flow.

        Lottery admins assign with reason=MANUAL. Partition owners with
        can_edit_assignments_in their partition assign with reason=PARTITION_GRANT;
        partition_id is locked to their grant's scope. Used by Marketplace,
        Belvedere, Panels, Accessibility to assign exhibitor/panelist rooms.
        """
        from uber.models import Attendee, RoomAssignment
        from uber.lottery_perms import is_lottery_admin, can_edit_assignments_in

        assignment = None
        if id and id not in ('None', ''):
            assignment = session.query(RoomAssignment).get(id)
        if assignment is None:
            assignment = RoomAssignment()

        # Permission gate: hotel-section access already required the user be
        # an admin; the partition-aware checks scope further.
        if cherrypy.request.method == 'POST':
            picked_attendee = params.get('attendee_id', '').strip()
            picked_inventory = params.get('inventory_id', '').strip()
            picked_partition = params.get('partition_id', '').strip() or None

            if not is_lottery_admin() and not can_edit_assignments_in(
                    session, picked_partition):
                message = "You don't have permission to edit assignments in this partition."
            elif not picked_attendee or not picked_inventory:
                message = "Attendee and inventory are required."
            else:
                assignment.attendee_id = picked_attendee
                assignment.inventory_id = picked_inventory
                assignment.partition_id = picked_partition
                if not assignment.assignment_reason or assignment.assignment_reason == c.MANUAL:
                    assignment.assignment_reason = (
                        c.PARTITION_GRANT if picked_partition and not is_lottery_admin()
                        else c.MANUAL)
                assignment.require_cc = params.get('require_cc') == 'true'

                ci = params.get('assigned_check_in_date', '').strip()
                co = params.get('assigned_check_out_date', '').strip()
                assignment.assigned_check_in_date = date.fromisoformat(ci) if ci else None
                assignment.assigned_check_out_date = date.fromisoformat(co) if co else None

                dcd = params.get('deposit_cutoff_date', '').strip()
                assignment.deposit_cutoff_date = date.fromisoformat(dcd) if dcd else None

                assignment.admin_notes = params.get('admin_notes', '').strip()
                session.add(assignment)
                try:
                    session.commit()
                except Exception as e:
                    session.rollback()
                    message = f"Could not save assignment: {e}"
                else:
                    raise HTTPRedirect(
                        'assign_room?id={}&message={}',
                        assignment.id,
                        'Assignment saved.')

        # Scope option lists to what the actor is allowed to see
        partitions = session.query(InventoryPartition).filter_by(active=True).order_by(
            InventoryPartition.name).all()
        if not is_lottery_admin():
            partitions = [
                p for p in partitions
                if can_edit_assignments_in(session, p.id)
            ]

        inventory_rows = session.query(HotelRoomInventory).filter_by(active=True).order_by(
            HotelRoomInventory.hotel_id, HotelRoomInventory.name).all()
        attendees = session.query(Attendee).order_by(
            Attendee.last_name, Attendee.first_name).all()

        return {
            'assignment': assignment,
            'partitions': partitions,
            'inventory_rows': inventory_rows,
            'attendees': attendees,
            'message': message,
        }

    def partition_owners(self, session, partition_id=None, message=''):
        """List PartitionOwner grants, optionally filtered to one partition."""
        from uber.models import AdminAccount

        query = session.query(PartitionOwner).options(
            sa.orm.joinedload(PartitionOwner.admin_account).joinedload(AdminAccount.attendee),
            sa.orm.joinedload(PartitionOwner.partition),
        )
        partition = None
        if partition_id:
            partition = session.query(InventoryPartition).get(partition_id)
            if partition:
                query = query.filter(PartitionOwner.partition_id == partition.id)
        grants = query.all()
        grants.sort(key=lambda g: (g.partition.name if g.partition else '',
                                   g.admin_account.attendee.full_name if g.admin_account and g.admin_account.attendee else ''))
        return {
            'grants': grants,
            'partition': partition,
            'message': message,
        }

    def edit_partition_owner(self, session, id=None, partition_id=None, message='', **params):
        """Create or edit one (admin, partition) grant with its flag bundle."""
        from uber.models import AdminAccount

        grant = None
        if id and id not in ('None', ''):
            grant = session.query(PartitionOwner).get(id)
        if grant is None:
            grant = PartitionOwner()
            if partition_id:
                grant.partition_id = partition_id

        if cherrypy.request.method == 'POST':
            picked_account = params.get('admin_account_id', '').strip()
            # `partition_id` is bound to the function arg (named in the
            # signature for GET pre-fill), so the form's partition_id field
            # lands there, not in **params.
            picked_partition = (partition_id or '').strip()
            if not picked_account or not picked_partition:
                message = "Admin account and partition are both required."
            else:
                is_new = grant.id is None
                grant.admin_account_id = picked_account
                grant.partition_id = picked_partition

                # Three scoped access levels are submitted as
                # `<scope>_level` = none | view | edit. We unpack each
                # into the underlying view/edit flag pair so the
                # invariant "edit implies view" is enforced at the UI
                # layer rather than runtime - view-without-edit is the
                # only intermediate state the dropdown can produce.
                level_scopes = [
                    ('inventory_level',  'can_view_inventory',  'can_edit_inventory'),
                    ('assignments_level', 'can_view_assignments', 'can_edit_assignments'),
                    ('guest_names_level', 'can_view_guest_names', 'can_edit_guest_names'),
                ]
                for level_field, view_flag, edit_flag in level_scopes:
                    level = (params.get(level_field) or '').strip()
                    if level == 'edit':
                        setattr(grant, view_flag, True)
                        setattr(grant, edit_flag, True)
                    elif level == 'view':
                        setattr(grant, view_flag, True)
                        setattr(grant, edit_flag, False)
                    else:
                        # 'none' or missing - both off.
                        setattr(grant, view_flag, False)
                        setattr(grant, edit_flag, False)

                grant.can_send_emails = params.get('can_send_emails') == 'true'
                session.add(grant)
                session.flush()
                record_partition_audit(
                    session, grant.partition_id,
                    action='partition_owner.granted' if is_new else 'partition_owner.updated',
                    description=("Granted partition access" if is_new
                                 else "Updated partition access capabilities"),
                    target_type='partition_owner', target_id=grant.id)
                try:
                    session.commit()
                except Exception as e:
                    session.rollback()
                    message = f"Could not save grant: {e}"
                else:
                    if params.get('return_to') == 'edit_partition':
                        raise HTTPRedirect(
                            'edit_partition?id={}&message={}',
                            grant.partition_id, 'Grant saved.')
                    raise HTTPRedirect('partition_owners?partition_id={}&message={}',
                                       grant.partition_id, 'Grant saved.')

        admin_accounts = session.query(AdminAccount).all()
        admin_accounts.sort(key=lambda a: a.attendee.full_name if a.attendee else '')
        partitions = session.query(InventoryPartition).filter_by(active=True).order_by(
            InventoryPartition.name).all()
        return {
            'grant': grant,
            'admin_accounts': admin_accounts,
            'partitions': partitions,
            'message': message,
        }

    def delete_partition_owner(self, session, id, csrf_token=None, return_to=''):
        from uber.utils import check_csrf
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('partition_owners')
        check_csrf(csrf_token)
        grant = session.query(PartitionOwner).get(id)
        if grant:
            partition_id = grant.partition_id
            record_partition_audit(
                session, partition_id,
                action='partition_owner.revoked',
                description="Revoked partition access",
                target_type='partition_owner', target_id=grant.id)
            session.delete(grant)
            session.commit()
            if return_to == 'edit_partition':
                raise HTTPRedirect('edit_partition?id={}&message={}',
                                   partition_id, 'Grant revoked.')
            raise HTTPRedirect('partition_owners?partition_id={}&message={}',
                               partition_id, 'Grant revoked.')
        raise HTTPRedirect('partition_owners?message={}', 'Grant not found.')

    def edit_partition(self, session, id=None, message='', **params):
        if id and id not in ('None', ''):
            partition = session.query(InventoryPartition).get(id)
        else:
            partition = None
        if not partition:
            partition = InventoryPartition()

        inventory_blocks = session.query(HotelRoomInventory).filter_by(active=True).order_by(
            HotelRoomInventory.hotel_id, HotelRoomInventory.name).all()
        existing_blocks = {str(pb.inventory_id): pb.quantity for pb in partition.blocks} if partition.id else {}

        if cherrypy.request.method == 'POST':
            partition.name = params.get('name', '').strip()
            partition.description = params.get('description', '').strip()
            partition.active = params.get('active') == 'true'
            session.add(partition)
            session.flush()

            existing_pb = {str(pb.inventory_id): pb for pb in partition.blocks}
            change_lines = []
            for inv in inventory_blocks:
                qty_str = params.get(f'block_qty_{inv.id}', '')
                inv_label = inv.name or 'inventory'
                if qty_str != '' and int(qty_str) > 0:
                    qty = int(qty_str)
                    if str(inv.id) in existing_pb:
                        old_qty = existing_pb[str(inv.id)].quantity
                        if old_qty != qty:
                            existing_pb[str(inv.id)].quantity = qty
                            change_lines.append(f"{inv_label}: {old_qty} -> {qty}")
                    else:
                        pb = InventoryPartitionBlock(
                            partition_id=partition.id, inventory_id=str(inv.id), quantity=qty)
                        session.add(pb)
                        change_lines.append(f"{inv_label}: added with {qty}")
                elif str(inv.id) in existing_pb:
                    old_qty = existing_pb[str(inv.id)].quantity
                    session.delete(existing_pb[str(inv.id)])
                    change_lines.append(f"{inv_label}: removed (was {old_qty})")

            session.commit()

            # Notify partition owners of block edits.
            if change_lines and partition.id:
                _notify_partition_owners_of_inventory_change(
                    session, partition,
                    change_description='; '.join(change_lines))

            raise HTTPRedirect('manage_partitions?message={}', f"Partition '{partition.name}' saved.")

        # Existing partition-owner grants + admin pool for the inline manager
        # at the bottom of the page.
        from uber.models import AdminAccount
        if partition.id:
            grants = (session.query(PartitionOwner)
                      .filter_by(partition_id=partition.id)
                      .options(
                          sa.orm.joinedload(PartitionOwner.admin_account)
                          .joinedload(AdminAccount.attendee))
                      .all())
            grants.sort(key=lambda g: (
                g.admin_account.attendee.full_name
                if g.admin_account and g.admin_account.attendee else ''))
        else:
            grants = []

        granted_account_ids = {g.admin_account_id for g in grants}
        ungranted_admins = [a for a in session.query(AdminAccount).all()
                            if a.id not in granted_account_ids and a.attendee]
        ungranted_admins.sort(key=lambda a: a.attendee.full_name if a.attendee else '')

        # Current usage per block within THIS partition, so admins can
        # see how much is already committed before changing the
        # allocation. `usage_by_block` is the PEAK per-night count of
        # live (ASSIGNED/SECURED) RoomAssignments tagged with this
        # partition for each inventory block - that's the number the
        # per-night `quantity` allocation has to cover, so reducing
        # below it would over-commit the busiest night.
        # `rooms_by_block` is the total distinct live rooms (handy
        # context, not the binding constraint).
        usage_by_block = {}
        rooms_by_block = {}
        if partition.id:
            live_ras = (session.query(RoomAssignment)
                        .filter(RoomAssignment.partition_id == partition.id,
                                RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
                                RoomAssignment.inventory_id.isnot(None))
                        .all())
            per_block_night = {}
            for ra in live_ras:
                bid = str(ra.inventory_id)
                rooms_by_block[bid] = rooms_by_block.get(bid, 0) + 1
                if ra.assigned_check_in_date and ra.assigned_check_out_date:
                    nights = per_block_night.setdefault(bid, {})
                    d = ra.assigned_check_in_date
                    while d < ra.assigned_check_out_date:
                        nights[d] = nights.get(d, 0) + 1
                        d += timedelta(days=1)
            for bid, total in rooms_by_block.items():
                nights = per_block_night.get(bid)
                # Peak per-night usage; fall back to the room count for
                # rows with no dates so they still register as in-use.
                usage_by_block[bid] = max(nights.values()) if nights else total

        return {
            'partition': partition,
            'inventory_blocks': inventory_blocks,
            'existing_blocks': existing_blocks,
            'usage_by_block': usage_by_block,
            'rooms_by_block': rooms_by_block,
            'grants': grants,
            'ungranted_admins': ungranted_admins,
            'message': message,
        }

    @ajax
    def reduce_awards(self, session, inventory_id, night_date, target_count):
        try:
            target_count = int(target_count)
            night = date.fromisoformat(night_date)
        except (ValueError, TypeError):
            return {"error": "Invalid target count or date."}

        # Reduce by ejecting RoomAssignment rows for the given block + night
        # at random. The owning LotteryApplication is rolled back to
        # COMPLETE once all its assignments are gone (RoomAssignment's
        # after_delete listener handles the status flip).
        candidate_ras = session.query(RoomAssignment).filter(
            RoomAssignment.inventory_id == inventory_id,
            RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]),
            RoomAssignment.assigned_check_in_date <= night,
            RoomAssignment.assigned_check_out_date > night,
        ).all()

        current_count = len(candidate_ras)
        if target_count >= current_count:
            return {"success": True, "message": f"No reduction needed ({current_count} currently assigned)."}

        # Prefer ejecting assignments whose source app has no group members.
        def _has_group(ra):
            app = ra.lottery_application
            return bool(app and app.group_members)

        ejectable = [ra for ra in candidate_ras if not _has_group(ra)]
        if len(ejectable) < current_count - target_count:
            ejectable = candidate_ras

        to_eject = random.sample(ejectable, min(len(ejectable), current_count - target_count))

        impacted_apps = {ra.lottery_application_id for ra in to_eject
                         if ra.lottery_application_id}
        for ra in to_eject:
            session.delete(ra)
        session.commit()

        # Clear partition + lottery_run linkage on apps whose last assignment
        # was just removed (the after_delete listener flips status back to
        # COMPLETE; we just clean the run linkage here).
        for app_id in impacted_apps:
            app = session.query(LotteryApplication).get(app_id)
            if not app:
                continue
            remaining = session.query(RoomAssignment).filter_by(
                lottery_application_id=app.id).count()
            if remaining == 0:
                app.partition_id = None
                app.lottery_run_id = None
                session.add(app)
        session.commit()

        return {"success": True, "message": f"Ejected {len(to_eject)} entries. {target_count} remain for {night_date}."}

    @ajax
    def unlock_application(self, session, id):
        app = session.lottery_application(id)
        app.export_locked = False
        session.add(app)
        session.commit()
        return {"success": True}

    # Wrappers around the underlying RoomAssignment CRUD that live in
    # partition_admin. Each redirects back to the application edit form
    # so the lottery admin stays on the same screen they came from.
    # Authority gate is HAS_HOTEL_LOTTERY_ADMIN_ACCESS (the @all_renderable
    # at the top of this class). The partition-scoped gating in the
    # partition_admin handlers ALSO short-circuits to True for lottery
    # admins via `can_edit_assignments_in`, so this just bypasses the
    # redirect dance.

    def add_room_assignment(self, session, application_id, inventory_id='',
                            partition_id='', csrf_token=None, **params):
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('form?id={}', application_id)
        check_csrf(csrf_token)
        app = session.lottery_application(application_id)
        if not inventory_id:
            raise HTTPRedirect('form?id={}&message={}', application_id,
                               'Inventory is required to add a room.')

        ra = RoomAssignment(
            attendee_id=app.attendee_id,
            lottery_application_id=app.id,
            inventory_id=inventory_id,
            partition_id=partition_id or None,
            assignment_reason=c.MANUAL,
            status=c.ASSIGNED,
            require_cc=params.get('require_cc') == 'true',
        )
        ci = params.get('assigned_check_in_date', '').strip()
        co = params.get('assigned_check_out_date', '').strip()
        if ci:
            try:
                ra.assigned_check_in_date = date.fromisoformat(ci)
            except ValueError:
                pass
        if co:
            try:
                ra.assigned_check_out_date = date.fromisoformat(co)
            except ValueError:
                pass
        session.add(ra)
        session.flush()
        if partition_id:
            record_partition_audit(
                session, partition_id,
                action='assignment.created',
                description=f"Manually added room to attendee {app.attendee_id}",
                target_type='assignment', target_id=ra.id)
        session.commit()
        raise HTTPRedirect('form?id={}&message={}', application_id, 'Room added.')

    def update_room_assignment(self, session, application_id, assignment_id,
                               csrf_token=None, **params):
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('form?id={}', application_id)
        check_csrf(csrf_token)
        ra = session.query(RoomAssignment).get(assignment_id)
        if not ra:
            raise HTTPRedirect('form?id={}&message={}', application_id,
                               'Assignment not found.')

        changes = []
        new_inv = params.get('inventory_id', '').strip()
        if new_inv and new_inv != ra.inventory_id:
            changes.append('inventory'); ra.inventory_id = new_inv
        new_part = params.get('partition_id', '').strip() or None
        if new_part != ra.partition_id:
            changes.append('partition'); ra.partition_id = new_part
        new_require_cc = params.get('require_cc') == 'true'
        if new_require_cc != ra.require_cc:
            changes.append('billing'); ra.require_cc = new_require_cc

        ci = params.get('assigned_check_in_date', '').strip()
        co = params.get('assigned_check_out_date', '').strip()
        try:
            new_ci = date.fromisoformat(ci) if ci else None
        except ValueError:
            new_ci = ra.assigned_check_in_date
        try:
            new_co = date.fromisoformat(co) if co else None
        except ValueError:
            new_co = ra.assigned_check_out_date
        if new_ci != ra.assigned_check_in_date:
            changes.append('check-in'); ra.assigned_check_in_date = new_ci
        if new_co != ra.assigned_check_out_date:
            changes.append('check-out'); ra.assigned_check_out_date = new_co

        if changes:
            session.add(ra)
            if ra.partition_id:
                record_partition_audit(
                    session, ra.partition_id,
                    action='assignment.updated',
                    description=f"Lottery admin updated {', '.join(changes)}",
                    target_type='assignment', target_id=ra.id)
            session.commit()
            msg = f"Updated {', '.join(changes)}."
        else:
            msg = 'No changes.'
        raise HTTPRedirect('form?id={}&message={}', application_id, msg)

    def delete_room_assignment(self, session, assignment_id,
                               application_id='', csrf_token=None):
        if cherrypy.request.method != 'POST':
            if application_id:
                raise HTTPRedirect('form?id={}', application_id)
            raise HTTPRedirect('rooms')
        check_csrf(csrf_token)
        ra = session.query(RoomAssignment).get(assignment_id)
        if not ra:
            if application_id:
                raise HTTPRedirect('form?id={}&message={}', application_id,
                                   'Assignment not found.')
            raise HTTPRedirect('rooms?message={}', 'Assignment not found.')

        # Cascade-delete connector children - they only exist as long as
        # their parent does.
        children = (session.query(RoomAssignment)
                    .filter_by(parent_assignment_id=ra.id).all())
        for child in children:
            if ra.partition_id:
                record_partition_audit(
                    session, ra.partition_id,
                    action='assignment.deleted',
                    description="Connector cascade",
                    target_type='assignment', target_id=child.id)
            session.delete(child)
        if ra.partition_id:
            record_partition_audit(
                session, ra.partition_id,
                action='assignment.deleted',
                description="Lottery admin removed assignment",
                target_type='assignment', target_id=ra.id)
        session.delete(ra)
        session.commit()
        if application_id:
            raise HTTPRedirect('form?id={}&message={}', application_id,
                               'Room removed.')
        raise HTTPRedirect('rooms?message={}', 'Room removed.')

    # Works for both lottery-tied and non-lottery (manual / partition
    # grant) RoomAssignments. The per-application form has its own modal
    # editor for rooms in the context of an application; this page is
    # the canonical "edit this specific room" surface, reachable from
    # the cross-section Rooms list.

    def edit_room_assignment(self, session, id, message=''):
        ra = session.query(RoomAssignment).get(id)
        if not ra:
            raise HTTPRedirect('rooms?message={}', 'Assignment not found.')

        partitions = (session.query(InventoryPartition)
                      .filter_by(active=True)
                      .order_by(InventoryPartition.name).all())
        inventory_blocks = (session.query(HotelRoomInventory)
                            .filter_by(active=True)
                            .order_by(HotelRoomInventory.hotel_id,
                                      HotelRoomInventory.name).all())
        # {inventory_id: [partition_id, ...]} for the partition-filter JS.
        inventory_partitions_map = {}
        for pb in session.query(InventoryPartitionBlock).all():
            inventory_partitions_map.setdefault(
                str(pb.inventory_id), []).append(str(pb.partition_id))

        return {
            'assignment': ra,
            'partitions': partitions,
            'inventory_blocks': inventory_blocks,
            'inventory_partitions_map': inventory_partitions_map,
            'message': message,
        }

    def save_room_assignment(self, session, assignment_id,
                             csrf_token=None, **params):
        """Standalone-page version of update_room_assignment that also
        accepts the per-room fields not exposed in the form modal
        (status, hotel confirmation, cancellation, deposit cutoff,
        special requests). Redirects back to the standalone edit page
        on success."""
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect('edit_room_assignment?id={}', assignment_id)
        check_csrf(csrf_token)
        ra = session.query(RoomAssignment).get(assignment_id)
        if not ra:
            raise HTTPRedirect('rooms?message={}', 'Assignment not found.')

        changes = []

        new_inv = params.get('inventory_id', '').strip()
        if new_inv and new_inv != ra.inventory_id:
            changes.append('inventory'); ra.inventory_id = new_inv
        new_part = params.get('partition_id', '').strip() or None
        if new_part != ra.partition_id:
            changes.append('partition'); ra.partition_id = new_part
        new_require_cc = params.get('require_cc') == 'true'
        if new_require_cc != ra.require_cc:
            changes.append('billing'); ra.require_cc = new_require_cc

        def _parse_date(name, current):
            raw = (params.get(name, '') or '').strip()
            if not raw:
                return None
            try:
                return date.fromisoformat(raw)
            except ValueError:
                return current

        new_ci = _parse_date('assigned_check_in_date', ra.assigned_check_in_date)
        if new_ci != ra.assigned_check_in_date:
            changes.append('check-in'); ra.assigned_check_in_date = new_ci
        new_co = _parse_date('assigned_check_out_date', ra.assigned_check_out_date)
        if new_co != ra.assigned_check_out_date:
            changes.append('check-out'); ra.assigned_check_out_date = new_co
        new_cutoff = _parse_date('deposit_cutoff_date', ra.deposit_cutoff_date)
        if new_cutoff != ra.deposit_cutoff_date:
            changes.append('deposit cutoff'); ra.deposit_cutoff_date = new_cutoff

        raw_status = params.get('status', '').strip()
        if raw_status:
            try:
                new_status = int(raw_status)
                if new_status != ra.status:
                    changes.append('status'); ra.status = new_status
            except ValueError:
                pass

        for field in ('hotel_confirmation_number',
                      'cancellation_confirmation_number',
                      'special_requests'):
            raw = (params.get(field, '') or '').strip()
            if raw != (getattr(ra, field) or ''):
                changes.append(field.replace('_', ' '))
                setattr(ra, field, raw or None)

        if changes:
            session.add(ra)
            if ra.partition_id:
                record_partition_audit(
                    session, ra.partition_id,
                    action='assignment.updated',
                    description=f"Edit page updated {', '.join(changes)}",
                    target_type='assignment', target_id=ra.id)
            session.commit()
            msg = f"Updated {', '.join(changes)}."
        else:
            msg = 'No changes.'
        raise HTTPRedirect('edit_room_assignment?id={}&message={}',
                           assignment_id, msg)

    @ajax
    def bulk_unlock(self, session, ids):
        id_list = [x.strip() for x in ids.split(',') if x.strip()]
        count = 0
        for app_id in id_list:
            app = session.query(LotteryApplication).get(app_id)
            if app and app.export_locked:
                app.export_locked = False
                session.add(app)
                count += 1
        session.commit()
        return {"success": True, "count": count}

    # Cross-application view: every RoomAssignment in the system,
    # paginated. Useful when an admin knows the room (hotel + date) but
    # not the attendee/application, or wants to scan the whole population
    # for sanity-check. Filters are deliberately minimal - for fine-grained
    # searching the application search box on the Applications page is
    # still the right tool.

    def rooms(self, session, message='', page='1', page_size='50',
              status='live', hotel_id='', partition_id='', search=''):
        try:
            page_num = max(1, int(page))
        except (TypeError, ValueError):
            page_num = 1
        try:
            ps = max(10, min(500, int(page_size)))
        except (TypeError, ValueError):
            ps = 50

        q = session.query(RoomAssignment)

        # `live` (default) = ASSIGNED + SECURED. `all` = no status filter.
        # Anything else = exact status int from the model's status enum.
        if status == 'live':
            q = q.filter(RoomAssignment.status.in_([c.ASSIGNED, c.SECURED]))
        elif status and status != 'all':
            try:
                q = q.filter(RoomAssignment.status == int(status))
            except (TypeError, ValueError):
                pass

        if hotel_id:
            inv_ids = [str(inv.id) for inv in
                       session.query(HotelRoomInventory)
                       .filter_by(hotel_id=hotel_id).all()]
            q = q.filter(RoomAssignment.inventory_id.in_(inv_ids))
        if partition_id:
            if partition_id == 'none':
                q = q.filter(RoomAssignment.partition_id.is_(None))
            else:
                q = q.filter(RoomAssignment.partition_id == partition_id)

        # Search hits the application's confirmation #, attendee email/
        # name, and the hotel confirmation #. We can't easily search the
        # attendee name across the RA join in raw SQLAlchemy without
        # joining, so do the join when needed.
        search_term = (search or '').strip()
        if search_term:
            from uber.models import Attendee as _Attendee
            like = f'%{search_term}%'
            q = (q.join(LotteryApplication,
                        LotteryApplication.id == RoomAssignment.lottery_application_id,
                        isouter=True)
                  .join(_Attendee,
                        _Attendee.id == RoomAssignment.attendee_id,
                        isouter=True)
                  .filter(or_(
                      LotteryApplication.confirmation_num.ilike(like),
                      RoomAssignment.hotel_confirmation_number.ilike(like),
                      _Attendee.email.ilike(like),
                      _Attendee.first_name.ilike(like),
                      _Attendee.last_name.ilike(like),
                  )))

        total = q.count()
        page_count = max(1, (total + ps - 1) // ps)
        if page_num > page_count:
            page_num = page_count

        # Sort: hotel name, then check-in date (nulls first so missing
        # dates pop to the top and get noticed), then created order.
        assignments = (q
                       .order_by(RoomAssignment.assigned_check_in_date.asc().nullsfirst(),
                                 RoomAssignment.created.asc())
                       .offset((page_num - 1) * ps)
                       .limit(ps)
                       .all())

        hotels = (session.query(LotteryHotel)
                  .filter_by(active=True)
                  .order_by(LotteryHotel.name).all())
        partitions = (session.query(InventoryPartition)
                      .filter_by(active=True)
                      .order_by(InventoryPartition.name).all())

        return {
            'message': message,
            'assignments': assignments,
            'page': page_num,
            'page_size': ps,
            'total': total,
            'page_count': page_count,
            'status': status,
            'hotel_id': hotel_id,
            'partition_id': partition_id,
            'search': search_term,
            'hotels': hotels,
            'partitions': partitions,
            'status_opts': c.HOTEL_ASSIGNMENT_STATUS_OPTS,
        }

    @ajax
    def process_waitlist(self, session, inventory_id='', night_date=''):
        inv_id = inventory_id if inventory_id else None
        nd = date.fromisoformat(night_date) if night_date else None
        result = _fulfill_waitlist(session, inventory_id=inv_id, night_date=nd)
        return result

    def waitlist(self, session, message='', page='1', search_text=''):
        """Admin Waitlist dashboard.

        Two views on the same data:

          1. **Per-block demand** - for every inventory block with at
             least one waitlisted RoomAssignment, the per-night
             waitlist count (how many distinct rooms are queued for
             that specific night). Helps the admin see which nights
             are oversubscribed.

          2. **Per-room queue** - every waitlisted RoomAssignment in
             FIFO order, with the inventory block, attendee, requested
             vs confirmed range, and waitlist start timestamp. Each
             row gets an Accept button that calls `accept_waitlist`
             to immediately extend `assigned_*` to cover the
             `waitlisted_*` range for that single row (admin override,
             no capacity check - the admin is explicitly choosing to
             accept this person off the queue).

        Process Waitlist (the cron-style fulfillment that respects
        capacity) also lives here now; the old button on the inventory
        overview was redundant once this page existed.

        **Search** (`search_text=`) - case-insensitive substring match
        against the attendee's first/last name, email, lottery
        confirmation number, hotel name, inventory block name, and
        room/suite type name. The block-demand histogram re-derives
        from the filtered set so both views stay consistent.

        **Pagination** (`page=`) - the FIFO queue paginates at 100 rows
        per page (matches the rest of the admin section's convention,
        e.g. `hotel_lottery_admin/index`). The histogram itself is
        small (one row per inventory block, capped by inventory size)
        so it's never paginated.
        """
        from uber.models.hotel import RoomAssignment

        PER_PAGE = 100

        search_text = (search_text or '').strip()

        # Base query - every waitlisted row, FIFO-ordered. We eagerly
        # load the relationships the search filter and the template
        # both need (attendee, inventory.hotel, inventory.room_type,
        # inventory.suite_type, lottery_application) in one shot so
        # the row-loop below doesn't trigger N+1 lookups.
        waitlist_filter = sa.or_(
            RoomAssignment.waitlisted_check_in_date.isnot(None),
            RoomAssignment.waitlisted_check_out_date.isnot(None))
        q = (session.query(RoomAssignment)
             .options(
                 joinedload(RoomAssignment.attendee),
                 joinedload(RoomAssignment.inventory)
                     .joinedload(HotelRoomInventory.hotel),
                 joinedload(RoomAssignment.lottery_application))
             .filter(waitlist_filter)
             .order_by(
                 RoomAssignment.waitlist_started_at.asc().nullsfirst(),
                 RoomAssignment.created.asc()))

        if search_text:
            # Push the search down to SQL via joins + ILIKE OR-chain so
            # the planner can use indexes on attendee/email and we
            # don't pull the entire waitlisted-row population into
            # Python just to throw most of it away.
            term = f'%{search_text}%'
            inv_alias = HotelRoomInventory
            from sqlalchemy.orm import aliased
            # Room and suite type aliases - same table, different
            # column on inventory points at them.
            room_type_a = aliased(LotteryRoomType)
            suite_type_a = aliased(LotteryRoomType)

            q = (q.outerjoin(Attendee,
                             RoomAssignment.attendee_id == Attendee.id)
                  .outerjoin(LotteryApplication,
                             RoomAssignment.lottery_application_id == LotteryApplication.id)
                  .outerjoin(inv_alias,
                             RoomAssignment.inventory_id == inv_alias.id)
                  .outerjoin(LotteryHotel,
                             inv_alias.hotel_id == LotteryHotel.id)
                  .outerjoin(room_type_a,
                             inv_alias.room_type_id == room_type_a.id)
                  .outerjoin(suite_type_a,
                             inv_alias.suite_type_id == suite_type_a.id)
                  .filter(sa.or_(
                      Attendee.first_name.ilike(term),
                      Attendee.last_name.ilike(term),
                      (Attendee.first_name + ' ' + Attendee.last_name).ilike(term),
                      Attendee.email.ilike(term),
                      LotteryApplication.confirmation_num.ilike(term),
                      LotteryHotel.name.ilike(term),
                      inv_alias.name.ilike(term),
                      room_type_a.name.ilike(term),
                      suite_type_a.name.ilike(term),
                  ))
                  # Duplicates appear if multiple joined rows match -
                  # one assignment with two LIKE-matching fields would
                  # come back twice. distinct() collapses them; the
                  # ordering above survives because both order columns
                  # are on RoomAssignment itself.
                  .distinct())

        filtered = q.all()
        total_count = len(filtered)

        # Unfiltered population - used by the template to distinguish
        # "the waitlist is genuinely empty" (show the friendly empty
        # state, hide the search form) from "your search has zero
        # matches" (keep the search form so the admin can revise or
        # clear it). Cheap separate count query when search is active;
        # we already have it when search is empty.
        if search_text:
            waitlist_size = session.query(RoomAssignment.id).filter(
                waitlist_filter).count()
        else:
            waitlist_size = total_count

        # Per-block per-night demand histogram. Rebuilt from the
        # filtered set so the histogram and the FIFO table represent
        # the same population - search "hampton" and you see only the
        # Hampton blocks light up.
        demand_by_block = defaultdict(lambda: defaultdict(list))
        for ra in filtered:
            block_id = str(ra.inventory_id) if ra.inventory_id else None
            if not block_id:
                continue
            wl_ci = ra.waitlisted_check_in_date or ra.assigned_check_in_date
            wl_co = ra.waitlisted_check_out_date or ra.assigned_check_out_date
            if wl_ci and ra.assigned_check_in_date and wl_ci < ra.assigned_check_in_date:
                d = wl_ci
                while d < ra.assigned_check_in_date:
                    demand_by_block[block_id][d].append(ra)
                    d += timedelta(days=1)
            if wl_co and ra.assigned_check_out_date and wl_co > ra.assigned_check_out_date:
                d = ra.assigned_check_out_date
                while d < wl_co:
                    demand_by_block[block_id][d].append(ra)
                    d += timedelta(days=1)

        block_ids = list(demand_by_block.keys())
        inventory_by_id = {}
        if block_ids:
            for inv in session.query(HotelRoomInventory).filter(
                    HotelRoomInventory.id.in_(block_ids)).all():
                inventory_by_id[str(inv.id)] = inv

        block_rows = []
        for block_id in block_ids:
            nights = demand_by_block[block_id]
            block_rows.append({
                'inventory': inventory_by_id.get(block_id),
                'inventory_id': block_id,
                'nights': sorted(((n, len(ras)) for n, ras in nights.items()),
                                 key=lambda p: p[0]),
                'total_demand': sum(len(ras) for ras in nights.values()),
            })
        block_rows.sort(key=lambda r: (
            r['inventory'].hotel.name if r['inventory'] and r['inventory'].hotel else '',
            r['inventory'].name if r['inventory'] else ''))

        # Pagination. `get_page` is the same helper the rest of the
        # admin uses (`uber.utils.get_page`), but its 100-per-page
        # default is baked in, so we mirror it here for the page
        # count math.
        try:
            page = max(1, int(page or 1))
        except (TypeError, ValueError):
            page = 1
        total_pages = max(1, math.ceil(total_count / PER_PAGE)) if total_count else 1
        if page > total_pages:
            page = total_pages
        page_slice = filtered[(page - 1) * PER_PAGE: page * PER_PAGE]
        pages = range(1, total_pages + 1)

        return {
            'message': message,
            'block_rows': block_rows,
            'queue': page_slice,
            'total_count': total_count,
            'waitlist_size': waitlist_size,
            'page': page,
            'pages': pages,
            'per_page': PER_PAGE,
            'search_text': search_text,
        }

    def export_waitlist_xlsx(self, session):
        """One-XLSX-per-call export of the current waitlist demand,
        with one worksheet per hotel that has any waitlisted rooms.

        Sheet layout (within each hotel):

            row 1:  Hotel name (merged across the night columns)
            row 2:  blank spacer
            row 3:  header - ["Room type", <night 1>, <night 2>, ..., "Total"]
            row 4+: one row per room type at this hotel, with the count of
                    distinct waitlisted RoomAssignment rows demanding each
                    (type, night) pair.
            last:   "Total" row summing each column.

        Built manually (no `@xlsx_file` decorator) because that helper
        only hands out a single worksheet, and we need one sheet per
        hotel. We still match the decorator's response shape: same
        Content-Type, a filename derived from the handler name plus a
        timestamp, and we participate in `track_report` so admin
        exports show up in the usage log.
        """
        from io import BytesIO
        from datetime import datetime as _dt
        import xlsxwriter
        from uber.models.hotel import RoomAssignment

        # Gather every waitlisted assignment.
        waitlisted = (session.query(RoomAssignment)
                      .filter(sa.or_(
                          RoomAssignment.waitlisted_check_in_date.isnot(None),
                          RoomAssignment.waitlisted_check_out_date.isnot(None)))
                      .all())

        # Build a {hotel_id: {(type_name, type_is_suite): {night: count}}}
        # nested histogram, plus parallel lookup tables for hotel display
        # names. Multiple inventory blocks of the same room type at the same
        # hotel collapse together so the report shows room type by night
        # within a hotel: two Standard King blocks at the same hotel roll
        # into one "Standard King" row.
        from collections import defaultdict as _dd
        per_hotel = _dd(lambda: _dd(lambda: _dd(int)))  # hotel_id -> type_label -> night -> count
        hotel_name_by_id = {}
        nights_by_hotel = _dd(set)
        type_order_by_hotel = _dd(list)  # preserve first-seen order per hotel
        type_seen_by_hotel = _dd(set)

        for ra in waitlisted:
            inv = ra.inventory
            if not inv or not inv.hotel:
                continue
            hotel = inv.hotel
            hotel_id = str(hotel.id)
            hotel_name_by_id[hotel_id] = hotel.name or '(unnamed hotel)'

            # Resolve a stable label for this room type. Suite types
            # and standard types both live in `LotteryRoomType`; an
            # inventory block points at one via `suite_type_id` or
            # `room_type_id` depending on `is_suite`.
            if inv.is_suite:
                rt = inv.suite_type
                label = (rt.name if rt else inv.name) + ' (suite)'
            else:
                rt = inv.room_type
                label = rt.name if rt else inv.name
            label = label or '(unnamed type)'

            if label not in type_seen_by_hotel[hotel_id]:
                type_seen_by_hotel[hotel_id].add(label)
                type_order_by_hotel[hotel_id].append(label)

            # Walk this assignment's waitlist gap and tick each
            # (type, night) cell. Front gap = nights before
            # assigned_check_in_date; back gap = nights at or after
            # assigned_check_out_date.
            wl_ci = ra.waitlisted_check_in_date or ra.assigned_check_in_date
            wl_co = ra.waitlisted_check_out_date or ra.assigned_check_out_date

            if wl_ci and ra.assigned_check_in_date and wl_ci < ra.assigned_check_in_date:
                d = wl_ci
                while d < ra.assigned_check_in_date:
                    per_hotel[hotel_id][label][d] += 1
                    nights_by_hotel[hotel_id].add(d)
                    d += timedelta(days=1)
            if wl_co and ra.assigned_check_out_date and wl_co > ra.assigned_check_out_date:
                d = ra.assigned_check_out_date
                while d < wl_co:
                    per_hotel[hotel_id][label][d] += 1
                    nights_by_hotel[hotel_id].add(d)
                    d += timedelta(days=1)

        # Build the workbook. Sheet name has a 31-char cap and can't
        # contain :\/?*[]; truncate and substitute.
        def _safe_sheet_name(name, taken):
            cleaned = ''
            for ch in (name or 'Waitlist'):
                cleaned += ' ' if ch in ':\\/?*[]' else ch
            cleaned = cleaned.strip()[:31] or 'Waitlist'
            # Disambiguate collisions (rare - two hotels with names
            # truncating to the same 31 chars).
            base = cleaned
            n = 2
            while cleaned in taken:
                suffix = f' ({n})'
                cleaned = base[:31 - len(suffix)] + suffix
                n += 1
            taken.add(cleaned)
            return cleaned

        rawoutput = BytesIO()
        with xlsxwriter.Workbook(rawoutput, {'in_memory': True}) as workbook:
            title_fmt = workbook.add_format(
                {'bold': True, 'font_size': 14, 'align': 'left'})
            header_fmt = workbook.add_format(
                {'bold': True, 'bg_color': '#EFEFEF', 'border': 1})
            total_fmt = workbook.add_format(
                {'bold': True, 'top': 1})
            date_fmt = workbook.add_format(
                {'bold': True, 'bg_color': '#EFEFEF', 'border': 1,
                 'align': 'center', 'num_format': 'ddd m/d'})
            cell_fmt = workbook.add_format({'align': 'center'})

            if not per_hotel:
                # Always produce at least one sheet so the file is
                # openable - an empty workbook would be confusing
                # output for an admin who clicked Export.
                ws = workbook.add_worksheet('Waitlist')
                ws.write(0, 0, 'No rooms are currently on the waitlist.',
                         title_fmt)
            else:
                taken_sheet_names = set()
                # Stable sheet order: alphabetical by hotel name so the
                # tab strip across the bottom of Excel is predictable.
                hotel_ids_sorted = sorted(
                    per_hotel.keys(),
                    key=lambda hid: hotel_name_by_id.get(hid, ''))

                for hotel_id in hotel_ids_sorted:
                    hotel_name = hotel_name_by_id[hotel_id]
                    sheet_name = _safe_sheet_name(
                        hotel_name, taken_sheet_names)
                    ws = workbook.add_worksheet(sheet_name)

                    nights_sorted = sorted(nights_by_hotel[hotel_id])
                    types_sorted = type_order_by_hotel[hotel_id]
                    n_cols = 1 + len(nights_sorted) + 1  # type + nights + total

                    # Row 0: hotel title spanning the night columns.
                    ws.merge_range(0, 0, 0, n_cols - 1,
                                   f'{hotel_name} - Waitlist demand',
                                   title_fmt)
                    # Row 1: spacer (left blank).

                    # Row 2: header.
                    ws.write(2, 0, 'Room type', header_fmt)
                    for i, night in enumerate(nights_sorted):
                        # Write as a real date so admins can re-sort
                        # or compute on it; format string handles
                        # display.
                        ws.write_datetime(
                            2, 1 + i, _dt.combine(night, _dt.min.time()),
                            date_fmt)
                    ws.write(2, 1 + len(nights_sorted), 'Total', header_fmt)

                    # Rows 3..N: one row per room type.
                    col_totals = [0] * len(nights_sorted)
                    for r_offset, label in enumerate(types_sorted):
                        row_idx = 3 + r_offset
                        ws.write(row_idx, 0, label)
                        row_total = 0
                        for c_offset, night in enumerate(nights_sorted):
                            count = per_hotel[hotel_id][label].get(night, 0)
                            if count:
                                ws.write_number(
                                    row_idx, 1 + c_offset, count, cell_fmt)
                                row_total += count
                                col_totals[c_offset] += count
                            else:
                                # Leave blank rather than writing 0
                                # so the sparse cells visually
                                # disappear and the populated ones
                                # pop.
                                ws.write_blank(
                                    row_idx, 1 + c_offset, None, cell_fmt)
                        ws.write_number(
                            row_idx, 1 + len(nights_sorted), row_total,
                            total_fmt)

                    # Final row: per-night totals.
                    total_row = 3 + len(types_sorted)
                    ws.write(total_row, 0, 'Total', total_fmt)
                    for c_offset, total in enumerate(col_totals):
                        ws.write_number(
                            total_row, 1 + c_offset, total, total_fmt)
                    ws.write_number(
                        total_row, 1 + len(nights_sorted), sum(col_totals),
                        total_fmt)

                    # Modest column widths so the sheet is readable
                    # without manual resizing.
                    ws.set_column(0, 0, 28)
                    ws.set_column(1, len(nights_sorted), 11)
                    ws.set_column(1 + len(nights_sorted),
                                  1 + len(nights_sorted), 8)
                    ws.freeze_panes(3, 1)

        output = rawoutput.getvalue()
        cherrypy.response.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Real `datetime.now()` rather than the project's `localized_now()`
        # because the file name doesn't need timezone fidelity - same
        # convention as `@xlsx_file`.
        from datetime import datetime as _dt2
        stamp = _dt2.now().strftime('%Y%m%d_%H%M')
        cherrypy.response.headers['Content-Disposition'] = (
            f'attachment; filename="export_waitlist_xlsx{stamp}.xlsx"')
        return output

    @ajax
    def accept_waitlist(self, session, assignment_id=''):
        """FIFO-bypass accept: serve a single waitlisted RoomAssignment
        out of order, but ONLY for the nights where its block has
        actual capacity.

        Difference from `process_waitlist`: that endpoint sweeps the
        whole queue in `waitlist_started_at` order (earliest entrants
        first) and extends nights up to capacity. This endpoint picks
        ONE row by `assignment_id` and extends that row's nights up to
        capacity, regardless of where it sits in the FIFO order. So
        admins can promote a specific attendee past the queue without
        also handing them nights that don't actually exist.

        Per-night capacity uses `_partition_capacity` (same helper the
        cron uses) so a partition-bound row only competes with other
        rows in the same partition, and the cron and this endpoint
        agree on what "full" means.

        If the row's full waitlisted range is satisfied, the model's
        `clear_waitlist_when_satisfied` presave zeros the waitlist
        columns and `waitlist_started_at` so the row drops out of the
        queue. If only some of the requested nights had capacity, the
        row keeps its tightened waitlist demand on whatever's left.
        """
        from uber.models.hotel import RoomAssignment

        if not assignment_id:
            return {'error': 'Missing assignment_id.'}
        ra = session.query(RoomAssignment).get(assignment_id)
        if not ra:
            return {'error': 'Assignment not found.'}
        if not (ra.waitlisted_check_in_date or ra.waitlisted_check_out_date):
            return {'error': 'That assignment is not currently on the waitlist.'}
        if ra.export_locked:
            return {'error': 'That assignment has been exported to the hotel '
                             'and cannot be edited from here.'}
        if not ra.inventory:
            return {'error': 'Assignment has no inventory block; cannot run '
                             'the capacity check.'}

        wl_ci = ra.waitlisted_check_in_date or ra.assigned_check_in_date
        wl_co = ra.waitlisted_check_out_date or ra.assigned_check_out_date

        # Walk the FRONT extension one night at a time, closest-to-now
        # first. `_partition_capacity` counts every currently-confirmed
        # assignment whose `assigned_check_in_date <= night <
        # assigned_check_out_date` - that excludes the row we're
        # extending (whose current `assigned_check_in_date` is strictly
        # after `night`), so we don't double-count ourselves.
        nights_extended_front = 0
        while ra.assigned_check_in_date and wl_ci and wl_ci < ra.assigned_check_in_date:
            candidate_night = ra.assigned_check_in_date - timedelta(days=1)
            _, _, open_slots = _partition_capacity(
                session, ra.inventory, candidate_night, ra.partition_id)
            if open_slots <= 0:
                break
            ra.assigned_check_in_date = candidate_night
            nights_extended_front += 1
            # Flush so the next iteration's `_partition_capacity` sees
            # the in-memory change (otherwise we'd over-extend by
            # racing our own writes).
            session.flush()

        # Walk the BACK extension one night at a time, earliest first.
        # Same self-exclusion logic: `assigned_check_out_date > night`
        # filters us out for any night >= our current check-out.
        nights_extended_back = 0
        while ra.assigned_check_out_date and wl_co and wl_co > ra.assigned_check_out_date:
            candidate_night = ra.assigned_check_out_date
            _, _, open_slots = _partition_capacity(
                session, ra.inventory, candidate_night, ra.partition_id)
            if open_slots <= 0:
                break
            ra.assigned_check_out_date = candidate_night + timedelta(days=1)
            nights_extended_back += 1
            session.flush()

        total_extended = nights_extended_front + nights_extended_back

        # Cascade the new confirmed range to any connector children
        # (their dates always mirror the parent), then sync the
        # remaining waitlist demand too so the cron and the children
        # stay consistent.
        for child in session.query(RoomAssignment).filter_by(
                parent_assignment_id=ra.id).all():
            child.assigned_check_in_date = ra.assigned_check_in_date
            child.assigned_check_out_date = ra.assigned_check_out_date
            # The parent's `waitlisted_*` is either cleared (fully satisfied)
            # or holds the original request (partial); copy it to the children.
            child.waitlisted_check_in_date = ra.waitlisted_check_in_date
            child.waitlisted_check_out_date = ra.waitlisted_check_out_date
            child.waitlist_started_at = ra.waitlist_started_at
            session.add(child)

        session.add(ra)
        session.commit()

        if total_extended == 0:
            return {
                'error': 'No capacity available on any of the requested '
                         'nights for this block. The row remains on the '
                         'waitlist for the cron to retry.',
            }

        # Notify the attendee that some/all of their requested nights
        # came through. Same template the cron uses.
        if ra.attendee and ra.lottery_application:
            try:
                EmailService.queue_email(
                    session, 'hotel_lottery_waitlist_fulfilled', ra.lottery_application,
                    subject=f'{c.EVENT_NAME} Hotel Lottery - Room Dates Updated',
                    data={'assignment': ra, 'app': ra.lottery_application})
            except Exception:
                log.exception('accept_waitlist: notification failed')

        still_waiting = bool(ra.waitlisted_check_in_date
                             or ra.waitlisted_check_out_date)
        msg = (f'Accepted {total_extended} night(s) off waitlist. '
               f'New range: {ra.assigned_check_in_date} - '
               f'{ra.assigned_check_out_date}.')
        if still_waiting:
            msg += ' Remaining nights still on waitlist (no capacity yet).'

        return {
            'success': True,
            'message': msg,
            'still_waiting': still_waiting,
        }

    # A read-only "what's wrong with our room data" report. Issues are
    # surfaced as a flat list, each one carrying severity (error/warning),
    # a human label, and a deep-link to wherever the admin can fix it
    # (usually the application's edit form). The scan is intentionally
    # all in Python - no SQL view - so it's easy to add new checks.

    def room_issues(self, session, message='', severity='all', kind='all',
                    search='', show_hidden=''):
        """Cross-check every live RoomAssignment and produce a list of
        validation issues.

        Issue types currently detected:
          - orphan_connector: connector's parent_assignment_id doesn't
            resolve to a live assignment for the same attendee.
          - childless_parent: parent suite is awarded but one or more
            of its required connector children are missing or short.
          - over_capacity: occupants count > inventory.capacity.
          - under_capacity: occupants count < inventory.min_capacity.
          - empty_room: zero occupants assigned (no booker name on file
            for the hotel - they'll reject the reservation).
          - missing_dates: assigned_check_in_date or
            assigned_check_out_date is null.
          - inverted_dates: check_in_date >= check_out_date.
          - too_short: stay length < 1 night.
          - out_of_range: check-in before HOTEL_LOTTERY_CHECKIN_START
            or check-out after HOTEL_LOTTERY_CHECKOUT_END.
          - status_mismatch: app.status doesn't match its rooms
            (e.g. has rooms but status==COMPLETE; no rooms but
            status==AWARDED). Caught by the model listener under normal
            use; this is the audit backstop.
          - double_booked: same attendee is an occupant on two rooms
            whose dates overlap.
          - secured_without_payment: self-pay room flipped to SECURED
            without a captured CC vault token. The hotel will treat the
            reservation as unguaranteed; master-bill rooms exempt.
        """
        from datetime import timedelta as _td
        from collections import defaultdict

        # All live assignments - these are what we audit.
        live = (session.query(RoomAssignment)
                .filter(RoomAssignment.status.in_(
                    [c.ASSIGNED, c.SECURED]))
                .all())

        by_id = {ra.id: ra for ra in live}
        # Build group-by-attendee for orphan detection (an orphan child
        # is one whose parent isn't tied to the SAME attendee).
        by_attendee = defaultdict(list)
        for ra in live:
            by_attendee[ra.attendee_id].append(ra)

        # Build group-by-parent for childless-parent detection.
        children_of = defaultdict(list)
        for ra in live:
            if ra.parent_assignment_id:
                children_of[ra.parent_assignment_id].append(ra)

        # Lookup: room_type_id -> list of (parent_type_id, qty) it must
        # follow. With the current single-parent model that's at most
        # one entry, but we model it as a list so the same code handles
        # any future multi-parent extension.
        room_type_parents = {}
        room_type_children_needed = defaultdict(list)
        for rt in session.query(LotteryRoomType).all():
            if rt.connects_to_type_id:
                room_type_parents[rt.id] = (rt.connects_to_type_id,
                                            rt.connector_quantity)
                room_type_children_needed[rt.connects_to_type_id].append(
                    (rt.id, rt.connector_quantity, rt.name))

        issues = []

        def _add(severity, kind, label, assignment, fix_url=None, extra=None):
            issues.append({
                'severity': severity,
                'kind': kind,
                'label': label,
                'assignment': assignment,
                'fix_url': fix_url,
                'extra': extra or {},
            })

        ci_start = c.HOTEL_LOTTERY_CHECKIN_START.date() if c.HOTEL_LOTTERY_CHECKIN_START else None
        co_end = c.HOTEL_LOTTERY_CHECKOUT_END.date() if c.HOTEL_LOTTERY_CHECKOUT_END else None

        for ra in live:
            app_id = ra.lottery_application_id
            fix = f'form?id={app_id}' if app_id else None
            inv = ra.inventory

            # Orphan connector
            if ra.parent_assignment_id:
                parent = by_id.get(ra.parent_assignment_id)
                if not parent or parent.attendee_id != ra.attendee_id:
                    _add('error', 'orphan_connector',
                         "Connector room without a matching parent suite "
                         "assigned to the same attendee.",
                         ra, fix)
                elif parent.status not in (c.ASSIGNED, c.SECURED):
                    _add('error', 'orphan_connector',
                         f"Connector's parent suite is in status "
                         f"{parent.status_label}, not live.",
                         ra, fix)
                else:
                    # Parent exists and is live - but does it actually
                    # require a connector of THIS type? If an admin edited
                    # the parent's inventory to a different room type, the
                    # connector is now hanging off something that doesn't
                    # need it.
                    child_inv = ra.inventory
                    child_type_id = (child_inv.room_type_id
                                     or child_inv.suite_type_id) if child_inv else None
                    parent_inv = parent.inventory
                    parent_type_id = (parent_inv.room_type_id
                                      or parent_inv.suite_type_id) if parent_inv else None
                    expected_parent_type_id = None
                    if child_type_id and child_type_id in room_type_parents:
                        expected_parent_type_id = room_type_parents[child_type_id][0]
                    if not expected_parent_type_id:
                        # This child's room type isn't configured as a
                        # connector at all anymore - somebody removed the
                        # `connects_to_type_id` mapping after the room
                        # was awarded.
                        _add('error', 'orphan_connector',
                             "Connector's room type is no longer "
                             "configured to follow any parent.",
                             ra, fix)
                    elif expected_parent_type_id != parent_type_id:
                        _add('error', 'orphan_connector',
                             "Connector's parent assignment is no longer "
                             "the correct room type to require this "
                             "connector.",
                             ra, fix)

            # Dates
            if not ra.assigned_check_in_date or not ra.assigned_check_out_date:
                _add('error', 'missing_dates',
                     "Assignment is missing check-in and/or check-out date.",
                     ra, fix)
            else:
                if ra.assigned_check_in_date >= ra.assigned_check_out_date:
                    _add('error', 'inverted_dates',
                         "Check-out date is on or before the check-in date.",
                         ra, fix)
                else:
                    nights = (ra.assigned_check_out_date
                              - ra.assigned_check_in_date).days
                    if nights < 1:
                        _add('error', 'too_short',
                             "Stay is less than one night long.",
                             ra, fix)
                if ci_start and ra.assigned_check_in_date < ci_start:
                    _add('warning', 'out_of_range',
                         f"Check-in {ra.assigned_check_in_date} is before "
                         f"the event window opens ({ci_start}).",
                         ra, fix)
                if co_end and ra.assigned_check_out_date > co_end:
                    _add('warning', 'out_of_range',
                         f"Check-out {ra.assigned_check_out_date} is after "
                         f"the event window closes ({co_end}).",
                         ra, fix)

            # Occupants vs inventory capacity. The booker (attendee_id)
            # is ALWAYS implicitly an occupant - they're the name on the
            # reservation - even when the room_assignment_occupant M2M
            # row is missing (the ensure_booker_is_occupant presave keeps
            # it in sync, but older/imported rows may lack it). Fold the
            # booker into the set so a room with a booker is never flagged
            # "empty", and the booker still counts toward capacity.
            occupant_ids = {o.id for o in (getattr(ra, 'occupants', None) or [])}
            if ra.attendee_id:
                occupant_ids.add(ra.attendee_id)
            occupant_count = len(occupant_ids)
            if inv:
                cap = inv.capacity or 0
                min_cap = inv.min_capacity or 0
                if occupant_count == 0:
                    _add('error', 'empty_room',
                         "No occupants assigned - the hotel needs a name "
                         "on the reservation.",
                         ra, fix)
                elif cap and occupant_count > cap:
                    _add('error', 'over_capacity',
                         f"{occupant_count} occupants in a room with "
                         f"capacity {cap}.",
                         ra, fix)
                elif min_cap and occupant_count < min_cap:
                    _add('warning', 'under_capacity',
                         f"{occupant_count} occupants in a room with "
                         f"minimum {min_cap}.",
                         ra, fix)

            # Secured-without-payment: status flipped to SECURED on a
            # self-pay room but no CC vault token was ever captured. The
            # secure flow normally requires the token before flipping, so
            # this means the row was edited around the flow (admin override,
            # imported state, etc.). Master-bill rooms (require_cc=False)
            # are exempt - they have no payment info by design.
            if ra.status == c.SECURED and ra.require_cc and not ra.cc_token:
                _add('error', 'secured_without_payment',
                     "Room is marked Secured but has no credit card on "
                     "file - the hotel will treat the reservation as "
                     "unguaranteed.",
                     ra, fix)

        # For each live primary whose room type is a connector parent,
        # tally the live children by type and warn when any required
        # type is short.
        for parent_ra in live:
            if parent_ra.parent_assignment_id:
                continue  # only check primaries
            inv = parent_ra.inventory
            if not inv:
                continue
            parent_type_id = inv.room_type_id or inv.suite_type_id
            specs = room_type_children_needed.get(parent_type_id, [])
            if not specs:
                continue
            kids_by_type = defaultdict(int)
            for child in children_of.get(parent_ra.id, []):
                ci = child.inventory
                if ci:
                    kt = ci.room_type_id or ci.suite_type_id
                    kids_by_type[kt] += 1
            for child_type_id, needed_qty, child_type_name in specs:
                got = kids_by_type.get(child_type_id, 0)
                if got < needed_qty:
                    fix = f'form?id={parent_ra.lottery_application_id}' if parent_ra.lottery_application_id else None
                    _add('error', 'childless_parent',
                         f"Suite is missing required connector "
                         f"'{child_type_name}': has {got}, needs "
                         f"{needed_qty}.",
                         parent_ra, fix)

        # Apps with rooms should be AWARDED; apps without any live rooms
        # should NOT be AWARDED. The listener flips these in real time,
        # so any hit here means data was edited around the listener.
        app_ids_with_rooms = {ra.lottery_application_id for ra in live
                              if ra.lottery_application_id}
        if app_ids_with_rooms:
            mismatched = (session.query(LotteryApplication)
                          .filter(LotteryApplication.id.in_(app_ids_with_rooms))
                          .filter(LotteryApplication.status != c.AWARDED)
                          .filter(LotteryApplication.status != c.PROCESSED)
                          .all())
            for app in mismatched:
                _add('warning', 'status_mismatch',
                     f"Application has live rooms but status is "
                     f"{app.status_label}.",
                     None,
                     fix_url=f'form?id={app.id}',
                     extra={'application': app})

        # Apps marked AWARDED with zero live rooms - same story, other side.
        awarded_apps = (session.query(LotteryApplication)
                        .filter(LotteryApplication.status == c.AWARDED).all())
        for app in awarded_apps:
            if app.id not in app_ids_with_rooms:
                _add('warning', 'status_mismatch',
                     "Application is marked AWARDED but has no live rooms.",
                     None,
                     fix_url=f'form?id={app.id}',
                     extra={'application': app})

        # An attendee listed as occupant of two assignments whose dates
        # overlap. We check across `occupants` (the M2M), not
        # `attendee_id` - the booker can hold many rooms intentionally,
        # but an occupant can't be in two rooms on the same night.
        occupant_rooms = defaultdict(list)
        for ra in live:
            if not ra.assigned_check_in_date or not ra.assigned_check_out_date:
                continue
            for occ in (getattr(ra, 'occupants', None) or []):
                occupant_rooms[occ.id].append(ra)
        for occ_id, ras in occupant_rooms.items():
            if len(ras) < 2:
                continue
            # Pairwise overlap check (n is small in practice).
            for i, a in enumerate(ras):
                for b in ras[i+1:]:
                    if (a.assigned_check_in_date < b.assigned_check_out_date
                            and b.assigned_check_in_date < a.assigned_check_out_date):
                        # Connector + its own parent overlapping doesn't count
                        # - the connector is part of the same block.
                        if (a.parent_assignment_id == b.id
                                or b.parent_assignment_id == a.id):
                            continue
                        fix = f'form?id={a.lottery_application_id}' if a.lottery_application_id else None
                        _add('warning', 'double_booked',
                             f"Occupant is on two overlapping rooms "
                             f"({a.assigned_check_in_date}->"
                             f"{a.assigned_check_out_date} and "
                             f"{b.assigned_check_in_date}->"
                             f"{b.assigned_check_out_date}).",
                             a, fix,
                             extra={'other_assignment': b})

        # These don't attach to a RoomAssignment - they describe a
        # mismatch between configured inventory and the lottery's load
        # (or static configuration mistakes). Collected into a separate
        # list so the template can put them on their own tab.
        inv_issues = []

        def _inv_add(severity, kind, label, inventory=None, partition=None,
                     room_type=None, fix_url=None, extra=None):
            inv_issues.append({
                'severity': severity,
                'kind': kind,
                'label': label,
                'inventory': inventory,
                'partition': partition,
                'room_type': room_type,
                'fix_url': fix_url,
                'extra': extra or {},
            })

        # All active inventory, plus the partition-block index keyed by
        # inventory_id so the per-partition oversubscription check is a
        # constant-time lookup per row.
        all_inventory = (session.query(HotelRoomInventory)
                         .filter_by(active=True).all())
        partition_blocks_by_inv = defaultdict(list)
        for pb in session.query(InventoryPartitionBlock).all():
            partition_blocks_by_inv[str(pb.inventory_id)].append(pb)

        # Live RAs grouped by inventory for the oversubscription scans.
        live_by_inv = defaultdict(list)
        for ra in live:
            if ra.inventory_id:
                live_by_inv[str(ra.inventory_id)].append(ra)

        for inv in all_inventory:
            inv_id = str(inv.id)
            ras = live_by_inv.get(inv_id, [])
            fix = f'edit_inventory_item?id={inv.id}'

            # Static config sanity:
            #   - active inventory at quantity 0 = mistake
            #   - active inventory whose hotel/type is inactive = stranded
            if inv.quantity == 0:
                _inv_add('warning', 'zero_quantity',
                         "Active inventory configured with quantity zero - "
                         "either deactivate it or set a non-zero quantity.",
                         inventory=inv, fix_url=fix)
            if inv.hotel and not inv.hotel.active:
                _inv_add('warning', 'inactive_parent',
                         f"Inventory points at inactive hotel "
                         f"'{inv.hotel.name}'.",
                         inventory=inv, fix_url=fix)
            rt = inv.suite_type if inv.is_suite else inv.room_type
            if rt and not rt.active:
                _inv_add('warning', 'inactive_parent',
                         f"Inventory points at inactive room type "
                         f"'{rt.name}'.",
                         inventory=inv, fix_url=fix)
            if inv.is_suite and not inv.suite_type:
                _inv_add('warning', 'type_mismatch',
                         "Inventory is flagged as a suite but has no "
                         "suite_type set.",
                         inventory=inv, fix_url=fix)
            if not inv.is_suite and not inv.room_type:
                _inv_add('warning', 'type_mismatch',
                         "Inventory is flagged as a standard room but has "
                         "no room_type set.",
                         inventory=inv, fix_url=fix)

            # Partition blocks summing to more than the inventory's cap:
            # a configuration bug, not a runtime overlap.
            blocks = partition_blocks_by_inv.get(inv_id, [])
            partition_total = sum(pb.quantity for pb in blocks)
            if inv.quantity and partition_total > inv.quantity:
                _inv_add('error', 'partition_overallocated',
                         f"Partition blocks sum to {partition_total} rooms "
                         f"but this inventory only has {inv.quantity}.",
                         inventory=inv, fix_url=fix,
                         extra={'partition_total': partition_total})

            # Night-level oversubscription: walk every night in the
            # event window covered by an assignment and tally occupancy
            # vs. capacity. inv.night_quantity_map (when populated)
            # overrides inv.quantity per-night.
            if not ras:
                continue
            nq_map = getattr(inv, 'night_quantity_map', None) or {}
            night_occupancy = defaultdict(int)
            night_partition_occupancy = defaultdict(lambda: defaultdict(int))
            for ra in ras:
                if not ra.assigned_check_in_date or not ra.assigned_check_out_date:
                    continue
                d = ra.assigned_check_in_date
                while d < ra.assigned_check_out_date:
                    night_occupancy[d] += 1
                    if ra.partition_id:
                        night_partition_occupancy[ra.partition_id][d] += 1
                    d = d + timedelta(days=1)

            # Inventory-wide oversubscription.
            bad_nights = []
            for night, used in sorted(night_occupancy.items()):
                cap = nq_map.get(night, inv.quantity)
                if used > cap:
                    bad_nights.append((night, used, cap))
            if bad_nights:
                _inv_add('error', 'oversubscribed_inventory',
                         f"Inventory is oversubscribed on "
                         f"{len(bad_nights)} night(s). First: "
                         f"{bad_nights[0][0]} ({bad_nights[0][1]} "
                         f"assigned vs cap {bad_nights[0][2]}).",
                         inventory=inv, fix_url=fix,
                         extra={'bad_nights': bad_nights})

            # Per-partition oversubscription. Each partition block has a
            # carve-out quantity, so an inventory can be fine in aggregate
            # but still exceed a single partition's slice.
            blocks_by_partition = {str(pb.partition_id): pb for pb in blocks}
            for part_id, by_night in night_partition_occupancy.items():
                pb = blocks_by_partition.get(str(part_id))
                if not pb:
                    # Live RAs assigned to a partition that has no block
                    # on this inventory - also a configuration issue.
                    bad_partition = session.query(InventoryPartition).get(part_id)
                    part_name = (bad_partition.name
                                 if bad_partition else f"id {part_id}")
                    _inv_add('error', 'partition_unconfigured',
                             f"Partition '{part_name}' has live "
                             f"assignments on this inventory but no "
                             f"matching partition block. Add a block "
                             f"(or move the assignments).",
                             inventory=inv,
                             partition=bad_partition,
                             fix_url=(f'edit_partition?id={part_id}'
                                      if bad_partition else fix))
                    continue
                bad = []
                for night, used in sorted(by_night.items()):
                    cap = pb.quantity
                    if used > cap:
                        bad.append((night, used, cap))
                if bad:
                    _inv_add('error', 'oversubscribed_partition',
                             f"Partition '{pb.partition.name}' is "
                             f"oversubscribed on {len(bad)} night(s). "
                             f"First: {bad[0][0]} ({bad[0][1]} "
                             f"assigned vs cap {bad[0][2]}).",
                             inventory=inv, partition=pb.partition,
                             fix_url=fix, extra={'bad_nights': bad})

        # If a room type follows another, the lottery (when it awards a
        # full parent inventory) needs enough child inventory to satisfy
        # the coupling. Check at two granularities: globally and per-hotel
        # (so an Exec Suite that's only offered at hotel X must have its
        # Standard King connectors also at hotel X).
        inv_qty_by_type = defaultdict(int)
        inv_qty_by_hotel_type = defaultdict(lambda: defaultdict(int))
        for inv in all_inventory:
            tid = inv.suite_type_id if inv.is_suite else inv.room_type_id
            if not tid:
                continue
            inv_qty_by_type[tid] += inv.quantity or 0
            inv_qty_by_hotel_type[inv.hotel_id][tid] += inv.quantity or 0

        all_types_by_id = {rt.id: rt for rt in
                           session.query(LotteryRoomType).all()}

        for rt in all_types_by_id.values():
            if not rt.connects_to_type_id or rt.connector_quantity <= 0:
                continue
            parent = all_types_by_id.get(rt.connects_to_type_id)
            if not parent:
                _inv_add('error', 'broken_connector_config',
                         f"Room type '{rt.name}' follows a parent that "
                         "no longer exists.",
                         room_type=rt,
                         fix_url=f'edit_room_type?id={rt.id}')
                continue

            parent_total = inv_qty_by_type.get(parent.id, 0)
            child_total = inv_qty_by_type.get(rt.id, 0)
            needed = parent_total * rt.connector_quantity
            if parent_total > 0 and child_total < needed:
                _inv_add('error', 'insufficient_connectors',
                         f"Configured inventory of connector '{rt.name}' "
                         f"({child_total}) is below what parent "
                         f"'{parent.name}' would need if fully awarded "
                         f"({parent_total} x {rt.connector_quantity} = "
                         f"{needed}).",
                         room_type=rt,
                         fix_url=f'edit_room_type?id={rt.id}',
                         extra={'parent_total': parent_total,
                                'child_total': child_total,
                                'needed': needed})

            # Per-hotel: every hotel offering the parent needs the
            # child inventory locally too - connectors are physical
            # neighbors, they can't follow across properties.
            for hotel_id, types_at_hotel in inv_qty_by_hotel_type.items():
                p_here = types_at_hotel.get(parent.id, 0)
                if p_here <= 0:
                    continue
                c_here = types_at_hotel.get(rt.id, 0)
                local_needed = p_here * rt.connector_quantity
                if c_here < local_needed:
                    hotel = session.query(LotteryHotel).get(hotel_id)
                    _inv_add('error', 'insufficient_connectors_at_hotel',
                             f"Hotel '{hotel.name if hotel else hotel_id}' "
                             f"offers {p_here} '{parent.name}' rooms but "
                             f"only has {c_here} '{rt.name}' connector "
                             f"rooms (needs {local_needed}).",
                             room_type=rt,
                             fix_url=f'edit_room_type?id={rt.id}',
                             extra={'hotel': hotel,
                                    'parent_total_here': p_here,
                                    'child_total_here': c_here,
                                    'needed_here': local_needed})

        # Issues are recomputed every load, so admin hide-flags + notes
        # live in HotelRoomIssueNote keyed to each issue's STABLE identity
        # (kind, target_type, target_id). Annotate every issue with its
        # note + hidden flag, then split shown vs hidden.
        notes_by_key = {
            (n.issue_kind, n.target_type, n.target_id): n
            for n in session.query(HotelRoomIssueNote).all()
        }

        def _issue_identity(iss):
            ra = iss.get('assignment')
            inv = iss.get('inventory')
            rt = iss.get('room_type')
            part = iss.get('partition')
            extra_app = (iss.get('extra') or {}).get('application')
            if ra is not None:
                return ('room_assignment', str(ra.id))
            if inv is not None:
                # Pair inventory + partition for partition-scoped issues so
                # hiding one partition's oversubscription doesn't hide the
                # whole block's.
                tid = str(inv.id)
                if part is not None:
                    tid += '|' + str(part.id)
                return ('inventory', tid)
            if rt is not None:
                return ('room_type', str(rt.id))
            if part is not None:
                return ('partition', str(part.id))
            if extra_app is not None:
                return ('lottery_application', str(extra_app.id))
            return ('other', iss.get('kind') or 'unknown')

        def _annotate(iss):
            ttype, tid = _issue_identity(iss)
            iss['target_type'] = ttype
            iss['target_id'] = tid
            note = notes_by_key.get((iss.get('kind'), ttype, tid))
            iss['note'] = note
            iss['admin_notes'] = note.admin_notes if note else ''
            iss['hidden'] = bool(note and note.hidden)

        for iss in issues:
            _annotate(iss)
        for iss in inv_issues:
            _annotate(iss)

        shown_room = [i for i in issues if not i['hidden']]
        hidden_room = [i for i in issues if i['hidden']]
        shown_inv = [i for i in inv_issues if not i['hidden']]
        hidden_inv = [i for i in inv_issues if i['hidden']]

        # Tab + kind counts reflect the SHOWN (non-hidden) issues, since
        # hidden ones are acknowledged; the hidden total gets its own
        # count. Filters then narrow the working lists before grouping.
        counts = {'error': 0, 'warning': 0}
        for iss in shown_room:
            counts[iss['severity']] = counts.get(iss['severity'], 0) + 1
        inv_counts = {'error': 0, 'warning': 0}
        for iss in shown_inv:
            inv_counts[iss['severity']] = inv_counts.get(iss['severity'], 0) + 1
        hidden_count = len(hidden_room) + len(hidden_inv)

        # Per-kind counts across the SHOWN issue sets, for the dropdown.
        kind_counts = {}
        for iss in shown_room + shown_inv:
            k = iss.get('kind') or 'unknown'
            kind_counts[k] = kind_counts.get(k, 0) + 1
        kind_options = sorted(kind_counts.items())

        # Free-text search builds a lowercase haystack per issue from the
        # human-relevant context (hotel, room type, attendee, conf #,
        # partition) plus the issue kind/label, so admins can search by
        # any of them. Keep the original-case string for redisplay; match
        # on the lowercased needle.
        search = (search or '').strip()
        needle = search.lower()

        def _haystack(iss):
            parts = [iss.get('kind') or '', iss.get('label') or '',
                     iss.get('severity') or '']
            ra = iss.get('assignment')
            if ra is not None:
                inv = ra.inventory
                if inv:
                    parts.append(inv.name or '')
                    if inv.hotel:
                        parts.append(inv.hotel.name or '')
                    rt = inv.suite_type if inv.is_suite else inv.room_type
                    if rt:
                        parts.append(rt.name or '')
                if ra.attendee:
                    parts.append(ra.attendee.full_name or '')
                    parts.append(ra.attendee.email or '')
                app = ra.lottery_application
                if app and app.confirmation_num:
                    parts.append(app.confirmation_num)
            # Inventory-level issue context.
            inv2 = iss.get('inventory')
            if inv2 is not None:
                parts.append(inv2.name or '')
                if inv2.hotel:
                    parts.append(inv2.hotel.name or '')
            rt2 = iss.get('room_type')
            if rt2 is not None:
                parts.append(rt2.name or '')
            part = iss.get('partition')
            if part is not None:
                parts.append(part.name or '')
            # status_mismatch and friends stash the application in extra.
            extra_app = (iss.get('extra') or {}).get('application')
            if extra_app is not None:
                if extra_app.confirmation_num:
                    parts.append(extra_app.confirmation_num)
                if extra_app.attendee:
                    parts.append(extra_app.attendee.full_name or '')
            return ' '.join(parts).lower()

        def _passes(iss):
            if severity in ('error', 'warning') and iss['severity'] != severity:
                return False
            if kind not in ('all', '', None) and iss.get('kind') != kind:
                return False
            if needle and needle not in _haystack(iss):
                return False
            return True

        shown_room = [i for i in shown_room if _passes(i)]
        shown_inv = [i for i in shown_inv if _passes(i)]
        hidden_room = [i for i in hidden_room if _passes(i)]
        hidden_inv = [i for i in hidden_inv if _passes(i)]

        # Room issues group by RA id (or `app:<id>` for application-level
        # issues); inventory issues group by inventory (or room type).
        # Each group inherits the worst severity of its issues. The same
        # helpers group the shown and hidden lists.
        def _group_sort_key(g):
            sev = 0 if g['severity'] == 'error' else 1
            ra = g['assignment']
            hotel_name = (ra.inventory.hotel.name
                          if ra and ra.inventory and ra.inventory.hotel else '~')
            conf = (g['application'].confirmation_num
                    if g['application'] else '~')
            return (sev, hotel_name, conf or '~')

        def _inv_group_sort_key(g):
            sev = 0 if g['severity'] == 'error' else 1
            inv = g['inventory']
            rt = g['room_type']
            hotel_name = (inv.hotel.name if inv and inv.hotel else '~')
            inv_name = (inv.name if inv else (rt.name if rt else '~'))
            return (sev, hotel_name, inv_name)

        def _group_rooms(issue_list):
            by_key = {}
            order = []
            for iss in issue_list:
                ra = iss.get('assignment')
                app = (iss.get('extra') or {}).get('application')
                if ra:
                    key = ra.id
                    group_app = app or ra.lottery_application
                elif app:
                    key = f'app:{app.id}'
                    group_app = app
                else:
                    key = f'orphaned-issue-{len(order)}'
                    group_app = None
                if key not in by_key:
                    by_key[key] = {
                        'key': key, 'assignment': ra, 'application': group_app,
                        'issues': [], 'severity': 'warning',
                    }
                    order.append(key)
                by_key[key]['issues'].append(iss)
                if iss['severity'] == 'error':
                    by_key[key]['severity'] = 'error'
            return sorted([by_key[k] for k in order], key=_group_sort_key)

        def _group_inv(issue_list):
            by_key = {}
            order = []
            for iss in issue_list:
                inv = iss.get('inventory')
                rt = iss.get('room_type')
                if inv:
                    key = f'inv:{inv.id}'
                elif rt:
                    key = f'rt:{rt.id}'
                else:
                    key = f'other:{len(order)}'
                if key not in by_key:
                    by_key[key] = {
                        'key': key, 'inventory': inv, 'room_type': rt,
                        'issues': [], 'severity': 'warning',
                    }
                    order.append(key)
                by_key[key]['issues'].append(iss)
                if iss['severity'] == 'error':
                    by_key[key]['severity'] = 'error'
            return sorted([by_key[k] for k in order], key=_inv_group_sort_key)

        groups = _group_rooms(shown_room)
        inv_groups = _group_inv(shown_inv)
        hidden_groups = _group_rooms(hidden_room)
        hidden_inv_groups = _group_inv(hidden_inv)

        return {
            'groups': groups,
            'counts': counts,
            'inv_groups': inv_groups,
            'inv_counts': inv_counts,
            'hidden_groups': hidden_groups,
            'hidden_inv_groups': hidden_inv_groups,
            'hidden_count': hidden_count,
            'show_hidden': str(show_hidden).lower() in ('1', 'true', 'on', 'yes'),
            'severity': severity,
            'kind': kind,
            'kind_options': kind_options,
            'search': search,
            'message': message,
        }

    def _get_or_make_issue_note(self, session, issue_kind, target_type,
                                target_id):
        """Fetch (or create, unsaved) the HotelRoomIssueNote for an
        issue's stable identity. Stamps the acting admin so we know who
        last touched it."""
        from uber.lottery_perms import _current_admin_account
        note = (session.query(HotelRoomIssueNote)
                .filter_by(issue_kind=issue_kind, target_type=target_type,
                           target_id=str(target_id))
                .one_or_none())
        if not note:
            note = HotelRoomIssueNote(
                issue_kind=issue_kind, target_type=target_type,
                target_id=str(target_id))
        admin = _current_admin_account(session)
        if admin:
            note.admin_account_id = admin.id
        return note

    def hide_issue(self, session, issue_kind='', target_type='',
                   target_id='', admin_notes='', severity='all', kind='all',
                   search='', show_hidden='', csrf_token=None):
        """Hide a single issue (by its stable identity) from the active
        report and optionally attach a note. Idempotent."""
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect(_room_issues_url())
        check_csrf(csrf_token)
        if not (issue_kind and target_type and target_id):
            raise HTTPRedirect(_room_issues_url(
                'Could not identify which issue to hide.',
                severity, kind, search, show_hidden))
        note = self._get_or_make_issue_note(
            session, issue_kind, target_type, target_id)
        note.hidden = True
        note.admin_notes = (admin_notes or '').strip()
        session.add(note)
        session.commit()
        raise HTTPRedirect(_room_issues_url(
            'Issue hidden.', severity, kind, search, show_hidden))

    def unhide_issue(self, session, issue_kind='', target_type='',
                     target_id='', severity='all', kind='all', search='',
                     show_hidden='1', csrf_token=None):
        """Un-hide an issue so it returns to the active report. The note
        text is preserved (the row stays, just `hidden=False`)."""
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect(_room_issues_url())
        check_csrf(csrf_token)
        note = (session.query(HotelRoomIssueNote)
                .filter_by(issue_kind=issue_kind, target_type=target_type,
                           target_id=str(target_id))
                .one_or_none())
        if note:
            note.hidden = False
            session.add(note)
            session.commit()
        raise HTTPRedirect(_room_issues_url(
            'Issue restored to the active report.',
            severity, kind, search, show_hidden))

    def save_issue_note(self, session, issue_kind='', target_type='',
                        target_id='', admin_notes='', severity='all',
                        kind='all', search='', show_hidden='',
                        csrf_token=None):
        """Save (or update) an issue's admin note without changing its
        hidden state. An empty note on a not-hidden issue deletes the
        row so we don't accumulate empty records."""
        if cherrypy.request.method != 'POST':
            raise HTTPRedirect(_room_issues_url())
        check_csrf(csrf_token)
        if not (issue_kind and target_type and target_id):
            raise HTTPRedirect(_room_issues_url(
                'Could not identify which issue to annotate.',
                severity, kind, search, show_hidden))
        notes = (admin_notes or '').strip()
        existing = (session.query(HotelRoomIssueNote)
                    .filter_by(issue_kind=issue_kind, target_type=target_type,
                               target_id=str(target_id))
                    .one_or_none())
        if not notes and (not existing or not existing.hidden):
            # Nothing to keep - drop an empty, non-hidden row.
            if existing:
                session.delete(existing)
                session.commit()
            raise HTTPRedirect(_room_issues_url(
                'Note cleared.', severity, kind, search, show_hidden))
        note = existing or self._get_or_make_issue_note(
            session, issue_kind, target_type, target_id)
        note.admin_notes = notes
        session.add(note)
        session.commit()
        raise HTTPRedirect(_room_issues_url(
            'Note saved.', severity, kind, search, show_hidden))
