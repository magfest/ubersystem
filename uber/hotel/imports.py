"""Shared hotel confirmation/cancellation file import.

Used by both the uber-vault hotel portal (uber.api.HotelLookup.import_confirmation_file)
and the hotel lottery admin upload. Parses a CSV or XLSX of confirmation and/or
cancellation numbers, applies them to room assignments keyed by confirmation_num,
and retains the raw uploaded file (on disk under UPLOADED_FILES_DIR) for later
debugging so odd hotel formats can be fixed after the fact.
"""
import csv
import io
import os
import uuid
from datetime import date, datetime

from pytz import UTC

from uber.config import c


def _normalize(value):
    return str(value if value is not None else '').strip().lower().replace(' ', '_')


def cell_to_str(value):
    """Render one spreadsheet cell as a string.

    openpyxl returns date/datetime objects for date-formatted XLSX cells;
    those are rendered as ISO 8601 so downstream date parsing
    (parse_iso_date) and string matching behave identically for CSV and
    XLSX uploads. None becomes ''.
    """
    if value is None:
        return ''
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def parse_iso_date(raw):
    """Parse an ISO 8601 date ('YYYY-MM-DD') or datetime string to a date.

    Hotels may quote back either a bare date or a full datetime with
    optional offset; datetimes are projected to their date. Returns None if
    the value is blank or unparseable.
    """
    if not raw:
        return None
    raw = raw.strip()
    if not raw:
        return None
    # date.fromisoformat handles 'YYYY-MM-DD' on its own. For datetimes we
    # take the leading date portion.
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def parse_spreadsheet(raw, filename, sheet_name=None, header_row=1):
    """Parse CSV or XLSX file bytes into (fieldnames, rows, error).

    fieldnames are the normalized header cells (lowercased, stripped, spaces
    treated as underscores) and rows are dicts keyed by those names, with
    every value a string (dates/datetimes rendered as ISO 8601 via
    cell_to_str). XLSX is detected by filename extension or the PK zip magic
    bytes. Parse failures come back as an error string rather than raising.

    `sheet_name` picks a worksheet by name when a hotel's format does not put
    the data on the first sheet; `header_row` skips leading junk rows above
    the header. Both default to the original behavior.
    """
    name = (filename or '').lower()
    fieldnames = []
    rows = []
    try:
        if name.endswith(('.xlsx', '.xlsm')) or raw[:2] == b'PK':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = (wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames
                     else wb.active)
            header = None
            skip = max(0, int(header_row or 1) - 1)
            for excel_row in sheet.iter_rows(values_only=True):
                if skip:
                    skip -= 1
                    continue
                if header is None:
                    header = [_normalize(cell) for cell in excel_row]
                    continue
                rows.append({header[i]: cell_to_str(v)
                             for i, v in enumerate(excel_row)
                             if i < len(header) and header[i]})
            fieldnames = [h for h in (header or []) if h]
        else:
            text = raw.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            for record in reader:
                rows.append({_normalize(k): cell_to_str(v)
                             for k, v in record.items() if k})
            fieldnames = [_normalize(f) for f in (reader.fieldnames or []) if f]
    except Exception as e:
        return [], [], f'Could not parse file: {e}'
    return fieldnames, rows, None


def parse_confirmation_rows(raw, filename):
    """Parse a CSV or XLSX file into a list of normalized-key dict rows.

    Returns (rows, error). Columns are matched case-insensitively with spaces
    treated as underscores. Parse failures come back as an error string rather
    than raising.
    """
    _fieldnames, rows, error = parse_spreadsheet(raw, filename)
    return rows, error


def _store_file(session, raw, filename, content_type, hotel, source, uploaded_by):
    from uber.models.hotel import HotelImportFile

    ext = os.path.splitext(filename or '')[1][:10]
    stored_name = f"hotel_import_{uuid.uuid4().hex}{ext}"
    os.makedirs(c.UPLOADED_FILES_DIR, exist_ok=True)
    filepath = os.path.join(c.UPLOADED_FILES_DIR, stored_name)
    with open(filepath, 'wb') as f:
        f.write(raw)

    record = HotelImportFile(
        hotel_id=hotel.id if hotel else None,
        filename=filename or stored_name,
        # CherryPy hands back a HeaderElement for an upload's type, which
        # psycopg can't adapt; the column just wants the text.
        content_type=str(content_type or ''),
        filepath=filepath,
        size=len(raw),
        source=source,
        uploaded_by=uploaded_by or '',
    )
    session.add(record)
    session.flush()
    return record


def import_confirmation_file(session, raw, filename, hotel=None, source='',
                             uploaded_by='', content_type=''):
    """Store the raw file, then apply confirmation/cancellation numbers.

    Recognized columns (case-insensitive, spaces treated as underscores):
      - confirmation_num (required key; identifies the attendee booking)
      - hotel_confirmation_number (optional)
      - hotel_cancellation_number (optional)

    Whichever hotel-number columns are present are applied, keyed by
    confirmation_num. Unknown columns are ignored, empty cells leave existing
    values alone, and rows that don't match a booking are skipped. The file is
    retained even when parsing fails. Returns {updated, unchanged, changes,
    error}.
    """
    from uber.models import LotteryApplication
    from uber.models.hotel import RoomAssignment, HotelExportLog

    record = _store_file(session, raw, filename, content_type, hotel, source, uploaded_by)

    rows, error = parse_confirmation_rows(raw, filename)
    if error:
        record.note = error
        session.flush()
        return {'updated': 0, 'unchanged': 0, 'changes': [], 'error': error,
                'record': record}

    # (file column, RoomAssignment attribute)
    fields = [('hotel_confirmation_number', 'hotel_confirmation_number'),
              ('hotel_cancellation_number', 'cancellation_confirmation_number')]

    # The change feed records no link back to this file, so bracket the
    # writes: everything tracked inside this window is this import's
    # doing (see uber.hotel.exports.import_changes).
    record.applied_from = datetime.now(UTC)

    updated = 0
    unchanged = 0
    changes = []
    hotels_imported = set()
    for row in rows:
        conf_num = (row.get('confirmation_num') or '').strip()
        if not conf_num:
            continue
        app = session.query(LotteryApplication).filter(
            LotteryApplication.confirmation_num == conf_num).one_or_none()
        ras = session.query(RoomAssignment).filter_by(
            lottery_application_id=app.id).all() if app else []
        if not ras:
            continue  # no matching booking; skip the row

        row_present = False
        row_changed = False
        for col, attr in fields:
            if col not in row:
                continue
            new_val = (row.get(col) or '').strip()
            if not new_val:
                continue  # empty cell: don't clear an existing value
            row_present = True
            old_val = getattr(ras[0], attr) or ''
            field_changed = False
            for ra in ras:
                if (getattr(ra, attr) or '') != new_val:
                    setattr(ra, attr, new_val)
                    session.add(ra)
                    field_changed = True
                    if ra.inventory and ra.inventory.hotel_id:
                        hotels_imported.add(str(ra.inventory.hotel_id))
            if field_changed:
                row_changed = True
                changes.append({'confirmation_num': conf_num, 'field': col,
                                'old': old_val, 'new': new_val})
        if row_present:
            updated += 1 if row_changed else 0
            unchanged += 0 if row_changed else 1

    session.flush()  # emit the row updates before closing the window
    record.applied_to = datetime.now(UTC)
    record.updated_count = updated
    record.unchanged_count = unchanged
    record.note = f"{updated} updated, {unchanged} unchanged"

    for hotel_id in hotels_imported:
        session.add(HotelExportLog(
            hotel_id=hotel_id, export_type='confirmation_import', record_count=updated))
    session.flush()

    return {'updated': updated, 'unchanged': unchanged, 'changes': changes,
            'error': None, 'record': record}


def match_assignments(session, app_id='', conf=''):
    """Every RoomAssignment identified by one import row.

    Precedence: `app_id` (lottery_application_id), then `conf` as
    LotteryApplication.confirmation_num, then `conf` as
    RoomAssignment.hotel_confirmation_number (expanded to every assignment
    on the same application, so multi-room bookings stay in sync).
    Returns [] when nothing matches.
    """
    from uber.models import LotteryApplication
    from uber.models.hotel import RoomAssignment

    app_id = (app_id or '').strip()
    conf = (conf or '').strip()

    if app_id:
        ras = session.query(RoomAssignment).filter_by(
            lottery_application_id=app_id).all()
        if ras:
            return ras
    if conf:
        app = session.query(LotteryApplication).filter(
            LotteryApplication.confirmation_num == conf).one_or_none()
        if app:
            ras = session.query(RoomAssignment).filter_by(
                lottery_application_id=app.id).all()
            if ras:
                return ras
        ra = session.query(RoomAssignment).filter_by(
            hotel_confirmation_number=conf).first()
        if ra:
            if ra.lottery_application_id:
                return session.query(RoomAssignment).filter_by(
                    lottery_application_id=ra.lottery_application_id).all()
            return [ra]
    return []


def apply_confirmation_rows(session, rows, apply_changes=True, on_update=None):
    """Shared row applier for hotel confirmation numbers.

    Row columns: `lottery_application_id` and/or `confirmation_num`
    identify the booking (see match_assignments); `new_confirmation_num`
    (preferred) or `hotel_confirmation_number` is the value to write.
    Rows without a value are skipped. The value is written to EVERY
    assignment on the matched booking, one preview entry per assignment.

    `on_update(assignment)` is invoked for each assignment actually
    changed while applying, so each controller keeps its own email
    behavior (the admin import pages queue a confirmation-updated email;
    the portal upload does not).

    Flushes, never commits - the caller owns the transaction.
    Returns {'new', 'changed', 'unchanged', 'unmatched', 'applied'}.
    """
    preview = {'new': [], 'changed': [], 'unchanged': [], 'unmatched': [],
               'applied': 0}

    for row in rows:
        app_id = (row.get('lottery_application_id') or '').strip()
        conf = (row.get('confirmation_num') or '').strip()
        new_conf = ((row.get('new_confirmation_num')
                     or row.get('hotel_confirmation_number') or '')).strip()
        if not new_conf:
            continue

        assignments = match_assignments(session, app_id, conf)
        if not assignments:
            preview['unmatched'].append({
                'lottery_application_id': app_id,
                'confirmation_num': conf,
                'new_confirmation_num': new_conf,
            })
            continue

        for assignment in assignments:
            existing = assignment.hotel_confirmation_number or ''
            if existing == new_conf:
                preview['unchanged'].append({'assignment': assignment})
                continue

            bucket = 'changed' if existing else 'new'
            preview[bucket].append({
                'assignment': assignment,
                'old': existing,
                'new': new_conf,
            })

            if apply_changes:
                assignment.hotel_confirmation_number = new_conf
                session.add(assignment)
                preview['applied'] += 1
                if on_update:
                    on_update(assignment)

    session.flush()
    return preview


def apply_cancellation_rows(session, rows, apply_changes=True):
    """Shared row applier for hotel cancellation numbers.

    A row's presence means "this booking was cancelled": callers whose
    file layout lists every booking (like the bookings-export re-import)
    must pre-filter to rows that actually carry a cancellation number.

    Row columns: `lottery_application_id` and/or `confirmation_num`
    identify the booking (see match_assignments);
    `cancellation_confirmation_number` (or `hotel_cancellation_number`)
    is the hotel's cancel record id, optional - it defaults to 'imported'
    when applying. Setting it triggers the model presave that flips
    status to CANCELLED. Applies to EVERY assignment on the matched
    booking, one preview entry per assignment.

    Flushes, never commits - the caller owns the transaction.
    Returns {'matched', 'already', 'unmatched', 'applied'}.
    """
    preview = {'matched': [], 'already': [], 'unmatched': [], 'applied': 0}

    for row in rows:
        app_id = (row.get('lottery_application_id') or '').strip()
        conf = (row.get('confirmation_num') or '').strip()
        cancel_num = ((row.get('cancellation_confirmation_number')
                       or row.get('hotel_cancellation_number') or '')).strip()
        if not app_id and not conf:
            continue

        assignments = match_assignments(session, app_id, conf)
        if not assignments:
            preview['unmatched'].append({
                'confirmation_num': conf,
                'cancellation_confirmation_number': cancel_num,
            })
            continue

        for assignment in assignments:
            if assignment.status == c.CANCELLED:
                preview['already'].append({
                    'assignment': assignment,
                    'cancellation_confirmation_number': cancel_num,
                })
                if apply_changes and cancel_num \
                        and not assignment.cancellation_confirmation_number:
                    assignment.cancellation_confirmation_number = cancel_num
                    session.add(assignment)
                    preview['applied'] += 1
                continue

            preview['matched'].append({
                'assignment': assignment,
                'cancellation_confirmation_number': cancel_num,
            })

            if apply_changes:
                assignment.cancellation_confirmation_number = cancel_num or 'imported'
                # The model's cancellation_flips_status presave takes care
                # of status = CANCELLED.
                session.add(assignment)
                preview['applied'] += 1

    session.flush()
    return preview
